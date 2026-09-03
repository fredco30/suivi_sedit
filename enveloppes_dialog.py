"""
Dialogue de saisie en masse des enveloppes contractuelles des marchés
"""

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QBrush, QColor, QFont
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QTableWidget, QTableWidgetItem, QDialogButtonBox, QMessageBox,
    QHeaderView, QFileDialog, QAbstractItemView
)

from enveloppes_marches import collecter_marches, parser_montant

COL_MARCHE, COL_FOURNISSEUR, COL_OPERATIONS, COL_ENVELOPPE, COL_ENGAGE, COL_FACTURE = range(6)

EN_TETES = [
    "N° MARCHÉ",
    "FOURNISSEUR",
    "OPÉRATION(S)",
    "ENVELOPPE CONTRACTUELLE TTC",
    "Engagé à ce jour\n(repère)",
    "Facturé à ce jour\n(repère)",
]

COULEUR_A_SAISIR = QColor("#FFF2CC")
COULEUR_REPERE = QColor("#F5F5F5")


def _format_euro(montant) -> str:
    if not montant:
        return ""
    return f"{montant:,.2f} €".replace(",", " ")


class EnveloppesMarchesDialog(QDialog):
    """Saisie des enveloppes notifiées de tous les marchés, en une passe.

    L'enveloppe d'un marché est son montant notifié : elle figure dans l'acte
    d'engagement et dans aucun export SEDIT. Les colonnes « engagé » et
    « facturé » ne sont donc que des repères de saisie — surtout pas une
    valeur à recopier dans la colonne enveloppe.
    """

    def __init__(self, db, analyzer, parent=None):
        super().__init__(parent)
        self.db = db
        self.analyzer = analyzer
        self.fiches = []

        self.setWindowTitle("Enveloppes contractuelles des marchés")
        self.resize(1100, 700)

        self._init_ui()
        self._charger_marches()

    def _init_ui(self):
        layout = QVBoxLayout(self)

        explication = QLabel(
            "Saisir le montant TTC <b>notifié</b> de chaque marché (acte d'engagement, "
            "avenants compris). Sans lui, la colonne « solde » des suivis financiers se "
            "calcule contre un total reconstitué depuis SEDIT, qui n'est pas l'enveloppe : "
            "les deux peuvent différer d'un facteur 100."
        )
        explication.setWordWrap(True)
        explication.setStyleSheet(
            "background-color: #FFF3CD; border: 1px solid #FFE69C; "
            "border-radius: 4px; padding: 8px;"
        )
        layout.addWidget(explication)

        barre = QHBoxLayout()
        self.edit_filtre = QLineEdit()
        self.edit_filtre.setPlaceholderText("Filtrer par marché, fournisseur ou opération...")
        self.edit_filtre.setClearButtonEnabled(True)
        self.edit_filtre.textChanged.connect(self._appliquer_filtre)
        barre.addWidget(self.edit_filtre)

        btn_exporter = QPushButton("📤 Exporter vers Excel")
        btn_exporter.clicked.connect(self.exporter_excel)
        barre.addWidget(btn_exporter)

        btn_importer = QPushButton("📥 Importer depuis Excel")
        btn_importer.clicked.connect(self.importer_excel)
        barre.addWidget(btn_importer)

        layout.addLayout(barre)

        self.table = QTableWidget(0, len(EN_TETES))
        self.table.setHorizontalHeaderLabels(EN_TETES)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        entete = self.table.horizontalHeader()
        entete.setSectionResizeMode(COL_OPERATIONS, QHeaderView.Stretch)
        for colonne in (COL_MARCHE, COL_FOURNISSEUR, COL_ENVELOPPE, COL_ENGAGE, COL_FACTURE):
            entete.setSectionResizeMode(colonne, QHeaderView.ResizeToContents)
        self.table.itemChanged.connect(self._on_item_change)
        layout.addWidget(self.table)

        self.label_etat = QLabel()
        layout.addWidget(self.label_etat)

        boutons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        boutons.button(QDialogButtonBox.Save).setText("Enregistrer en base")
        boutons.accepted.connect(self.save)
        boutons.rejected.connect(self.reject)
        layout.addWidget(boutons)

    def _charger_marches(self):
        """Dresse l'inventaire des marchés et de ce qu'on sait déjà d'eux."""
        self.fiches = collecter_marches(self.analyzer)

        self.table.blockSignals(True)
        self.table.setRowCount(len(self.fiches))

        for ligne, fiche in enumerate(self.fiches):
            enregistrement = self.db.get_marche(fiche["marche"])
            montant = enregistrement["montant_initial_manuel"] if enregistrement else None
            fiche["montant_initial"] = montant

            valeurs = {
                COL_MARCHE: fiche["marche"],
                COL_FOURNISSEUR: fiche["fournisseur"],
                COL_OPERATIONS: ", ".join(sorted(fiche["operations"])),
                COL_ENVELOPPE: _format_euro(montant),
                COL_ENGAGE: _format_euro(fiche["engage"]),
                COL_FACTURE: _format_euro(fiche["facture"]),
            }

            for colonne, valeur in valeurs.items():
                item = QTableWidgetItem(valeur)
                if colonne == COL_ENVELOPPE:
                    item.setBackground(QBrush(COULEUR_A_SAISIR))
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    item.setFont(QFont("", -1, QFont.Bold))
                else:
                    # Seule la colonne enveloppe se saisit.
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    if colonne in (COL_ENGAGE, COL_FACTURE):
                        item.setBackground(QBrush(COULEUR_REPERE))
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(ligne, colonne, item)

        self.table.blockSignals(False)
        self._rafraichir_etat()

    def _ligne_du_marche(self, code_marche):
        for ligne in range(self.table.rowCount()):
            if self.table.item(ligne, COL_MARCHE).text() == code_marche:
                return ligne
        return None

    def _on_item_change(self, item):
        """Reformate la saisie et signale un montant illisible."""
        if item.column() != COL_ENVELOPPE:
            return

        try:
            montant = parser_montant(item.text())
        except ValueError:
            QMessageBox.warning(
                self, "Montant illisible",
                f"« {item.text()} » n'est pas un montant.\n\n"
                "Formats acceptés : 138108, 138108.00, 138 108,00 €"
            )
            self.table.blockSignals(True)
            item.setText(_format_euro(self.fiches[item.row()].get("montant_initial")))
            self.table.blockSignals(False)
            return

        self.fiches[item.row()]["montant_initial"] = montant
        self.table.blockSignals(True)
        item.setText(_format_euro(montant))
        self.table.blockSignals(False)
        self._rafraichir_etat()

    def _appliquer_filtre(self, texte):
        texte = (texte or "").strip().lower()
        for ligne in range(self.table.rowCount()):
            contenu = " ".join(
                self.table.item(ligne, colonne).text().lower()
                for colonne in (COL_MARCHE, COL_FOURNISSEUR, COL_OPERATIONS)
            )
            self.table.setRowHidden(ligne, bool(texte) and texte not in contenu)

    def _rafraichir_etat(self):
        renseignes = sum(1 for f in self.fiches if f.get("montant_initial"))
        total = len(self.fiches)
        manquants = total - renseignes
        self.label_etat.setText(
            f"<b>{renseignes}</b> / {total} enveloppes renseignées"
            + (f" — <span style='color:#b8860b;'>{manquants} à saisir</span>" if manquants else "")
        )

    def exporter_excel(self):
        """Sort le tableau pour une saisie hors application."""
        from enveloppes_marches import exporter

        chemin, _ = QFileDialog.getSaveFileName(
            self, "Exporter les enveloppes", "enveloppes_a_saisir.xlsx", "Excel Files (*.xlsx)"
        )
        if not chemin:
            return

        try:
            # L'analyzer est déjà chargé : on réutilise ses données plutôt que
            # de relire les sources.
            self._ecrire_classeur(chemin)
        except Exception as e:
            QMessageBox.critical(self, "Export impossible", str(e))
            return

        QMessageBox.information(
            self, "Export réussi",
            f"{len(self.fiches)} marchés exportés vers :\n{chemin}\n\n"
            "Remplir la colonne « ENVELOPPE CONTRACTUELLE TTC » puis réimporter."
        )

    def _ecrire_classeur(self, chemin):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from enveloppes_marches import EN_TETES as EN_TETES_FICHIER

        wb = Workbook()
        ws = wb.active
        ws.title = "Enveloppes"
        ws.cell(1, 1, "ENVELOPPES CONTRACTUELLES DES MARCHÉS — À RENSEIGNER").font = Font(
            bold=True, size=14
        )
        ws.cell(2, 1,
                "Saisir dans la colonne B le montant TTC notifié de chaque marché "
                "(acte d'engagement, avenants compris). La colonne F ne donne qu'un "
                "repère : c'est le montant engagé à ce jour, pas l'enveloppe.")
        ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=8)

        fill_entete = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        fill_saisie = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for colonne, entete in enumerate(EN_TETES_FICHIER, 1):
            cellule = ws.cell(5, colonne, entete)
            cellule.font = Font(bold=True, size=11)
            cellule.fill = fill_entete
            cellule.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for index, fiche in enumerate(self.fiches):
            ligne = 6 + index
            montant = fiche.get("montant_initial")
            ws.cell(ligne, 1, fiche["marche"])
            ws.cell(ligne, 2, montant or None).fill = fill_saisie
            ws.cell(ligne, 3, ", ".join(sorted(fiche["operations"])))
            ws.cell(ligne, 4, fiche["fournisseur"])
            ws.cell(ligne, 5, montant or None)
            ws.cell(ligne, 6, round(fiche["engage"], 2))
            ws.cell(ligne, 7, round(fiche["facture"], 2))
            ws.cell(ligne, 8, len(fiche["bdc"]))
            for colonne in (2, 5, 6, 7):
                ws.cell(ligne, colonne).number_format = '#,##0.00 €'

        for colonne, largeur in enumerate([18, 26, 28, 30, 24, 30, 20, 10], 1):
            ws.column_dimensions[get_column_letter(colonne)].width = largeur
        ws.freeze_panes = "A6"
        wb.save(chemin)

    def importer_excel(self):
        """Reprend les enveloppes saisies dans un classeur, sans écrire en base.

        Les valeurs alimentent le tableau : c'est « Enregistrer » qui décide.
        """
        from openpyxl import load_workbook
        from enveloppes_marches import COLONNE_ENVELOPPE, COLONNE_MARCHE

        chemin, _ = QFileDialog.getOpenFileName(
            self, "Importer les enveloppes", "", "Excel Files (*.xlsx *.xlsm)"
        )
        if not chemin:
            return

        try:
            wb = load_workbook(chemin, data_only=True)
            ws = wb["Enveloppes"] if "Enveloppes" in wb.sheetnames else wb.active

            ligne_entete, colonnes = None, {}
            for ligne in range(1, min(ws.max_row, 20) + 1):
                valeurs = {
                    str(ws.cell(ligne, colonne).value).strip(): colonne
                    for colonne in range(1, ws.max_column + 1)
                    if ws.cell(ligne, colonne).value is not None
                }
                if COLONNE_MARCHE in valeurs and COLONNE_ENVELOPPE in valeurs:
                    ligne_entete, colonnes = ligne, valeurs
                    break

            if ligne_entete is None:
                QMessageBox.warning(
                    self, "Fichier non reconnu",
                    f"Colonnes « {COLONNE_MARCHE} » et « {COLONNE_ENVELOPPE} » introuvables."
                )
                return

            reprises, inconnus, illisibles = 0, [], []
            for ligne in range(ligne_entete + 1, ws.max_row + 1):
                code = ws.cell(ligne, colonnes[COLONNE_MARCHE]).value
                if not code:
                    continue
                code = str(code).strip()

                try:
                    montant = parser_montant(ws.cell(ligne, colonnes[COLONNE_ENVELOPPE]).value)
                except ValueError:
                    illisibles.append(code)
                    continue
                if montant is None or montant <= 0:
                    continue

                index = self._ligne_du_marche(code)
                if index is None:
                    inconnus.append(code)
                    continue

                self.fiches[index]["montant_initial"] = montant
                self.table.blockSignals(True)
                self.table.item(index, COL_ENVELOPPE).setText(_format_euro(montant))
                self.table.blockSignals(False)
                reprises += 1

        except Exception as e:
            QMessageBox.critical(self, "Import impossible", str(e))
            return

        self._rafraichir_etat()

        message = f"{reprises} enveloppe(s) reprise(s) dans le tableau."
        if illisibles:
            message += f"\n\n{len(illisibles)} montant(s) illisible(s) : " \
                       f"{', '.join(illisibles[:5])}"
        if inconnus:
            message += f"\n\n{len(inconnus)} marché(s) inconnu(s) ignoré(s) : " \
                       f"{', '.join(inconnus[:5])}"
        message += "\n\nRien n'est encore écrit en base : cliquer « Enregistrer »."
        QMessageBox.information(self, "Import terminé", message)

    def save(self):
        """Écrit en base les enveloppes modifiées, en préservant le reste."""
        a_ecrire = []
        for fiche in self.fiches:
            montant = fiche.get("montant_initial")
            if not montant:
                continue
            enregistrement = self.db.get_marche(fiche["marche"])
            ancien = enregistrement["montant_initial_manuel"] if enregistrement else None
            if ancien is None or abs((ancien or 0) - montant) > 0.005:
                a_ecrire.append((fiche, enregistrement, montant))

        if not a_ecrire:
            QMessageBox.information(self, "Rien à enregistrer",
                                    "Aucune enveloppe n'a été modifiée.")
            return

        creations = sum(1 for _, enregistrement, _ in a_ecrire if enregistrement is None)
        confirmation = QMessageBox.question(
            self, "Confirmer l'enregistrement",
            f"{len(a_ecrire)} marché(s) à enregistrer "
            f"({creations} création(s), {len(a_ecrire) - creations} modification(s)).\n\n"
            "Continuer ?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes
        )
        if confirmation != QMessageBox.Yes:
            return

        for fiche, enregistrement, montant in a_ecrire:
            # upsert_marche réécrit toutes les colonnes : on repasse les valeurs
            # existantes pour ne pas effacer libellé, dates, type ou notes.
            donnees = {
                "libelle": "",
                "fournisseur": fiche["fournisseur"],
                "type_marche": "CLASSIQUE",
                "montant_initial_manuel": montant,
                "date_notification": None,
                "date_debut": None,
                "date_fin_prevue": None,
                "notes": "",
            }
            if enregistrement is not None:
                for cle in donnees:
                    if cle == "montant_initial_manuel":
                        continue
                    try:
                        valeur = enregistrement[cle]
                    except (KeyError, IndexError):
                        continue
                    if valeur is not None:
                        donnees[cle] = valeur

            self.db.upsert_marche(fiche["marche"], donnees)

        QMessageBox.information(
            self, "Enregistrement réussi",
            f"{len(a_ecrire)} marché(s) enregistré(s).\n\n"
            "Régénérer les suivis financiers pour que les soldes se calculent "
            "contre ces enveloppes."
        )
        self.accept()
