"""
Analyse du fichier _Suivi_financier_op.xlsx pour comprendre sa structure
"""

import pandas as pd
import openpyxl

# Lire le fichier
file_path = '_Suivi_financier_op.xlsx'

print("=" * 80)
print("ANALYSE DU FICHIER SUIVI FINANCIER D'OPÉRATION")
print("=" * 80)

# Ouvrir avec openpyxl pour voir les feuilles
wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"\n📊 Nombre de feuilles: {len(wb.sheetnames)}")
print(f"📋 Noms des feuilles: {wb.sheetnames}")

# Analyser chaque feuille
for sheet_name in wb.sheetnames:
    print(f"\n" + "=" * 80)
    print(f"FEUILLE: {sheet_name}")
    print("=" * 80)

    # Lire avec pandas
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    print(f"\n📏 Dimensions: {df.shape[0]} lignes × {df.shape[1]} colonnes")

    # Afficher les premières lignes
    print(f"\n📄 Aperçu des données (premières 20 lignes):")
    print("-" * 80)

    # Afficher avec les index de colonnes
    for i, row in df.head(20).iterrows():
        # Afficher ligne par ligne avec les valeurs non-NaN
        values = []
        for col_idx, val in enumerate(row):
            if pd.notna(val) and str(val).strip() != '':
                values.append(f"[{col_idx}]:{val}")
        if values:
            print(f"Ligne {i}: {' | '.join(values)}")

    # Identifier les lignes qui ressemblent à des en-têtes
    print(f"\n🔍 Recherche d'en-têtes potentiels:")
    for i in range(min(15, len(df))):
        row = df.iloc[i]
        non_empty = row.dropna()
        if len(non_empty) > 3:  # Si plus de 3 colonnes non vides
            print(f"  Ligne {i}: {non_empty.to_dict()}")

    # Analyser la structure des données
    print(f"\n📊 Statistiques par colonne:")
    for col in range(min(20, df.shape[1])):
        non_empty = df[col].dropna()
        if len(non_empty) > 0:
            print(f"  Col {col}: {len(non_empty)} valeurs non vides | "
                  f"Exemple: {str(non_empty.iloc[0])[:50]}")

print("\n" + "=" * 80)
print("ANALYSE TERMINÉE")
print("=" * 80)
