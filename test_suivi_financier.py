"""Tests de non-regression du suivi financier par marche.

Couvre le correctif du double comptage de l'enveloppe : le generateur emettait
une ligne d'engagement *et* une ligne de facture pour un meme BDC, sans jamais
solder l'engagement par la facture.

Deux niveaux :
  * `TestReglesAgregation` / `TestScenariosDocumentes` : regles metier sur des
    jeux synthetiques reproduisant les cas releves dans le fichier defectueux ;
  * `TestExportOperation` : export Excel complet sur les donnees du depot, qui
    verifie les invariants structurels (#7 a #11 du correctif) quel que soit le
    jeu de donnees ;
  * `TestValeursReference2020_14G3P` : les valeurs chiffrees attendues sur le
    jeu de reference `2020_14G3P`. Ignore si ce jeu n'est pas celui charge --
    l'instantane versionne dans le depot n'est pas celui du fichier diffuse.

Lancement :
    python -m unittest test_suivi_financier.py -v
ou
    python test_suivi_financier.py
"""
from __future__ import annotations

import glob
import os
import shutil
import sqlite3
import tempfile
import unittest
from collections import defaultdict
from typing import Dict, List

from suivi_financier_agg import (
    ETAT_NON_FACTURE,
    ETAT_PARTIEL,
    ETAT_SOLDE,
    LONGUEUR_DESIGNATION,
    STATUT_ENGAGEMENT,
    STATUT_FACTURE,
    TOLERANCE,
    Ecriture,
    ResultatSuivi,
    agreger_ecritures,
    nettoyer_designation,
)

MARCHE = "2020_14G3P"
# Index de colonnes SEDIT touches par les tests de resolution du n° de BDC.
MARCHE_COLONNES = (4, 38)
FOURNISSEUR = "BOUYGUES ENERGIES ET SERVICES"


def engagement(bdc: str, montant: float, libelle: str = "travaux", initial: float = None) -> Ecriture:
    """Ecriture d'engagement : BDC emis, rien de facture."""
    return Ecriture(
        marche=MARCHE, fournisseur=FOURNISSEUR, num_commande=bdc, libelle=libelle,
        montant_ttc=montant, montant_initial=montant if initial is None else initial,
    )


def facture(bdc: str, montant: float, num_facture: str, num_mandat: str,
            libelle: str = "travaux", initial: float = None) -> Ecriture:
    """Ecriture de realisation : la facture arrive et est mandatee."""
    return Ecriture(
        marche=MARCHE, fournisseur=FOURNISSEUR, num_commande=bdc, libelle=libelle,
        num_facture=num_facture, num_mandat=num_mandat, date_sf="2024-04-26",
        montant_ttc=montant, montant_sf=montant,
        montant_initial=montant if initial is None else initial,
    )


class TestNettoyageDesignation(unittest.TestCase):
    """Le suffixe (REPORT) sortait de la designation, tronquee avant concatenation."""

    def test_suffixe_report_supprime(self):
        self.assertEqual(
            nettoyer_designation("G3P extension du reseau video su(REPORT)"),
            "G3P extension du reseau video su",
        )

    def test_troncature_apres_nettoyage(self):
        libelle = "A" * 100
        self.assertEqual(len(nettoyer_designation(libelle)), LONGUEUR_DESIGNATION)
        self.assertEqual(LONGUEUR_DESIGNATION, 60)

    def test_libelle_vide_ou_absent(self):
        self.assertEqual(nettoyer_designation(None), "")
        self.assertEqual(nettoyer_designation(""), "")

    def test_marqueur_report_tronque_par_sedit(self):
        # SEDIT coupe le libelle a largeur fixe : le marqueur arrive ampute.
        for fragment in ("(REPORT", "(REPOR", "(REPO", "(REP"):
            self.assertEqual(
                nettoyer_designation("georeferencement obligatoire" + fragment),
                "georeferencement obligatoire",
                fragment,
            )

    def test_parenthese_legitime_conservee(self):
        # Une parenthese ouverte par une troncature ordinaire n'est pas un
        # marqueur de report : elle reste, faute de quoi on amputerait le texte.
        libelle = "G3P - CABLE HS Oree du Golf (allee de l'"
        self.assertEqual(nettoyer_designation(libelle), libelle)


class TestReglesAgregation(unittest.TestCase):
    """Regle : une ligne par etat de BDC, pas une ligne par ecriture SEDIT."""

    def test_engagement_solde_par_la_facture(self):
        # §1.1 : le BDC 22AA01789 de 2 779,20 € consommait 5 558,40 €.
        resultat = agreger_ecritures([
            engagement("22AA01789", 2779.20, "mise en place coquilles betons"),
            facture("22AA01789", 2779.20, "F2401609", "1423", "mise en place coquilles betons"),
        ], {"22AA01789": 2779.20})

        self.assertEqual(len(resultat.lignes), 1)
        self.assertEqual(resultat.lignes[0].statut, STATUT_FACTURE)
        self.assertAlmostEqual(resultat.total_impute, 2779.20, places=2)

    def test_report_exercice_ne_cree_pas_de_seconde_ligne(self):
        # §1.2 : deux engagements identiques pour un seul BDC.
        resultat = agreger_ecritures([
            engagement("25AA04392", 92351.76, "SAINT LOUIS- G3P-ECLAIRAGE ET FEUX DE CI"),
            engagement("25AA04392", 92351.76, "SAINT LOUIS- G3P-ECLAIRAGE ET FE(REPORT)"),
        ], {"25AA04392": 92351.76})

        self.assertEqual(len(resultat.lignes), 1)
        self.assertEqual(resultat.lignes[0].statut, STATUT_ENGAGEMENT)
        self.assertAlmostEqual(resultat.total_impute, 92351.76, places=2)
        self.assertNotIn("REPORT", resultat.lignes[0].designation)

    def test_engagement_ramene_au_reliquat(self):
        # BDC de 14 619,96 € facture pour 8 421,96 € : reliquat 6 198,00 €.
        resultat = agreger_ecritures([
            engagement("22AA02376", 8421.96, "georeferencement obligatoire", initial=14619.96),
            facture("22AA02376", 8421.96, "F2400578", "454", "georeferencement obligatoire",
                    initial=14619.96),
        ])

        statuts = [ligne.statut for ligne in resultat.lignes]
        self.assertEqual(statuts, [STATUT_FACTURE, STATUT_ENGAGEMENT])
        self.assertAlmostEqual(resultat.lignes[1].montant_impute, 6198.00, places=2)
        self.assertAlmostEqual(resultat.total_impute, 14619.96, places=2)

    def test_aucune_ligne_engagement_pour_bdc_solde(self):
        # §2.1 : un BDC soldé n'a pas de ligne d'engagement (assertion #8).
        resultat = agreger_ecritures([
            engagement("22AA03536", 83323.20, "REHABILITATION ALLEE DES GOELANDS"),
            facture("22AA03536", 886.80, "F2304108", "3683"),
            facture("22AA03536", 34638.00, "F2302126", "1822"),
            facture("22AA03536", 47798.40, "F2302126", "1822"),
        ], {"22AA03536": 83323.20})

        self.assertEqual(len(resultat.lignes), 3)
        self.assertTrue(all(l.statut == STATUT_FACTURE for l in resultat.lignes))
        self.assertEqual(resultat.bdcs[0].etat, ETAT_SOLDE)

    def test_un_seul_engagement_par_bdc(self):
        # Assertion #7 : quel que soit le nombre d'ecritures d'engagement.
        resultat = agreger_ecritures([
            engagement("23AA03103", 7315.20, "G3P Modification quai Bus", initial=8455.20),
            engagement("23AA03103", 8455.20, "G3P Modification quai(REPORT)", initial=8455.20),
            engagement("23AA03103", 8455.20, "G3P Modification quai Bus", initial=8455.20),
        ])

        engagements = [l for l in resultat.lignes if l.statut == STATUT_ENGAGEMENT]
        self.assertEqual(len(engagements), 1)
        self.assertAlmostEqual(engagements[0].montant_impute, 8455.20, places=2)

    def test_montant_ttc_bdc_identique_sur_toutes_les_lignes(self):
        # §1.3 / §2.2 : la colonne ne porte jamais un montant de facture.
        resultat = agreger_ecritures([
            engagement("23AA01991", 109875.60, "REQUALIFICATION ALLEE"),
            facture("23AA01991", 36089.64, "F2400012", "12"),
            facture("23AA01991", 73785.96, "F2302417", "2085"),
        ], {"23AA01991": 109875.60})

        montants_ref = {round(l.montant_ref, 2) for l in resultat.lignes}
        self.assertEqual(montants_ref, {109875.60})

    def test_bdc_jamais_facture_reste_engage_pour_son_montant(self):
        resultat = agreger_ecritures([
            engagement("26AA00890", 5000.00),
        ], {"26AA00890": 5000.00})

        self.assertEqual(len(resultat.lignes), 1)
        self.assertEqual(resultat.lignes[0].statut, STATUT_ENGAGEMENT)
        self.assertEqual(resultat.bdcs[0].etat, ETAT_NON_FACTURE)

    def test_facture_superieure_au_bdc_leve_une_anomalie(self):
        # §2.2 : le maximum est retenu, mais l'ecart remonte pour arbitrage.
        resultat = agreger_ecritures([
            facture("22AA03211", 17244.52, "F2400900", "900", initial=17052.52),
        ], {"22AA03211": 17052.52})

        self.assertEqual(len(resultat.anomalies), 1)
        self.assertIn("Facturé supérieur", resultat.anomalies[0].message)
        self.assertAlmostEqual(resultat.total_bdc_distincts, 17244.52, places=2)

    def test_montants_bdc_divergents_leve_une_anomalie(self):
        resultat = agreger_ecritures([
            engagement("25AA00614", 2100.87, initial=2100.87),
            engagement("25AA00614", 2899.13, initial=5000.00),
        ])

        self.assertEqual(len(resultat.anomalies), 1)
        self.assertIn("Plusieurs montants", resultat.anomalies[0].message)
        self.assertAlmostEqual(resultat.total_bdc_distincts, 5000.00, places=2)

    def test_etat_partiellement_facture(self):
        resultat = agreger_ecritures([
            facture("24AA00679", 1000.00, "F1", "1", initial=2500.00),
        ])
        self.assertEqual(resultat.bdcs[0].etat, ETAT_PARTIEL)
        self.assertAlmostEqual(resultat.bdcs[0].reliquat, 1500.00, places=2)

    def test_trace_des_lignes_neutralisees(self):
        # §4 : chaque ligne supprimee ou ajustee laisse une trace.
        resultat = agreger_ecritures([
            engagement("22AA01789", 2779.20, "coquilles betons"),
            facture("22AA01789", 2779.20, "F2401609", "1423", "coquilles betons"),
            engagement("22AA02376", 8421.96, "georeferencement", initial=14619.96),
            facture("22AA02376", 8421.96, "F2400578", "454", "georeferencement",
                    initial=14619.96),
        ])

        par_bdc = {n.cle_bdc: n for n in resultat.neutralisations}
        self.assertAlmostEqual(par_bdc["22AA01789"].montant_neutralise, 2779.20, places=2)
        self.assertAlmostEqual(par_bdc["22AA02376"].montant_retenu, 6198.00, places=2)
        self.assertAlmostEqual(par_bdc["22AA02376"].montant_neutralise, 2223.96, places=2)


class TestProprieteGenerique(unittest.TestCase):
    """Propriete a verifier sur tout marche, pas seulement sur 2020_14G3P."""

    JEUX = {
        "engagement puis facture": [
            engagement("A1", 1000.0),
            facture("A1", 1000.0, "F1", "M1"),
        ],
        "engagement reporte deux fois": [
            engagement("B1", 500.0),
            engagement("B1", 500.0, "libelle(REPORT)"),
        ],
        "bdc paye en trois factures": [
            engagement("C1", 900.0),
            facture("C1", 300.0, "F1", "M1"),
            facture("C1", 300.0, "F2", "M2"),
            facture("C1", 300.0, "F3", "M3"),
        ],
        "facture sans mandat (non mandatee)": [
            engagement("D1", 700.0),
            Ecriture(marche=MARCHE, num_commande="D1", num_facture="F9", num_mandat="",
                     montant_ttc=700.0, montant_initial=700.0),
        ],
        "facture superieure au bdc": [
            engagement("E1", 100.0, initial=100.0),
            facture("E1", 150.0, "F1", "M1", initial=100.0),
        ],
        "ligne sans bdc rattachee a la tranche": [
            Ecriture(marche=MARCHE, tranche_libelle="TF", montant_ttc=250.0,
                     montant_initial=250.0),
        ],
    }

    def test_somme_imputee_egale_max_bdc_factures(self):
        for nom, ecritures in self.JEUX.items():
            with self.subTest(jeu=nom):
                resultat = agreger_ecritures(ecritures)

                impute_par_bdc: Dict[str, float] = defaultdict(float)
                for ligne in resultat.lignes:
                    impute_par_bdc[ligne.cle_bdc] += ligne.montant_impute

                for bdc in resultat.bdcs:
                    attendu = max(bdc.montant_ref, bdc.montant_facture)
                    self.assertAlmostEqual(
                        impute_par_bdc[bdc.cle_bdc], attendu, places=2,
                        msg=f"BDC {bdc.cle_bdc} du jeu « {nom} »",
                    )

                self.assertAlmostEqual(
                    resultat.total_impute, resultat.total_bdc_distincts, places=2
                )

    def test_au_plus_un_engagement_par_bdc(self):
        for nom, ecritures in self.JEUX.items():
            with self.subTest(jeu=nom):
                resultat = agreger_ecritures(ecritures)
                compte: Dict[str, int] = defaultdict(int)
                for ligne in resultat.lignes:
                    if ligne.statut == STATUT_ENGAGEMENT:
                        compte[ligne.cle_bdc] += 1
                self.assertTrue(all(n <= 1 for n in compte.values()), compte)

    def test_aucun_engagement_sur_bdc_solde(self):
        for nom, ecritures in self.JEUX.items():
            with self.subTest(jeu=nom):
                resultat = agreger_ecritures(ecritures)
                soldes = {b.cle_bdc for b in resultat.bdcs if b.etat == ETAT_SOLDE}
                engages = {l.cle_bdc for l in resultat.lignes
                           if l.statut == STATUT_ENGAGEMENT}
                self.assertEqual(soldes & engages, set())


# ---------------------------------------------------------------------------
# Export Excel complet
# ---------------------------------------------------------------------------

_ANALYZER_CACHE: List = []


def _charger_analyzer():
    """Analyzer sur le jeu de reference, ou None si les sources sont absentes.

    Les exports SEDIT sont annuels et se recouvrent ; le suivi d'une operation
    pluriannuelle se lit sur leur reunion. On les charge tous, dans un cache
    temporaire, sans toucher au cache de travail de l'application.
    """
    if _ANALYZER_CACHE:
        return _ANALYZER_CACHE[0]

    sources = sorted(glob.glob(os.path.join("data_sources", "factures*.xls")))
    if not sources or not os.path.exists("suivi_commandes.db"):
        return None
    try:
        from marches_module import MarchesAnalyzer
        from regenerer_suivis import BaseSuiviLectureSeule
    except ImportError:
        return None

    analyzer = MarchesAnalyzer(
        sources,
        database=BaseSuiviLectureSeule("suivi_commandes.db"),
        use_cache=True,
        cache_path=os.path.join(tempfile.mkdtemp(), "cache_test.db"),
    )
    if not analyzer.load_data(force_reload=True):
        return None

    _ANALYZER_CACHE.append(analyzer)
    return analyzer


def _lignes_financier(ws) -> List[dict]:
    """Lit les lignes de donnees de l'onglet FINANCIER (hors sous-totaux)."""
    lignes = []
    for row in range(6, ws.max_row + 1):
        statut = ws.cell(row, 15).value
        if statut in (None, "TOTAL"):
            continue
        lignes.append({
            "row": row,
            "bdc": ws.cell(row, 5).value,
            "montant_ref": ws.cell(row, 6).value,
            "num_facture": ws.cell(row, 7).value,
            "montant_impute": ws.cell(row, 11).value,
            "restant": ws.cell(row, 14).value,
            "statut": statut,
        })
    return lignes


class TestExportOperation(unittest.TestCase):
    """Invariants structurels de l'export, verifies sur les donnees du depot."""

    OPERATION = MARCHE

    @classmethod
    def setUpClass(cls):
        analyzer = _charger_analyzer()
        if analyzer is None:
            raise unittest.SkipTest("données du dépôt indisponibles")

        cls._tmpdir = tempfile.TemporaryDirectory()
        chemin = os.path.join(cls._tmpdir.name, "suivi.xlsx")
        if not analyzer.export_suivi_financier_operation(
            cls.OPERATION, chemin, exercice_filter=None, special_export=False
        ):
            raise unittest.SkipTest(f"export de {cls.OPERATION} impossible")

        import openpyxl
        cls.wb = openpyxl.load_workbook(chemin)
        cls.ws = cls.wb["FINANCIER"]
        cls.lignes = _lignes_financier(cls.ws)
        cls.ligne_soustotal = cls.ws.max_row

    @classmethod
    def tearDownClass(cls):
        if hasattr(cls, "_tmpdir"):
            cls._tmpdir.cleanup()

    def test_onglets_produits(self):
        self.assertEqual(
            self.wb.sheetnames,
            ["FINANCIER", "A jour", "Anomalies", "Lignes neutralisées"],
        )

    def test_enveloppe_initiale_ecrite_en_clair(self):
        # §2.4 : l'enveloppe ne doit plus se deduire de N6 + M6.
        self.assertIn("Enveloppe initiale", self.ws.cell(4, 1).value)
        self.assertGreater(self.ws.cell(4, 6).value, 0)

    def test_7_un_seul_engagement_par_bdc(self):
        compte: Dict[str, int] = defaultdict(int)
        for ligne in self.lignes:
            if ligne["statut"] == STATUT_ENGAGEMENT:
                compte[ligne["bdc"]] += 1
        doublons = {bdc: n for bdc, n in compte.items() if n > 1}
        self.assertEqual(doublons, {})

    def test_8_aucun_bdc_solde_ne_porte_d_engagement(self):
        ws_ajour = self.wb["A jour"]
        soldes = set()
        for row in range(5, ws_ajour.max_row):
            if ws_ajour.cell(row, 9).value == ETAT_SOLDE:
                soldes.add(ws_ajour.cell(row, 5).value)
        engages = {l["bdc"] for l in self.lignes if l["statut"] == STATUT_ENGAGEMENT}
        self.assertEqual(soldes & engages, set())

    def test_9_totaux_facture_identiques_entre_onglets(self):
        # §2.5 : les deux onglets sortent du meme jeu de donnees.
        ws_ajour = self.wb["A jour"]
        total_ajour = ws_ajour.cell(ws_ajour.max_row, 7).value
        total_financier = sum(
            l["montant_impute"] for l in self.lignes if l["statut"] == STATUT_FACTURE
        )
        self.assertAlmostEqual(total_ajour, total_financier, places=2)

    def test_10_aucun_couple_facture_montant_en_double(self):
        vus = defaultdict(int)
        for ligne in self.lignes:
            if ligne["statut"] == STATUT_FACTURE:
                vus[(ligne["num_facture"], round(ligne["montant_impute"], 2))] += 1
        self.assertEqual({k: n for k, n in vus.items() if n > 1}, {})

    def test_11_dernier_solde_glissant_egale_le_sous_total(self):
        # §2.4 : un seul mode de calcul, glissant et sous-total confondus.
        dernier = self.lignes[-1]["restant"]
        sous_total = self.ws.cell(self.ligne_soustotal, 14).value
        self.assertAlmostEqual(dernier, sous_total, places=2)

    def test_colonne_montant_bdc_constante_par_bdc(self):
        # §2.2 : jamais un montant de facture dans cette colonne.
        valeurs = defaultdict(set)
        for ligne in self.lignes:
            valeurs[ligne["bdc"]].add(round(ligne["montant_ref"], 2))
        multiples = {bdc: v for bdc, v in valeurs.items() if len(v) > 1}
        self.assertEqual(multiples, {})

    def test_sous_total_montant_bdc_somme_sur_bdc_distincts(self):
        # §2.2 : jamais une somme de colonne.
        distincts = {}
        for ligne in self.lignes:
            distincts[ligne["bdc"]] = ligne["montant_ref"]
        attendu = sum(distincts.values())
        self.assertAlmostEqual(self.ws.cell(self.ligne_soustotal, 6).value, attendu, places=2)

    def test_invariant_somme_imputee_egale_somme_bdc(self):
        impute = sum(l["montant_impute"] for l in self.lignes)
        distincts = {l["bdc"]: l["montant_ref"] for l in self.lignes}
        self.assertAlmostEqual(impute, sum(distincts.values()), places=2)

    def test_aucun_suffixe_report_dans_les_designations(self):
        # §2.3 : le suffixe est remplace par la colonne STATUT, y compris
        # lorsque SEDIT ne l'a remonte qu'ampute (« ...obligatoire(REP »).
        for row in range(6, self.ws.max_row + 1):
            designation = self.ws.cell(row, 2).value or ""
            self.assertNotIn("REPORT)", designation)
            for fragment in ("(REPORT", "(REPOR", "(REPO", "(REP"):
                self.assertFalse(designation.endswith(fragment), designation)

    def test_numeros_de_mandat_sans_decimale(self):
        # SEDIT remonte les mandats comme des nombres : « 1470.0 » n'est pas
        # un n° de mandat, le mandat s'appelle « 1470 ».
        mandats = [
            str(self.ws.cell(row, 8).value)
            for row in range(6, self.ws.max_row)
            if self.ws.cell(row, 8).value
        ]
        self.assertTrue(mandats, "aucun mandat dans l'export de reference")
        for mandat in mandats:
            self.assertFalse(mandat.endswith(".0"), mandat)


class TestResolutionNumeroBdc(unittest.TestCase):
    """La colonne « Commande » est vide sur pres de 40 % des lignes SEDIT.

    Le n° d'engagement se lit alors dans « Code mouvement ». Sans ce repli, ces
    lignes se retrouvent sans BDC et echappent a toute agregation par BDC.
    """

    @classmethod
    def setUpClass(cls):
        try:
            import pandas as pd
            from marches_module import MarchesAnalyzer
        except ImportError:
            raise unittest.SkipTest("pandas indisponible")
        cls.pd = pd
        cls.analyzer = MarchesAnalyzer("inexistant.xls", use_cache=False)

    def _df(self, lignes):
        largeur = max(MARCHE_COLONNES) + 1
        return self.pd.DataFrame(lignes, columns=range(largeur))

    def test_repli_sur_le_code_mouvement(self):
        colonnes = [None] * (max(MARCHE_COLONNES) + 1)
        ligne_liee = list(colonnes)
        ligne_liee[4], ligne_liee[38] = "24AA02259", "24AA02259"
        ligne_orpheline = list(colonnes)
        ligne_orpheline[4], ligne_orpheline[38] = "25AA03133", None
        ligne_vide = list(colonnes)
        ligne_vide[4], ligne_vide[38] = "26AA00890", "   "

        resolus = self.analyzer._resoudre_num_bdc(
            self._df([ligne_liee, ligne_orpheline, ligne_vide])
        )
        self.assertEqual(
            list(resolus), ["24AA02259", "25AA03133", "26AA00890"]
        )

    def test_toutes_les_ecritures_du_jeu_de_reference_portent_un_bdc(self):
        analyzer = _charger_analyzer()
        if analyzer is None:
            self.skipTest("données du dépôt indisponibles")
        groupes, _, _, _, info = analyzer.collecter_ecritures_operation(MARCHE)
        if info is None:
            self.skipTest(f"opération {MARCHE} absente")
        sans_bdc = [
            e for ecritures in groupes.values() for e in ecritures if not e.num_commande
        ]
        self.assertEqual(sans_bdc, [])


class TestValeursReference2020_14G3P(unittest.TestCase):
    """Valeurs chiffrees attendues sur le jeu de reference `2020_14G3P`.

    Le jeu est la reunion des exports SEDIT annuels de `data_sources/`, soit
    172 ecritures pour ce marche -- le meme perimetre que le fichier diffuse.

    Ecart assume avec le fichier corrige fourni comme modele
    ------------------------------------------------------
    Ce fichier a ete construit par retraitement du tableau defectueux, sans la
    colonne « Montant initial » (O) de l'export SEDIT. Faute de cette source, il
    retient le plus gros montant *de ligne* de chaque BDC et signale sept BDC
    « Montant BDC à confirmer ». La colonne O porte le montant total du BDC :
    elle en tranche quatre, tous sous-evalues par le fichier modele.

      BDC          modele       colonne O   ecart      lecture
      21AA01640    2 316,00     3 777,60    1 461,60   BDC en 2 lignes (n° 3 et 4), aucune facturee
      22AA02376    8 421,96    14 619,96    6 198,00   ligne 2 facturee, ligne 1 jamais facturee
      25AA00614   57 899,13    60 000,00    2 100,87   BDC annuel, 13e mensualite non facturee
      25AA02351   32 839,20    46 048,20   13 209,00   ligne 1 facturee, ligne 2 non facturee
                                           ---------
                                           22 969,47

    La colonne « Reste engagé » (F) de SEDIT confirme la lecture par la colonne
    O : sur les 19 BDC ou elle est renseignee, le reliquat calcule ici la
    reproduit exactement (les trois apparents ecarts sont des lignes soldees
    dans un export plus recent que celui qui portait le reste engage).

    Totaux du fichier modele, pour memoire : 134 lignes, 1 803 139,47 € imputes,
    261 094,21 € engages, solde 2 372 586,53 €.
    """

    # Valeurs du fichier modele, conservees pour tracer l'ecart de 22 969,47 €.
    MODELE = {
        "nb_lignes": 134,
        "nb_lignes_engagement": 21,
        "total_impute": 1803139.47,
        "total_engagement": 261094.21,
        "solde_final": 2372586.53,
    }

    ATTENDU = {
        "nb_lignes": 137,
        "nb_lignes_facturees": 113,
        "nb_lignes_engagement": 24,
        "total_impute": 1826108.94,
        "total_facture": 1542045.26,
        "total_engagement": 284063.68,
        "total_bdc_distincts": 1826108.94,
        "solde_final": 2349617.06,
        "enveloppe_initiale": 4175726.00,
    }

    # Empreinte du jeu de reference : deux valeurs sur lesquelles le calcul et
    # le fichier modele s'accordent, donc neutres vis-a-vis de l'ecart assume.
    EMPREINTE = {"nb_bdc": 100, "total_facture": 1542045.26}

    @classmethod
    def setUpClass(cls):
        analyzer = _charger_analyzer()
        if analyzer is None:
            raise unittest.SkipTest("données du dépôt indisponibles")

        groupes, montants, enveloppes, _, info = analyzer.collecter_ecritures_operation(MARCHE)
        if info is None:
            raise unittest.SkipTest(f"opération {MARCHE} absente")

        cls.resultat = ResultatSuivi()
        for ecritures in groupes.values():
            cls.resultat.etendre(
                agreger_ecritures(ecritures, montants, trier_par_bdc=True)
            )
        cls.enveloppe = sum(enveloppes.values())

        empreinte = (len(cls.resultat.bdcs), round(cls.resultat.total_facture, 2))
        attendue = (cls.EMPREINTE["nb_bdc"], cls.EMPREINTE["total_facture"])
        if empreinte != attendue:
            raise unittest.SkipTest(
                f"jeu de données différent de la référence (empreinte {empreinte} "
                f"au lieu de {attendue}) — "
                "rejouer ces assertions sur l'export SEDIT de référence"
            )

    def test_1_nombre_de_lignes_emises(self):
        lignes = self.resultat.lignes
        self.assertEqual(len(lignes), self.ATTENDU["nb_lignes"])
        self.assertEqual(
            sum(1 for l in lignes if l.statut == STATUT_FACTURE),
            self.ATTENDU["nb_lignes_facturees"],
        )
        self.assertEqual(
            sum(1 for l in lignes if l.statut == STATUT_ENGAGEMENT),
            self.ATTENDU["nb_lignes_engagement"],
        )

    def test_2_somme_montants_imputes(self):
        self.assertAlmostEqual(
            self.resultat.total_impute, self.ATTENDU["total_impute"], places=2
        )

    def test_ecart_assume_avec_le_fichier_modele(self):
        """L'écart avec le fichier modèle vaut exactement les 4 BDC sous-évalués."""
        ecart = self.resultat.total_impute - self.MODELE["total_impute"]
        self.assertAlmostEqual(ecart, 22969.47, places=2)

        # Les quatre BDC concernés, montant de référence lu en colonne O.
        attendus = {
            "21AA01640": 3777.60,
            "22AA02376": 14619.96,
            "25AA00614": 60000.00,
            "25AA02351": 46048.20,
        }
        par_bdc = {b.cle_bdc: round(b.montant_ref, 2) for b in self.resultat.bdcs}
        for bdc, montant in attendus.items():
            self.assertAlmostEqual(par_bdc[bdc], montant, places=2, msg=f"BDC {bdc}")

    def test_3_somme_lignes_facturees(self):
        self.assertAlmostEqual(
            self.resultat.total_facture, self.ATTENDU["total_facture"], places=2
        )

    def test_4_somme_lignes_engagement(self):
        self.assertAlmostEqual(
            self.resultat.total_engagement, self.ATTENDU["total_engagement"], places=2
        )

    def test_5_somme_montants_bdc_distincts(self):
        self.assertAlmostEqual(
            self.resultat.total_bdc_distincts, self.ATTENDU["total_bdc_distincts"], places=2
        )

    def test_6_solde_final(self):
        self.assertAlmostEqual(self.enveloppe, self.ATTENDU["enveloppe_initiale"], places=2)
        self.assertAlmostEqual(
            self.enveloppe - self.resultat.total_impute,
            self.ATTENDU["solde_final"], places=2,
        )


class TestSynchronisationMultiFichiers(unittest.TestCase):
    """Un export SEDIT est annuel : le suivi se lit sur la reunion des exports.

    La synchronisation devait donc cesser de traiter un fichier comme la
    totalite du cache -- elle supprimait les lignes de tous les autres.
    """

    LARGEUR = 50

    @classmethod
    def setUpClass(cls):
        try:
            import pandas as pd
            from marches_module import MarchesAnalyzer
        except ImportError:
            raise unittest.SkipTest("pandas indisponible")
        cls.pd = pd
        cls.MarchesAnalyzer = MarchesAnalyzer

    def setUp(self):
        self.repertoire = tempfile.mkdtemp()
        self.cache = os.path.join(self.repertoire, "cache.db")

    def tearDown(self):
        shutil.rmtree(self.repertoire, ignore_errors=True)

    def _ecrire_export(self, nom, lignes):
        """Ecrit un export SEDIT minimal : (bdc, libelle, ttc, facture, mandat)."""
        colonnes = self.MarchesAnalyzer
        contenu = []
        for bdc, libelle, ttc, facture, mandat in lignes:
            ligne = [None] * self.LARGEUR
            ligne[colonnes.COL_CODE_MOUVEMENT] = bdc
            ligne[colonnes.COL_FOURNISSEUR] = FOURNISSEUR
            ligne[colonnes.COL_LIBELLE] = libelle
            ligne[colonnes.COL_MONTANT_INITIAL] = ttc
            ligne[colonnes.COL_MONTANT_TTC] = ttc
            ligne[colonnes.COL_MONTANT_SF] = ttc if facture else 0
            ligne[colonnes.COL_FACTURE] = facture
            ligne[colonnes.COL_MANDAT] = mandat
            ligne[colonnes.COL_MARCHE] = MARCHE
            contenu.append(ligne)

        chemin = os.path.join(self.repertoire, nom)
        # La 1re ligne du fichier sert d'en-tete a pandas, comme un export SEDIT.
        entete = [f"col{i}" for i in range(self.LARGEUR)]
        self.pd.DataFrame(contenu, columns=entete).to_excel(chemin, index=False)
        return chemin

    def _charger(self, sources, force=False):
        analyzer = self.MarchesAnalyzer(sources, use_cache=True, cache_path=self.cache)
        self.assertTrue(analyzer.load_data(force_reload=force))
        return analyzer

    def test_les_exports_annuels_se_cumulent(self):
        a = self._ecrire_export("2024.xlsx", [("24AA00001", "borne", 1000.0, None, None)])
        b = self._ecrire_export("2025.xlsx", [("25AA00002", "mât", 2000.0, "F1", "M1")])

        analyzer = self._charger([a, b])
        self.assertEqual(len(analyzer.df_marches), 2)
        self.assertEqual(sorted(analyzer.sync.list_sources()), sorted([a, b]))

    def test_une_ligne_presente_dans_deux_exports_n_est_stockee_qu_une_fois(self):
        # Cas reel : un engagement non solde est reconduit d'un exercice sur
        # l'autre et figure a l'identique dans les deux exports.
        commune = ("24AA00001", "borne", 1000.0, None, None)
        a = self._ecrire_export("2024.xlsx", [commune])
        b = self._ecrire_export("2025.xlsx", [commune, ("25AA00002", "mât", 2000.0, "F1", "M1")])

        analyzer = self._charger([a, b])
        self.assertEqual(len(analyzer.df_marches), 2)

    def test_synchroniser_un_export_ne_supprime_pas_les_lignes_des_autres(self):
        a = self._ecrire_export("2024.xlsx", [("24AA00001", "borne", 1000.0, None, None)])
        b = self._ecrire_export("2025.xlsx", [("25AA00002", "mât", 2000.0, "F1", "M1")])
        self._charger([a, b])

        # L'export 2025 est remplacé par une version modifiée : seules ses
        # propres lignes doivent bouger.
        os.remove(b)
        b = self._ecrire_export("2025.xlsx", [("25AA00003", "coffret", 3000.0, "F2", "M2")])

        analyzer = self._charger([a, b], force=True)
        bdcs = set(analyzer.df_marches.iloc[:, self.MarchesAnalyzer.COL_COMMANDE])
        self.assertEqual(bdcs, {"24AA00001", "25AA00003"})

    def test_retirer_un_export_de_la_selection_retire_ses_lignes(self):
        a = self._ecrire_export("2024.xlsx", [("24AA00001", "borne", 1000.0, None, None)])
        b = self._ecrire_export("2025.xlsx", [("25AA00002", "mât", 2000.0, "F1", "M1")])
        self._charger([a, b])

        analyzer = self._charger([a])
        self.assertEqual(len(analyzer.df_marches), 1)
        self.assertEqual(analyzer.sync.list_sources(), [a])

    def test_une_ligne_partagee_survit_au_retrait_d_un_seul_export(self):
        commune = ("24AA00001", "borne", 1000.0, None, None)
        a = self._ecrire_export("2024.xlsx", [commune])
        b = self._ecrire_export("2025.xlsx", [commune])
        self._charger([a, b])

        analyzer = self._charger([a])
        self.assertEqual(len(analyzer.df_marches), 1)

    def test_second_chargement_ne_resynchronise_rien(self):
        a = self._ecrire_export("2024.xlsx", [("24AA00001", "borne", 1000.0, None, None)])
        self._charger([a])

        analyzer = self._charger([a])
        self.assertEqual(analyzer.sync_stats["nb_inserted"], 0)
        self.assertEqual(analyzer.sync_stats["nb_deleted"], 0)

    def test_source_introuvable_signalee(self):
        analyzer = self.MarchesAnalyzer(
            os.path.join(self.repertoire, "aucun*.xlsx"), use_cache=True, cache_path=self.cache
        )
        self.assertFalse(analyzer.load_data())


class TestResolutionSources(unittest.TestCase):
    """Une source peut etre un fichier, un motif, un repertoire ou une liste."""

    @classmethod
    def setUpClass(cls):
        try:
            from marches_module import MarchesAnalyzer
        except ImportError:
            raise unittest.SkipTest("marches_module indisponible")
        cls.resoudre = MarchesAnalyzer.resoudre_sources

    def setUp(self):
        self.repertoire = tempfile.mkdtemp()
        for nom in ("factures_2024.xlsx", "factures_2025.xlsx", "notes.txt"):
            open(os.path.join(self.repertoire, nom), "w").close()

    def tearDown(self):
        shutil.rmtree(self.repertoire, ignore_errors=True)

    def test_repertoire_developpe_en_fichiers_excel(self):
        resolus = self.resoudre(self.repertoire)
        self.assertEqual([os.path.basename(c) for c in resolus],
                         ["factures_2024.xlsx", "factures_2025.xlsx"])

    def test_motif_developpe(self):
        resolus = self.resoudre(os.path.join(self.repertoire, "factures_*.xlsx"))
        self.assertEqual(len(resolus), 2)

    def test_liste_sans_doublon_et_ordre_stable(self):
        fichier = os.path.join(self.repertoire, "factures_2024.xlsx")
        resolus = self.resoudre([fichier, self.repertoire])
        self.assertEqual([os.path.basename(c) for c in resolus],
                         ["factures_2024.xlsx", "factures_2025.xlsx"])

    def test_mot_cle_database_sync_ne_designe_aucun_fichier(self):
        self.assertEqual(self.resoudre("database_sync"), [])


class TestEnveloppesMarches(unittest.TestCase):
    """Saisie en lot des enveloppes contractuelles.

    L'enveloppe d'un marche est son montant notifie : elle ne figure dans aucun
    export SEDIT et ne s'en deduit pas. Le module ne devine donc rien -- il
    produit un tableau a remplir et reinjecte la colonne saisie.
    """

    @classmethod
    def setUpClass(cls):
        try:
            import openpyxl
            import enveloppes_marches
        except ImportError:
            raise unittest.SkipTest("openpyxl indisponible")
        cls.openpyxl = openpyxl
        cls.module = enveloppes_marches

    def setUp(self):
        self.repertoire = tempfile.mkdtemp()
        self.db = os.path.join(self.repertoire, "suivi.db")
        conn = sqlite3.connect(self.db)
        conn.execute("""
            CREATE TABLE marches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code_marche TEXT UNIQUE NOT NULL,
                libelle TEXT, fournisseur TEXT, montant_initial_manuel REAL,
                date_notification TEXT, date_debut TEXT, date_fin_prevue TEXT,
                notes TEXT, last_update TEXT, type_marche TEXT DEFAULT 'CLASSIQUE'
            )
        """)
        conn.execute(
            "INSERT INTO marches (code_marche, montant_initial_manuel) VALUES ('DEJA_LA', 1000.0)"
        )
        conn.commit()
        conn.close()

    def tearDown(self):
        shutil.rmtree(self.repertoire, ignore_errors=True)

    def _tableau(self, lignes):
        """Écrit un tableau au format attendu par l'importateur."""
        chemin = os.path.join(self.repertoire, "enveloppes.xlsx")
        wb = self.openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enveloppes"
        for col, entete in enumerate(self.module.EN_TETES, 1):
            ws.cell(5, col, entete)
        for idx, (code, montant) in enumerate(lignes, start=6):
            ws.cell(idx, 1, code)
            ws.cell(idx, 2, montant)
        wb.save(chemin)
        return chemin

    def _montants(self):
        conn = sqlite3.connect(self.db)
        try:
            return dict(conn.execute(
                "SELECT code_marche, montant_initial_manuel FROM marches"
            ))
        finally:
            conn.close()

    def test_simulation_n_ecrit_rien(self):
        fichier = self._tableau([("NOUVEAU", 5000.0)])
        self.assertEqual(self.module.importer(fichier, self.db, appliquer=False), 0)
        self.assertNotIn("NOUVEAU", self._montants())

    def test_application_cree_et_modifie(self):
        fichier = self._tableau([("NOUVEAU", 5000.0), ("DEJA_LA", 2500.0)])
        self.assertEqual(self.module.importer(fichier, self.db, appliquer=True), 0)
        montants = self._montants()
        self.assertAlmostEqual(montants["NOUVEAU"], 5000.0, places=2)
        self.assertAlmostEqual(montants["DEJA_LA"], 2500.0, places=2)

    def test_lignes_sans_montant_ignorees(self):
        fichier = self._tableau([("NOUVEAU", None), ("VIDE", ""), ("ZERO", 0)])
        self.assertEqual(self.module.importer(fichier, self.db, appliquer=True), 0)
        montants = self._montants()
        self.assertEqual(set(montants), {"DEJA_LA"})

    def test_montant_saisi_a_la_francaise(self):
        fichier = self._tableau([("NOUVEAU", "12 345,67 €")])
        self.assertEqual(self.module.importer(fichier, self.db, appliquer=True), 0)
        self.assertAlmostEqual(self._montants()["NOUVEAU"], 12345.67, places=2)

    def test_montant_illisible_refuse_sans_rien_ecrire(self):
        fichier = self._tableau([("NOUVEAU", "à voir avec le service")])
        self.assertEqual(self.module.importer(fichier, self.db, appliquer=True), 1)
        self.assertNotIn("NOUVEAU", self._montants())

    def test_colonnes_deplacees_toujours_reperees(self):
        # Le fichier revient d'un tableur : ses colonnes ont pu bouger.
        chemin = os.path.join(self.repertoire, "deplace.xlsx")
        wb = self.openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enveloppes"
        ws.cell(3, 4, self.module.COLONNE_ENVELOPPE)
        ws.cell(3, 7, self.module.COLONNE_MARCHE)
        ws.cell(4, 7, "NOUVEAU")
        ws.cell(4, 4, 7500.0)
        wb.save(chemin)

        self.assertEqual(self.module.importer(chemin, self.db, appliquer=True), 0)
        self.assertAlmostEqual(self._montants()["NOUVEAU"], 7500.0, places=2)

    def test_en_tetes_absents_signales(self):
        chemin = os.path.join(self.repertoire, "vide.xlsx")
        wb = self.openpyxl.Workbook()
        wb.active.cell(1, 1, "rien à voir")
        wb.save(chemin)
        self.assertEqual(self.module.importer(chemin, self.db, appliquer=True), 1)

    def test_enveloppe_saisie_reprise_par_l_export(self):
        """Une fois saisie, l'enveloppe est reprise sans autre intervention."""
        analyzer = _charger_analyzer()
        if analyzer is None:
            self.skipTest("données du dépôt indisponibles")

        _, _, enveloppes, provenances, info = analyzer.collecter_ecritures_operation(MARCHE)
        if info is None:
            self.skipTest(f"opération {MARCHE} absente")

        from marches_module import MarchesAnalyzer
        self.assertEqual(set(provenances.values()), {MarchesAnalyzer.ENVELOPPE_BASE})
        self.assertAlmostEqual(sum(enveloppes.values()), 4175726.00, places=2)


class TestGroupementParLot(unittest.TestCase):
    """Un bloc par marche : le lot est l'unite qui porte une enveloppe."""

    @classmethod
    def setUpClass(cls):
        cls.analyzer = _charger_analyzer()
        if cls.analyzer is None:
            raise unittest.SkipTest("données du dépôt indisponibles")
        cls.operations = cls.analyzer.get_vision_operations()

    def test_un_groupe_par_marche_de_l_operation(self):
        for operation in self.operations:
            groupes, _, _, _, info = self.analyzer.collecter_ecritures_operation(
                operation["operation"]
            )
            if info is None:
                continue
            colonne = self.analyzer.df_marches.iloc[:, self.analyzer.COL_MARCHE]
            avec_ecritures = {m for m in operation["marches"] if (colonne == m).any()}
            self.assertEqual(set(groupes), avec_ecritures, operation["operation"])

    def test_enveloppe_identique_a_l_onglet_operations(self):
        """L'export et l'onglet Opérations ne peuvent plus afficher deux montants.

        Les deux lisent l'enveloppe du marché — base si elle y est, sinon la
        somme SEDIT de *toutes* ses tranches. Retenir la seule tranche de la
        première ligne sous-évaluait 5 opérations de 767 935,72 € au total.
        """
        for operation in self.operations:
            _, _, enveloppes, _, info = self.analyzer.collecter_ecritures_operation(
                operation["operation"]
            )
            if info is None:
                continue
            self.assertAlmostEqual(
                sum(enveloppes.values()),
                operation["montant_initial_total"],
                places=2,
                msg=operation["operation"],
            )

    def test_provenance_identique_a_l_onglet_operations(self):
        """La colonne « Enveloppe » du tableau annonce ce que dira l'export.

        L'une et l'autre passent par `_enveloppe_marche` : le tableau ne peut
        pas afficher « notifiée » là où le fichier écrira « reconstituée ».
        """
        for operation in self.operations:
            _, _, _, provenances, info = self.analyzer.collecter_ecritures_operation(
                operation["operation"]
            )
            if info is None or not provenances:
                continue
            pire = max(
                provenances.values(), key=lambda p: self.analyzer.RANG_PROVENANCE[p]
            )
            self.assertEqual(
                operation["provenance_enveloppe"], pire, operation["operation"]
            )

    def test_prestataire_lu_sur_l_ecriture(self):
        """Un marche tenu par un groupement garde ses cotraitants.

        Le prestataire etait lu sur la premiere ligne du marche et recopie sur
        toutes les autres, ce qui reattribuait au mandataire les lignes de ses
        cotraitants — et dependait de l'ordre de lecture des exports SEDIT.
        """
        import pandas as pd

        analyzer = self.analyzer
        colonne_marche = analyzer.df_marches.iloc[:, analyzer.COL_MARCHE]
        colonne_fournisseur = analyzer.df_marches.iloc[:, analyzer.COL_FOURNISSEUR]

        controle = 0
        for operation in self.operations:
            groupes, _, _, _, info = analyzer.collecter_ecritures_operation(
                operation["operation"]
            )
            if info is None:
                continue
            for marche, ecritures in groupes.items():
                attendus = {
                    "" if pd.isna(v) else str(v).strip()
                    for v in colonne_fournisseur[colonne_marche == marche]
                }
                obtenus = {e.fournisseur for e in ecritures}
                self.assertTrue(obtenus <= attendus, f"{marche} : {obtenus - attendus}")
                if len(attendus) > 1:
                    controle += 1
                    self.assertEqual(obtenus, attendus, marche)

        self.assertGreater(controle, 0, "aucun marché à plusieurs prestataires")

    def test_lots_d_un_meme_titulaire_gardent_chacun_leur_sous_total(self):
        """Bout en bout : deux lots du meme titulaire font deux blocs, pas un.

        Regroupes sur (fournisseur, tranche), ils n'en formaient qu'un, avec un
        sous-total et une enveloppe additionnes — le solde de chaque lot, qui
        est un marche notifie pour son propre montant, etait alors introuvable.
        """
        import pandas as pd

        analyzer = self.analyzer
        colonne_marche = analyzer.df_marches.iloc[:, analyzer.COL_MARCHE]
        colonne_fournisseur = analyzer.df_marches.iloc[:, analyzer.COL_FOURNISSEUR]

        def _titulaires(marche):
            return {
                "" if pd.isna(v) else str(v).strip()
                for v in colonne_fournisseur[colonne_marche == marche]
            }

        cible = None
        for operation in self.operations:
            lots = [m for m in operation["marches"] if (colonne_marche == m).any()]
            if len(lots) < 2:
                continue
            for rang, lot in enumerate(lots):
                if any(_titulaires(lot) & _titulaires(autre) for autre in lots[rang + 1:]):
                    cible = (operation["operation"], lots)
                    break
            if cible:
                break

        if cible is None:
            self.skipTest("aucune opération dont deux lots partagent un titulaire")

        code_operation, lots = cible
        import openpyxl

        with tempfile.TemporaryDirectory() as repertoire:
            chemin = os.path.join(repertoire, "suivi.xlsx")
            self.assertTrue(analyzer.export_suivi_financier_operation(
                code_operation, chemin, exercice_filter=None, special_export=False
            ))
            feuille = openpyxl.load_workbook(chemin)["FINANCIER"]
            sous_totaux = [
                feuille.cell(row, 2).value
                for row in range(6, feuille.max_row + 1)
                if feuille.cell(row, 15).value == "TOTAL"
            ]

        self.assertEqual(len(sous_totaux), len(lots), sous_totaux)
        for lot in lots:
            self.assertTrue(
                any(f"marché {lot}" in (libelle or "") for libelle in sous_totaux),
                f"{lot} absent des sous-totaux : {sous_totaux}",
            )


class TestOptionsExport(unittest.TestCase):
    """Le tri et le journal se choisissent separement, sans changer le resultat."""

    @classmethod
    def setUpClass(cls):
        cls.analyzer = _charger_analyzer()
        if cls.analyzer is None:
            raise unittest.SkipTest("données du dépôt indisponibles")

    def _feuille(self, **options):
        import openpyxl

        with tempfile.TemporaryDirectory() as repertoire:
            chemin = os.path.join(repertoire, "suivi.xlsx")
            options.setdefault("exercice_filter", None)
            self.assertTrue(self.analyzer.export_suivi_financier_operation(
                MARCHE, chemin, **options
            ))
            feuille = openpyxl.load_workbook(chemin)["FINANCIER"]
            return [
                tuple(feuille.cell(ligne, colonne).value for colonne in range(1, 16))
                for ligne in range(6, feuille.max_row + 1)
            ]

    def test_les_deux_options_reproduisent_l_ancien_export_special(self):
        """Le bouton en dur ne faisait rien d'autre que cocher ces deux cases.

        Il etait le seul chemin vers le tri par BDC et le journal de controle ;
        la fenetre de choix les offre pour n'importe quelle operation, et doit
        rendre exactement le meme tableau.
        """
        self.assertEqual(
            self._feuille(trier_par_bdc=True, journal=True),
            self._feuille(special_export=True),
        )

    def test_le_tri_par_bdc_ordonne_les_lignes(self):
        bdc_tries = [ligne[4] for ligne in self._feuille(trier_par_bdc=True) if ligne[14] != "TOTAL"]
        self.assertEqual(bdc_tries, sorted(bdc_tries))

    def test_le_journal_est_ecrit_a_part_du_tri(self):
        import shutil

        journal = os.path.join("run_logs", f"export_{MARCHE}.log")
        sauvegarde = journal + ".test"
        existait = os.path.exists(journal)
        if existait:
            shutil.copy2(journal, sauvegarde)
        try:
            if existait:
                os.remove(journal)
            self._feuille(trier_par_bdc=True, journal=False)
            self.assertFalse(os.path.exists(journal))

            self._feuille(trier_par_bdc=False, journal=True)
            self.assertTrue(os.path.exists(journal))
        finally:
            if existait:
                shutil.move(sauvegarde, journal)
            elif os.path.exists(journal):
                os.remove(journal)

    def test_exercices_disponibles_couvre_tout_le_jeu(self):
        globaux = self.analyzer.exercices_disponibles()
        self.assertEqual(globaux[0], "Tous")
        self.assertGreater(len(globaux), 1)

        # La liste d'une operation est incluse dans la liste globale.
        pour_operation = self.analyzer.get_exercices_for_operation(MARCHE)
        self.assertEqual(pour_operation[0], "Tous")
        self.assertTrue(set(pour_operation) <= set(globaux))

    def test_exercice_filtre_reduit_le_tableau(self):
        exercices = [e for e in self.analyzer.get_exercices_for_operation(MARCHE) if e != "Tous"]
        self.assertTrue(exercices, "aucun exercice sur l'opération de référence")

        complet = self._feuille()
        partiel = self._feuille(exercice_filter=exercices[0])
        self.assertLess(len(partiel), len(complet))


class TestLibelleSousTotal(unittest.TestCase):
    """L'intitule du sous-total n'enonce que ce que le bloc contient."""

    @staticmethod
    def _resultat(paires):
        from suivi_financier_agg import LigneSuivi, ResultatSuivi

        resultat = ResultatSuivi()
        resultat.lignes = [
            LigneSuivi(tranche_libelle=tranche, fournisseur=fournisseur)
            for tranche, fournisseur in paires
        ]
        return resultat

    def _libelle(self, marche, paires):
        from marches_module import MarchesAnalyzer

        return MarchesAnalyzer._libelle_sous_total(marche, self._resultat(paires))

    def test_marche_mono_tranche_mono_prestataire(self):
        self.assertEqual(
            self._libelle("2020_11_1", [("TF", "WURTH FRANCE")] * 3),
            "Sous-total marché 2020_11_1 — TF — WURTH FRANCE",
        )

    def test_tranches_dans_l_ordre_du_contrat(self):
        libelle = self._libelle("2024_1_5", [("TO2", "X"), ("TO3", "X"), ("TF", "X")])
        self.assertIn("TF, TO2, TO3", libelle)

    def test_groupement_au_dela_de_trois_titulaires(self):
        paires = [(f"TF", f"E{n}") for n in range(4)]
        self.assertTrue(self._libelle("M", paires).endswith("4 prestataires"))

    def test_tranche_absente_omise(self):
        self.assertEqual(
            self._libelle("MC146_1", [("", "DISPANO")]),
            "Sous-total marché MC146_1 — DISPANO",
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)
