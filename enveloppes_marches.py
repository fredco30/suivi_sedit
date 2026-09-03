#!/usr/bin/env python3
"""Saisie en lot des enveloppes contractuelles des marches.

L'enveloppe initiale d'un marche est son montant notifie : elle figure dans
l'acte d'engagement, jamais dans les exports SEDIT. Ceux-ci ne portent que les
montants *engages* a ce jour, qui n'en sont pas une approximation utilisable --
sur les trois marches ou les deux valeurs sont connues, l'ecart va de 1,06 a
115 fois.

Ce module ne devine donc aucun montant. Il produit un tableau des marches a
renseigner, avec tout ce que le logiciel sait deja d'eux, puis reinjecte en base
la colonne remplie a la main.

    python enveloppes_marches.py exporter --sortie enveloppes_a_saisir.xlsx
    # ... saisir la colonne « ENVELOPPE CONTRACTUELLE TTC » ...
    python enveloppes_marches.py importer enveloppes_a_saisir.xlsx        # simulation
    python enveloppes_marches.py importer enveloppes_a_saisir.xlsx --appliquer
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from datetime import datetime
from typing import Dict, List, Optional

from marches_module import MarchesAnalyzer
from regenerer_suivis import (
    BaseSuiviLectureSeule,
    CACHE_REGENERATION,
    DB_PAR_DEFAUT,
    SOURCE_PAR_DEFAUT,
    cache_par_defaut,
)
from suivi_financier_agg import STATUT_FACTURE, agreger_ecritures

COLONNE_ENVELOPPE = "ENVELOPPE CONTRACTUELLE TTC"
COLONNE_MARCHE = "N° MARCHÉ"

EN_TETES = [
    COLONNE_MARCHE,
    COLONNE_ENVELOPPE,
    "OPÉRATION(S)",
    "FOURNISSEUR",
    "MONTANT ACTUELLEMENT EN BASE",
    "ENGAGÉ À CE JOUR (repère, PAS l'enveloppe)",
    "FACTURÉ À CE JOUR",
    "NB BDC",
]


def collecter_marches(analyzer: MarchesAnalyzer) -> List[Dict]:
    """Dresse l'inventaire des marchés vus dans les écritures, avec leurs repères."""
    par_marche: Dict[str, Dict] = {}

    for operation in analyzer.get_vision_operations():
        code_operation = operation["operation"]
        groupes, montants, _, _, info = analyzer.collecter_ecritures_operation(code_operation)
        if info is None:
            continue

        for ecritures in groupes.values():
            resultat = agreger_ecritures(ecritures, montants)
            for ligne in resultat.lignes:
                fiche = par_marche.setdefault(ligne.marche, {
                    "marche": ligne.marche,
                    "operations": set(),
                    "fournisseur": ligne.fournisseur,
                    "engage": 0.0,
                    "facture": 0.0,
                    "bdc": set(),
                })
                fiche["operations"].add(code_operation)
                fiche["engage"] += ligne.montant_impute
                if ligne.statut == STATUT_FACTURE:
                    fiche["facture"] += ligne.montant_impute
                fiche["bdc"].add(ligne.cle_bdc)

    return [par_marche[cle] for cle in sorted(par_marche)]


def montants_en_base(db_path: str) -> Dict[str, float]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        return {
            row["code_marche"]: row["montant_initial_manuel"]
            for row in conn.execute(
                "SELECT code_marche, montant_initial_manuel FROM marches"
            )
        }
    finally:
        conn.close()


def exporter(sortie: str, source, db_path: str, cache_path: Optional[str] = None) -> int:
    """Produit le tableau des enveloppes à renseigner."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    analyzer = MarchesAnalyzer(
        source, database=BaseSuiviLectureSeule(db_path), use_cache=True,
        cache_path=cache_path or cache_par_defaut(source),
    )
    if not analyzer.load_data():
        raise RuntimeError(f"Chargement des données impossible depuis « {source} »")

    fiches = collecter_marches(analyzer)
    deja = montants_en_base(db_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Enveloppes"

    ws.cell(1, 1, "ENVELOPPES CONTRACTUELLES DES MARCHÉS — À RENSEIGNER").font = Font(
        bold=True, size=14
    )
    ws.cell(2, 1,
            "Saisir dans la colonne B le montant TTC notifié de chaque marché "
            "(acte d'engagement, avenants compris). La colonne F ne donne qu'un "
            "repère : c'est le montant engagé à ce jour, pas l'enveloppe.")
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=8)

    fill_entete = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_saisie = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for col, entete in enumerate(EN_TETES, 1):
        cellule = ws.cell(5, col, entete)
        cellule.font = Font(bold=True, size=11)
        cellule.fill = fill_entete
        cellule.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ligne_idx = 6
    a_saisir = 0
    for fiche in fiches:
        montant_base = deja.get(fiche["marche"])
        if not montant_base:
            a_saisir += 1

        ws.cell(ligne_idx, 1, fiche["marche"])
        ws.cell(ligne_idx, 2, montant_base or None).fill = fill_saisie
        ws.cell(ligne_idx, 3, ", ".join(sorted(fiche["operations"])))
        ws.cell(ligne_idx, 4, fiche["fournisseur"])
        ws.cell(ligne_idx, 5, montant_base or None)
        ws.cell(ligne_idx, 6, round(fiche["engage"], 2))
        ws.cell(ligne_idx, 7, round(fiche["facture"], 2))
        ws.cell(ligne_idx, 8, len(fiche["bdc"]))
        for col in (2, 5, 6, 7):
            ws.cell(ligne_idx, col).number_format = '#,##0.00 €'
        ligne_idx += 1

    for col, largeur in enumerate([18, 26, 28, 30, 24, 30, 20, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = largeur
    ws.freeze_panes = "A6"

    wb.save(sortie)
    print(f"[OK] {len(fiches)} marchés dans « {sortie} » "
          f"({a_saisir} enveloppes à saisir, {len(fiches) - a_saisir} déjà en base)")
    return 0


def importer(fichier: str, db_path: str, appliquer: bool = False) -> int:
    """Réinjecte en base la colonne d'enveloppes saisie à la main."""
    from openpyxl import load_workbook

    wb = load_workbook(fichier, data_only=True)
    ws = wb["Enveloppes"] if "Enveloppes" in wb.sheetnames else wb.active

    # Repérer les colonnes par leur en-tête plutôt que par leur position :
    # le fichier revient d'un tableur, ses colonnes ont pu bouger.
    ligne_entete = None
    colonnes: Dict[str, int] = {}
    for ligne in range(1, min(ws.max_row, 20) + 1):
        valeurs = {
            str(ws.cell(ligne, col).value).strip(): col
            for col in range(1, ws.max_column + 1)
            if ws.cell(ligne, col).value is not None
        }
        if COLONNE_MARCHE in valeurs and COLONNE_ENVELOPPE in valeurs:
            ligne_entete, colonnes = ligne, valeurs
            break

    if ligne_entete is None:
        print(f"[ERREUR] Colonnes « {COLONNE_MARCHE} » et « {COLONNE_ENVELOPPE} » "
              f"introuvables dans {fichier}")
        return 1

    col_marche = colonnes[COLONNE_MARCHE]
    col_montant = colonnes[COLONNE_ENVELOPPE]

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    existants = {
        row["code_marche"]: row["montant_initial_manuel"]
        for row in conn.execute("SELECT code_marche, montant_initial_manuel FROM marches")
    }

    creations: List[tuple] = []
    modifications: List[tuple] = []
    inchanges = 0
    ignores: List[str] = []

    for ligne in range(ligne_entete + 1, ws.max_row + 1):
        code = ws.cell(ligne, col_marche).value
        if not code:
            continue
        code = str(code).strip()

        brut = ws.cell(ligne, col_montant).value
        if brut is None or str(brut).strip() == "":
            ignores.append(code)
            continue
        try:
            montant = float(str(brut).replace("€", "").replace(" ", "").replace(",", "."))
        except ValueError:
            print(f"[ERREUR] Ligne {ligne} ({code}) : montant illisible {brut!r}")
            return 1
        if montant <= 0:
            ignores.append(code)
            continue

        if code not in existants:
            creations.append((code, montant))
        elif abs((existants[code] or 0) - montant) > 0.005:
            modifications.append((code, existants[code], montant))
        else:
            inchanges += 1

    print(f"{len(creations)} marché(s) à créer, {len(modifications)} à modifier, "
          f"{inchanges} inchangé(s), {len(ignores)} sans montant saisi")
    for code, montant in creations[:10]:
        print(f"   + {code:<16}{montant:>16,.2f} €".replace(",", " "))
    if len(creations) > 10:
        print(f"   … {len(creations) - 10} autres créations")
    for code, avant, apres in modifications[:10]:
        print(f"   ~ {code:<16}{(avant or 0):>16,.2f} € → {apres:,.2f} €".replace(",", " "))

    if not appliquer:
        print("\nSimulation : relancer avec --appliquer pour écrire en base.")
        conn.close()
        return 0

    horodatage = datetime.now().isoformat(timespec="seconds")
    conn.executemany(
        "INSERT INTO marches (code_marche, montant_initial_manuel, type_marche, last_update) "
        "VALUES (?, ?, 'CLASSIQUE', ?)",
        [(code, montant, horodatage) for code, montant in creations]
    )
    conn.executemany(
        "UPDATE marches SET montant_initial_manuel = ?, last_update = ? WHERE code_marche = ?",
        [(apres, horodatage, code) for code, _, apres in modifications]
    )
    conn.commit()
    conn.close()

    print(f"\n[OK] Base mise à jour : {len(creations)} créations, "
          f"{len(modifications)} modifications.")
    print("Régénérer ensuite les suivis :")
    print('   python regenerer_suivis.py --tout --source "data_sources/factures*.xls" '
          '--sortie exports')
    return 0


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    sous = parser.add_subparsers(dest="commande", required=True)

    p_export = sous.add_parser("exporter", help="Produire le tableau des enveloppes à saisir")
    p_export.add_argument("--sortie", default="enveloppes_a_saisir.xlsx")
    p_export.add_argument("--source", nargs="+", default=[SOURCE_PAR_DEFAUT])
    p_export.add_argument("--db", default=DB_PAR_DEFAUT)
    p_export.add_argument("--cache", default=None)

    p_import = sous.add_parser("importer", help="Réinjecter le tableau rempli en base")
    p_import.add_argument("fichier")
    p_import.add_argument("--db", default=DB_PAR_DEFAUT)
    p_import.add_argument("--appliquer", action="store_true",
                          help="Écrire en base (sinon simulation)")

    args = parser.parse_args(argv)

    if args.commande == "exporter":
        return exporter(args.sortie, args.source, args.db, args.cache)
    return importer(args.fichier, args.db, args.appliquer)


if __name__ == "__main__":
    sys.exit(main())
