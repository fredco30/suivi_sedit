
import sys
import os
import sqlite3
from datetime import datetime, date, timedelta

import pandas as pd

# Import des modules pour le suivi des marchés
from marches_module import MarchesAnalyzer
from marches_models import (
    MarchesGlobauxTableModel, MarchesTranchesTableModel,
    MarchesGlobauxProxy, MarchesTranchesProxy,
    OperationsTableModel, OperationsProxy,
    HistoriqueTableModel, HistoriqueProxy
)
from marches_dialogs import EditMarcheDialog

from PyQt5.QtCore import (
    Qt,
    QAbstractTableModel,
    QModelIndex,
    QSortFilterProxyModel,
    QTimer,
    QTime,
    QEvent,
)
from PyQt5.QtGui import QIcon, QBrush, QColor, QFont, QStandardItemModel, QStandardItem
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QFileDialog,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QToolBar,
    QToolButton,
    QAction,
    QMessageBox,
    QInputDialog,
    QSystemTrayIcon,
    QMenu,
    QStyle,
    QComboBox,
    QDialog,
    QHeaderView,
    QFormLayout,
    QPushButton,
    QSpinBox,
    QDialogButtonBox,
    QWidget,
    QHBoxLayout,
    QLabel,
    QVBoxLayout,
    QLineEdit,
    QCheckBox,    QTimeEdit,
 )


DB_NAME = "suivi_commandes.db"


# ============== FONCTION UTILITAIRE WORD WRAP ==============


# ============== WIDGET COMBOBOX À CHOIX MULTIPLES ==============

class CheckableComboBox(QComboBox):
    """ComboBox avec checkboxes pour sélection multiple."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        # Utiliser un modèle standard pour les items
        self._model = QStandardItemModel(self)
        self.setModel(self._model)
        
        # Connecter les événements
        self.view().pressed.connect(self.handle_item_pressed)
        self._changed = False
        
        # Empêcher la fermeture au clic
        self.view().viewport().installEventFilter(self)
        
        # Ajouter l'item "Tous" par défaut
        self._add_all_item()
    
    def _add_all_item(self):
        """Ajoute l'item spécial '[Tous]' en première position."""
        item = QStandardItem("[Tous]")
        item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        item.setCheckState(Qt.Unchecked)
        font = item.font()
        font.setBold(True)
        item.setFont(font)
        self._model.insertRow(0, item)
    
    def eventFilter(self, obj, event):
        """Empêche la fermeture de la liste au clic sans dupliquer le traitement du clic."""
        # On consomme simplement le MouseButtonRelease sur la vue pour éviter
        # que la popup ne se ferme, mais le changement d'état de la case
        # est déjà géré par le signal 'pressed' connecté à handle_item_pressed.
        if obj == self.view().viewport() and event.type() == QEvent.MouseButtonRelease:
            return True
        return super().eventFilter(obj, event)
    
    def handle_item_pressed(self, index):
        """Gère le clic sur un item."""
        item = self._model.itemFromIndex(index)
        if item and item.isCheckable():
            # Si c'est l'item "[Tous]"
            if index.row() == 0 and item.text() == "[Tous]":
                new_state = Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked
                item.setCheckState(new_state)
                # Cocher/décocher tous les autres items
                for i in range(1, self._model.rowCount()):
                    other_item = self._model.item(i, 0)
                    if other_item:
                        other_item.setCheckState(new_state)
            else:
                # Item normal
                if item.checkState() == Qt.Checked:
                    item.setCheckState(Qt.Unchecked)
                else:
                    item.setCheckState(Qt.Checked)
                # Mettre à jour l'état de "[Tous]"
                self._update_all_item()
            
            self._changed = True
            # Émettre un signal de changement
            self.currentIndexChanged.emit(index.row())
    
    def _update_all_item(self):
        """Met à jour l'état de l'item '[Tous]' selon les autres items."""
        all_item = self._model.item(0, 0)
        if not all_item or all_item.text() != "[Tous]":
            return
        
        # Vérifier si tous les items (sauf [Tous]) sont cochés
        all_checked = True
        any_checked = False
        for i in range(1, self._model.rowCount()):
            item = self._model.item(i, 0)
            if item:
                if item.checkState() == Qt.Checked:
                    any_checked = True
                else:
                    all_checked = False
        
        # Mettre à jour [Tous]
        if all_checked and any_checked:
            all_item.setCheckState(Qt.Checked)
        else:
            all_item.setCheckState(Qt.Unchecked)
    
    def item_checked(self, index):
        """Vérifie si un item est coché."""
        item = self._model.item(index, 0)
        return item.checkState() == Qt.Checked if item else False
    
    def checked_items(self):
        """Retourne la liste des items cochés (sauf [Tous])."""
        checked = []
        for i in range(1, self._model.rowCount()):  # Commencer à 1 pour ignorer [Tous]
            item = self._model.item(i, 0)
            if item and item.checkState() == Qt.Checked:
                checked.append(item.text())
        return checked
    
    def clear_selection(self):
        """Décoche tous les items."""
        for i in range(self._model.rowCount()):
            item = self._model.item(i, 0)
            if item:
                item.setCheckState(Qt.Unchecked)
    
    def addItem(self, text):
        """Ajoute un item avec checkbox."""
        item = QStandardItem(text)
        item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        item.setCheckState(Qt.Unchecked)
        self._model.appendRow(item)
    
    def addItems(self, texts):
        """Ajoute plusieurs items avec checkboxes."""
        for text in texts:
            self.addItem(text)
    
    def clear(self):
        """Vide le combobox et recrée l'item [Tous]."""
        self._model.clear()
        self._add_all_item()
    
    def currentText(self):
        """Retourne le texte des items cochés."""
        checked = self.checked_items()
        if not checked:
            return "Aucun"
        elif len(checked) == 1:
            return checked[0]
        else:
            return f"{len(checked)} sélectionnés"


def today_iso():
    return date.today().isoformat()


def parse_date_safe(value):
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            d = datetime.strptime(str(value), fmt).date()
            return d.isoformat()
        except ValueError:
            continue
    try:
        d = pd.to_datetime(value, dayfirst=True).date()
        return d.isoformat()
    except Exception:
        return None


def smart_word_wrap(text, max_width=40):
    """Découpe intelligente du texte en respectant les mots."""
    if not text or len(text) <= max_width:
        return text
    
    words = text.split()
    lines = []
    current_line = []
    current_length = 0
    
    for word in words:
        word_length = len(word)
        # Si le mot seul dépasse la largeur max, on le coupe
        if word_length > max_width:
            if current_line:
                lines.append(" ".join(current_line))
                current_line = []
                current_length = 0
            # Découper le mot long
            for i in range(0, len(word), max_width):
                lines.append(word[i:i+max_width])
        else:
            # Si ajouter ce mot dépasse la largeur
            if current_length + word_length + len(current_line) > max_width:
                if current_line:
                    lines.append(" ".join(current_line))
                current_line = [word]
                current_length = word_length
            else:
                current_line.append(word)
                current_length += word_length
    
    # Ajouter la dernière ligne
    if current_line:
        lines.append(" ".join(current_line))
    
    return "\n".join(lines)


def compute_facture_status(num_facture, montant_service_fait, date_facture):
    """Calcule le statut intelligent d'une facture."""
    # Convertir et nettoyer le numéro de facture
    num_facture_str = str(num_facture).strip().lower() if num_facture is not None else ""
    has_facture = num_facture_str and num_facture_str not in ("", "none", "nan", "nat", "na")
    
    # Vérifier le montant service fait
    try:
        msf = float(montant_service_fait) if montant_service_fait is not None else 0.0
        has_service_fait = msf > 0.01  # Tolérance pour les arrondis
    except (ValueError, TypeError):
        has_service_fait = False
    
    # Vérifier la date
    date_str = str(date_facture).strip().lower() if date_facture is not None else ""
    has_date = date_str and date_str not in ("", "none", "nan", "nat", "na")
    
    if has_facture and has_service_fait:
        return "Facturée"
    elif has_service_fait and not has_facture:
        return "Service fait"
    elif has_date and not has_service_fait:
        return "En attente de paiement"
    else:
        return "A vérifier"


class Database:
    def __init__(self, path):
        self.path = path
        self.conn = sqlite3.connect(self.path)
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

    def _init_schema(self):
        cur = self.conn.cursor()

        # Commandes
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS commandes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exercice TEXT,
                num_commande TEXT,
                fournisseur TEXT,
                libelle TEXT,
                date_commande TEXT,
                marche TEXT,
                service_emetteur TEXT,
                montant_ttc REAL,
                section TEXT,
                article_fonction TEXT,
                article_nature TEXT,
                statut TEXT,
                rappel_actif INTEGER,
                frequence_rappel_jours INTEGER,
                prochaine_date_rappel TEXT,
                date_envoi TEXT,
                notes TEXT,
                last_update TEXT,
                montant_facture REAL DEFAULT 0,
                reste_a_facturer REAL DEFAULT 0,
                statut_facturation TEXT,
                UNIQUE(exercice, num_commande)
            )
            """
        )

        # Factures
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS factures (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exercice TEXT,
                num_facture TEXT,
                code_mouvement TEXT,
                fournisseur TEXT,
                libelle TEXT,
                date_facture TEXT,
                montant_ttc REAL,
                montant_service_fait REAL DEFAULT 0,
                marche TEXT,
                statut_facture TEXT,
                rappel_facture_actif INTEGER,
                frequence_rappel_facture_jours INTEGER,
                prochaine_date_rappel_facture TEXT,
                notes TEXT,
                last_update TEXT
            )
            """
        )

        # Migration: ajouter montant_service_fait si inexistant
        try:
            cur.execute("SELECT montant_service_fait FROM factures LIMIT 1")
        except sqlite3.OperationalError:
            cur.execute("ALTER TABLE factures ADD COLUMN montant_service_fait REAL DEFAULT 0")
            self.conn.commit()

        # Migration: ajouter les colonnes pour l'analyse des marchés
        try:
            cur.execute("SELECT tranche FROM factures LIMIT 1")
        except sqlite3.OperationalError:
            print("[MIGRATION] Ajout de la colonne 'tranche'")
            cur.execute("ALTER TABLE factures ADD COLUMN tranche TEXT")
            self.conn.commit()

        try:
            cur.execute("SELECT commande FROM factures LIMIT 1")
        except sqlite3.OperationalError:
            print("[MIGRATION] Ajout de la colonne 'commande'")
            cur.execute("ALTER TABLE factures ADD COLUMN commande TEXT")
            self.conn.commit()

        try:
            cur.execute("SELECT num_mandat FROM factures LIMIT 1")
        except sqlite3.OperationalError:
            print("[MIGRATION] Ajout de la colonne 'num_mandat'")
            cur.execute("ALTER TABLE factures ADD COLUMN num_mandat TEXT")
            self.conn.commit()

        try:
            cur.execute("SELECT montant_initial FROM factures LIMIT 1")
        except sqlite3.OperationalError:
            print("[MIGRATION] Ajout de la colonne 'montant_initial'")
            cur.execute("ALTER TABLE factures ADD COLUMN montant_initial REAL")
            self.conn.commit()

        # Config
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS config (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """
        )

        # Config Exports
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS config_exports (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                logo_path TEXT,
                nom_entreprise TEXT,
                adresse_1 TEXT,
                adresse_2 TEXT,
                code_postal TEXT,
                ville TEXT,
                inclure_couleurs INTEGER DEFAULT 1,
                lignes_filtrees_uniquement INTEGER DEFAULT 1
            )
            """
        )
        
        # Initialiser les valeurs par défaut si la table est vide
        cur.execute("SELECT COUNT(*) as count FROM config_exports")
        if cur.fetchone()["count"] == 0:
            cur.execute(
                """
                INSERT INTO config_exports
                (id, nom_entreprise, adresse_1, adresse_2, code_postal, ville, inclure_couleurs, lignes_filtrees_uniquement)
                VALUES (1, '', '', '', '', '', 1, 1)
                """
            )

        # Table Marchés - Stocke les ajustements manuels et métadonnées
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS marches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code_marche TEXT UNIQUE NOT NULL,
                libelle TEXT,
                fournisseur TEXT,
                montant_initial_manuel REAL,
                date_notification TEXT,
                date_debut TEXT,
                date_fin_prevue TEXT,
                notes TEXT,
                last_update TEXT
            )
            """
        )

        # Table Avenants - Historique des modifications du marché
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS avenants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code_marche TEXT NOT NULL,
                numero_avenant INTEGER,
                libelle TEXT,
                montant REAL,
                type_modification TEXT,
                date_avenant TEXT,
                motif TEXT,
                last_update TEXT,
                FOREIGN KEY (code_marche) REFERENCES marches(code_marche)
            )
            """
        )

        # Table Tranches - Définition des tranches fermes et optionnelles
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS tranches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code_marche TEXT NOT NULL,
                code_tranche TEXT NOT NULL,
                libelle TEXT,
                montant REAL,
                ordre INTEGER,
                last_update TEXT,
                FOREIGN KEY (code_marche) REFERENCES marches(code_marche),
                UNIQUE(code_marche, code_tranche)
            )
            """
        )

        # Table Import Tracking - Suivi des imports de fichiers Excel
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS import_tracking (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT UNIQUE NOT NULL,
                file_path TEXT,
                file_hash TEXT,
                file_size INTEGER,
                last_modified_date TEXT,
                import_date TEXT,
                import_type TEXT,
                records_imported INTEGER,
                status TEXT,
                error_message TEXT
            )
            """
        )

        # Migration: ajouter source_file aux commandes
        try:
            cur.execute("SELECT source_file FROM commandes LIMIT 1")
        except sqlite3.OperationalError:
            cur.execute("ALTER TABLE commandes ADD COLUMN source_file TEXT")
            self.conn.commit()

        # Migration: ajouter source_file aux factures
        try:
            cur.execute("SELECT source_file FROM factures LIMIT 1")
        except sqlite3.OperationalError:
            cur.execute("ALTER TABLE factures ADD COLUMN source_file TEXT")
            self.conn.commit()

        # Migration: ajouter type_marche aux marchés
        try:
            cur.execute("SELECT type_marche FROM marches LIMIT 1")
        except sqlite3.OperationalError:
            print("[MIGRATION] Ajout de la colonne 'type_marche'")
            # Par défaut, tous les marchés existants sont considérés comme "CLASSIQUE"
            cur.execute("ALTER TABLE marches ADD COLUMN type_marche TEXT DEFAULT 'CLASSIQUE'")
            self.conn.commit()

        self.conn.commit()

    # -------------- Config --------------

    def get_config(self, key, default=None):
        cur = self.conn.cursor()
        cur.execute("SELECT value FROM config WHERE key = ?", (key,))
        row = cur.fetchone()
        return row["value"] if row else default

    def set_config(self, key, value):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO config(key, value) VALUES(?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, str(value)),
        )
        self.conn.commit()

    # -------------- Config Exports --------------

    def get_config_exports(self):
        """Récupère la configuration des exports."""
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM config_exports WHERE id = 1")
        row = cur.fetchone()
        if row:
            return {
                "logo_path": row["logo_path"] or "",
                "nom_entreprise": row["nom_entreprise"] or "",
                "adresse_1": row["adresse_1"] or "",
                "adresse_2": row["adresse_2"] or "",
                "code_postal": row["code_postal"] or "",
                "ville": row["ville"] or "",
                "inclure_couleurs": bool(row["inclure_couleurs"]),
                "lignes_filtrees_uniquement": bool(row["lignes_filtrees_uniquement"])
            }
        # Valeurs par défaut
        return {
            "logo_path": "",
            "nom_entreprise": "",
            "adresse_1": "",
            "adresse_2": "",
            "code_postal": "",
            "ville": "",
            "inclure_couleurs": True,
            "lignes_filtrees_uniquement": True
        }

    def save_config_exports(self, config):
        """Sauvegarde la configuration des exports."""
        cur = self.conn.cursor()
        cur.execute(
            """
            UPDATE config_exports SET
                logo_path = ?,
                nom_entreprise = ?,
                adresse_1 = ?,
                adresse_2 = ?,
                code_postal = ?,
                ville = ?,
                inclure_couleurs = ?,
                lignes_filtrees_uniquement = ?
            WHERE id = 1
            """,
            (
                config.get("logo_path", ""),
                config.get("nom_entreprise", ""),
                config.get("adresse_1", ""),
                config.get("adresse_2", ""),
                config.get("code_postal", ""),
                config.get("ville", ""),
                1 if config.get("inclure_couleurs", True) else 0,
                1 if config.get("lignes_filtrees_uniquement", True) else 0
            )
        )
        self.conn.commit()

    # -------------- Marchés et Avenants --------------

    def get_marche(self, code_marche):
        """Récupère les données d'un marché."""
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM marches WHERE code_marche = ?", (code_marche,))
        return cur.fetchone()

    def upsert_marche(self, code_marche, data):
        """Crée ou met à jour un marché."""
        cur = self.conn.cursor()
        now = datetime.now().isoformat(timespec="seconds")

        cur.execute("SELECT * FROM marches WHERE code_marche = ?", (code_marche,))
        row = cur.fetchone()

        if row is None:
            cur.execute(
                """
                INSERT INTO marches (
                    code_marche, libelle, fournisseur, type_marche, montant_initial_manuel,
                    date_notification, date_debut, date_fin_prevue, notes, last_update
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    code_marche,
                    data.get("libelle", ""),
                    data.get("fournisseur", ""),
                    data.get("type_marche", "CLASSIQUE"),
                    data.get("montant_initial_manuel"),
                    data.get("date_notification"),
                    data.get("date_debut"),
                    data.get("date_fin_prevue"),
                    data.get("notes", ""),
                    now
                )
            )
        else:
            cur.execute(
                """
                UPDATE marches SET
                    libelle = ?,
                    fournisseur = ?,
                    type_marche = ?,
                    montant_initial_manuel = ?,
                    date_notification = ?,
                    date_debut = ?,
                    date_fin_prevue = ?,
                    notes = ?,
                    last_update = ?
                WHERE code_marche = ?
                """,
                (
                    data.get("libelle", ""),
                    data.get("fournisseur", ""),
                    data.get("type_marche", "CLASSIQUE"),
                    data.get("montant_initial_manuel"),
                    data.get("date_notification"),
                    data.get("date_debut"),
                    data.get("date_fin_prevue"),
                    data.get("notes", ""),
                    now,
                    code_marche
                )
            )

        self.conn.commit()

    def get_avenants(self, code_marche):
        """Récupère tous les avenants d'un marché."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT * FROM avenants WHERE code_marche = ? ORDER BY numero_avenant",
            (code_marche,)
        )
        return cur.fetchall()

    def add_avenant(self, code_marche, data):
        """Ajoute un avenant à un marché."""
        cur = self.conn.cursor()
        now = datetime.now().isoformat(timespec="seconds")

        cur.execute(
            """
            INSERT INTO avenants (
                code_marche, numero_avenant, libelle, montant,
                type_modification, date_avenant, motif, last_update
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                code_marche,
                data.get("numero_avenant"),
                data.get("libelle", ""),
                data.get("montant", 0),
                data.get("type_modification", "Augmentation"),
                data.get("date_avenant"),
                data.get("motif", ""),
                now
            )
        )

        self.conn.commit()
        return cur.lastrowid

    def update_avenant(self, avenant_id, data):
        """Met à jour un avenant existant."""
        cur = self.conn.cursor()
        now = datetime.now().isoformat(timespec="seconds")

        cur.execute(
            """
            UPDATE avenants SET
                numero_avenant = ?,
                libelle = ?,
                montant = ?,
                type_modification = ?,
                date_avenant = ?,
                motif = ?,
                last_update = ?
            WHERE id = ?
            """,
            (
                data.get("numero_avenant"),
                data.get("libelle", ""),
                data.get("montant", 0),
                data.get("type_modification", "Augmentation"),
                data.get("date_avenant"),
                data.get("motif", ""),
                now,
                avenant_id
            )
        )

        self.conn.commit()

    def delete_avenant(self, avenant_id):
        """Supprime un avenant."""
        cur = self.conn.cursor()
        cur.execute("DELETE FROM avenants WHERE id = ?", (avenant_id,))
        self.conn.commit()

    # -------------- Tranches --------------

    def get_tranches(self, code_marche):
        """Récupère toutes les tranches d'un marché."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT * FROM tranches WHERE code_marche = ? ORDER BY ordre",
            (code_marche,)
        )
        return cur.fetchall()

    def add_tranche(self, code_marche, data):
        """Ajoute une tranche à un marché."""
        cur = self.conn.cursor()
        now = datetime.now().isoformat(timespec="seconds")

        cur.execute(
            """
            INSERT INTO tranches (
                code_marche, code_tranche, libelle, montant, ordre, last_update
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                code_marche,
                data.get("code_tranche"),
                data.get("libelle", ""),
                data.get("montant", 0),
                data.get("ordre", 0),
                now
            )
        )

        self.conn.commit()
        return cur.lastrowid

    def update_tranche(self, tranche_id, data):
        """Met à jour une tranche existante."""
        cur = self.conn.cursor()
        now = datetime.now().isoformat(timespec="seconds")

        cur.execute(
            """
            UPDATE tranches SET
                code_tranche = ?,
                libelle = ?,
                montant = ?,
                ordre = ?,
                last_update = ?
            WHERE id = ?
            """,
            (
                data.get("code_tranche"),
                data.get("libelle", ""),
                data.get("montant", 0),
                data.get("ordre", 0),
                now,
                tranche_id
            )
        )

        self.conn.commit()

    def delete_tranche(self, tranche_id):
        """Supprime une tranche."""
        cur = self.conn.cursor()
        cur.execute("DELETE FROM tranches WHERE id = ?", (tranche_id,))
        self.conn.commit()

    def get_montant_total_marche(self, code_marche):
        """
        Calcule le montant total d'un marché.

        Pour les marchés CLASSIQUES :
        - Montant initial (TF) + Somme des tranches optionnelles (TO1, TO2...) + avenants
        - montant_initial_manuel = TF
        - tranches table = TO1, TO2, TO3...

        Pour les marchés BDC :
        - montant_initial_manuel + avenants
        """
        marche = self.get_marche(code_marche)
        if not marche:
            return 0.0

        # Vérifier le type de marché
        try:
            type_marche = marche["type_marche"] if marche["type_marche"] else "CLASSIQUE"
        except (KeyError, IndexError):
            type_marche = "CLASSIQUE"

        # Montant de base
        montant_initial = marche["montant_initial_manuel"] or 0.0

        # Pour les marchés CLASSIQUES : ajouter les tranches optionnelles (TOs)
        if type_marche == "CLASSIQUE":
            tranches = self.get_tranches(code_marche)
            if tranches:
                # Les tranches sont les TOs (TO1, TO2, TO3...)
                # On les ajoute au montant_initial qui représente la TF
                montant_tos = sum((t["montant"] or 0.0) for t in tranches)
                print(f"[MONTANT TOTAL] {code_marche}: TF={montant_initial} € + TOs={montant_tos} €")
                montant_initial += montant_tos

        # Ajouter les avenants
        avenants = self.get_avenants(code_marche)
        for avenant in avenants:
            montant_avenant = avenant["montant"] or 0.0
            if avenant["type_modification"] == "Diminution":
                montant_avenant = -montant_avenant
            montant_initial += montant_avenant

        return montant_initial

    # -------------- Commandes --------------

    def upsert_commande(self, data):
        cur = self.conn.cursor()
        exercice = data["exercice"]
        num_commande = data["num_commande"]

        cur.execute(
            "SELECT * FROM commandes WHERE exercice = ? AND num_commande = ?",
            (exercice, num_commande),
        )
        row = cur.fetchone()

        now = datetime.now().isoformat(timespec="seconds")

        if row is None:
            statut = "A suivre"
            rappel_actif = 1
            try:
                freq = int(self.get_config("global_reminder_days", "7") or 7)
            except ValueError:
                freq = 7
            if freq < 0:
                freq = 0

            t_val = self.get_config("global_reminder_time", "09:00") or "09:00"
            try:
                hh, mm = map(int, t_val.split(":"))
                hh = max(0, min(23, hh))
                mm = max(0, min(59, mm))
            except Exception:
                hh, mm = 9, 0

            date_base = date.today() + timedelta(days=freq)
            prochaine_dt = datetime.combine(date_base, datetime.min.time()).replace(hour=hh, minute=mm)
            prochaine = prochaine_dt.strftime("%Y-%m-%d %H:%M")

            cur.execute(
                """
                INSERT INTO commandes (
                    exercice, num_commande, fournisseur, libelle, date_commande,
                    marche, service_emetteur, montant_ttc, section,
                    article_fonction, article_nature,
                    statut, rappel_actif, frequence_rappel_jours,
                    prochaine_date_rappel, date_envoi,
                    notes, last_update, montant_facture, reste_a_facturer, statut_facturation,
                    source_file
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    exercice,
                    num_commande,
                    data.get("fournisseur") or "",
                    data.get("libelle") or "",
                    data.get("date_commande"),
                    data.get("marche") or "",
                    data.get("service_emetteur") or "",
                    data.get("montant_ttc"),
                    data.get("section") or "",
                    data.get("article_fonction") or "",
                    data.get("article_nature") or "",
                    statut,
                    int(rappel_actif),
                    freq,
                    prochaine,
                    None,
                    "",
                    now,
                    0.0,
                    data.get("montant_ttc") or 0.0,
                    "Non facturée",
                    data.get("source_file"),
                ),
            )
        else:
            cur.execute(
                """
                UPDATE commandes SET
                    fournisseur = ?,
                    libelle = ?,
                    date_commande = ?,
                    marche = ?,
                    service_emetteur = ?,
                    montant_ttc = ?,
                    section = ?,
                    article_fonction = ?,
                    article_nature = ?,
                    last_update = ?,
                    source_file = ?
                WHERE exercice = ? AND num_commande = ?
                """,
                (
                    data.get("fournisseur") or "",
                    data.get("libelle") or "",
                    data.get("date_commande"),
                    data.get("marche") or "",
                    data.get("service_emetteur") or "",
                    data.get("montant_ttc"),
                    data.get("section") or "",
                    data.get("article_fonction") or "",
                    data.get("article_nature") or "",
                    now,
                    data.get("source_file"),
                    exercice,
                    num_commande,
                ),
            )

        self.conn.commit()

    def fetch_all_commandes(self):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT * FROM commandes
            ORDER BY date_commande DESC, fournisseur ASC
            """
        )
        return cur.fetchall()

    def update_statut_for_ids(self, ids, statut, disable_rappel=False):
        if not ids:
            return
        cur = self.conn.cursor()
        date_envoi = today_iso() if statut == "Envoyée" else None
        if disable_rappel:
            cur.execute(
                f"""
                UPDATE commandes
                SET statut = ?, rappel_actif = 0, prochaine_date_rappel = NULL,
                    date_envoi = ?
                WHERE id IN ({",".join("?" * len(ids))})
                """,
                (statut, date_envoi, *ids),
            )
        else:
            cur.execute(
                f"""
                UPDATE commandes
                SET statut = ?, date_envoi = ?
                WHERE id IN ({",".join("?" * len(ids))})
                """,
                (statut, date_envoi, *ids),
            )
        self.conn.commit()

    def reschedule_rappel_for_ids(self, ids):
        if not ids:
            return
        try:
            days = int(self.get_config("global_reminder_days", "7") or 7)
        except ValueError:
            days = 7
        if days < 0:
            days = 0

        t_val = self.get_config("global_reminder_time", "09:00") or "09:00"
        try:
            hh, mm = map(int, t_val.split(":"))
            hh = max(0, min(23, hh))
            mm = max(0, min(59, mm))
        except Exception:
            hh, mm = 9, 0

        date_base = date.today() + timedelta(days=days)
        prochaine_dt = datetime.combine(date_base, datetime.min.time()).replace(hour=hh, minute=mm)
        date_next = prochaine_dt.strftime("%Y-%m-%d %H:%M")
        cur = self.conn.cursor()
        cur.execute(
            f"""
            UPDATE commandes
            SET rappel_actif = 1,
                frequence_rappel_jours = ?,
                prochaine_date_rappel = ?
            WHERE id IN ({",".join("?" * len(ids))})
            """,
            (days, date_next, *ids),
        )
        self.conn.commit()

    def due_reminders(self):
        cur = self.conn.cursor()
        # On travaille en date+heure : les rappels sont déclenchés si l'échéance est passée.
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        cur.execute(
            """
            SELECT * FROM commandes
            WHERE statut != 'Envoyée'
              AND rappel_actif = 1
              AND prochaine_date_rappel IS NOT NULL
              AND prochaine_date_rappel <= ?
            ORDER BY prochaine_date_rappel ASC
            """,
            (now_str,),
        )
        return cur.fetchall()

    def all_active_reminders(self):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT * FROM commandes
            WHERE statut != 'Envoyée'
              AND rappel_actif = 1
              AND prochaine_date_rappel IS NOT NULL
            ORDER BY prochaine_date_rappel ASC
            """
        )
        return cur.fetchall()

    # -------------- Factures --------------

    def upsert_facture(self, data):
        cur = self.conn.cursor()
        num_facture = data.get("num_facture") or ""
        code_mouvement = data.get("code_mouvement") or ""
        exercice = data.get("exercice") or ""

        cur.execute(
            """
            SELECT * FROM factures
            WHERE exercice = ? AND num_facture = ? AND code_mouvement = ?
            """,
            (exercice, num_facture, code_mouvement),
        )
        row = cur.fetchone()
        now = datetime.now().isoformat(timespec="seconds")
        
        # Calcul du statut intelligent
        statut = compute_facture_status(
            data.get("num_facture"),
            data.get("montant_service_fait"),
            data.get("date_facture")
        )

        if row is None:
            cur.execute(
                """
                INSERT INTO factures (
                    exercice, num_facture, code_mouvement,
                    fournisseur, libelle, date_facture,
                    montant_ttc, montant_service_fait, marche,
                    statut_facture,
                    rappel_facture_actif, frequence_rappel_facture_jours,
                    prochaine_date_rappel_facture, notes, last_update
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    exercice,
                    num_facture,
                    code_mouvement,
                    data.get("fournisseur") or "",
                    data.get("libelle") or "",
                    data.get("date_facture"),
                    data.get("montant_ttc"),
                    data.get("montant_service_fait") or 0.0,
                    data.get("marche") or "",
                    statut,
                    0,
                    None,
                    None,
                    "",
                    now,
                ),
            )
        else:
            cur.execute(
                """
                UPDATE factures SET
                    fournisseur = ?,
                    libelle = ?,
                    date_facture = ?,
                    montant_ttc = ?,
                    montant_service_fait = ?,
                    marche = ?,
                    statut_facture = ?,
                    last_update = ?
                WHERE id = ?
                """,
                (
                    data.get("fournisseur") or "",
                    data.get("libelle") or "",
                    data.get("date_facture"),
                    data.get("montant_ttc"),
                    data.get("montant_service_fait") or 0.0,
                    data.get("marche") or "",
                    statut,
                    now,
                    row["id"],
                ),
            )

        self.conn.commit()

    def fetch_all_factures(self):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT * FROM factures
            ORDER BY date_facture DESC, fournisseur ASC
            """
        )
        return cur.fetchall()

    # -------------- Agrégations facturation --------------


    def recompute_facturation(self):
        """Calcule et met à jour pour chaque commande :
        - montant_facture : somme des Montant service fait des factures liées (code_mouvement = num_commande)
        - reste_a_facturer : montant_commande - montant_facture
        - statut_facturation : Non / Partiellement / Totalement facturée

        Règle complémentaire :
        - si statut_facturation est "Partiellement facturée" ou "Totalement facturée",
          le statut de la commande est automatiquement positionné à "Envoyée".
        """
        cur = self.conn.cursor()

        # Somme des montants service fait par commande (code_mouvement)
        cur.execute(
            """
            SELECT code_mouvement, SUM(montant_service_fait) AS total
            FROM factures
            WHERE code_mouvement IS NOT NULL AND code_mouvement != ''
            GROUP BY code_mouvement
            """
        )
        totals = {row["code_mouvement"]: (row["total"] or 0.0) for row in cur.fetchall()}

        # Recalcule pour chaque commande
        cur.execute("SELECT id, num_commande, montant_ttc, statut FROM commandes")
        rows = cur.fetchall()
        for row in rows:
            num = row["num_commande"]
            mt_cmd = row["montant_ttc"] or 0.0
            mt_fact = totals.get(num, 0.0)
            reste = mt_cmd - mt_fact

            # Statut de facturation
            if mt_cmd <= 0:
                statut_fact = "Non facturée" if mt_fact <= 0 else "Totalement facturée"
            else:
                if mt_fact <= 0:
                    statut_fact = "Non facturée"
                elif mt_fact + 1e-6 < mt_cmd:
                    statut_fact = "Partiellement facturée"
                else:
                    statut_fact = "Totalement facturée"

            # Statut commande auto si la facturation est engagée
            statut_cmd = row["statut"]
            if statut_fact in ("Partiellement facturée", "Totalement facturée"):
                new_statut = "Envoyée"
            else:
                new_statut = statut_cmd or "A suivre"

            cur.execute(
                """
                UPDATE commandes
                SET montant_facture = ?, reste_a_facturer = ?, statut_facturation = ?, statut = ?
                WHERE id = ?
                """,
                (mt_fact, reste, statut_fact, new_statut, row["id"]),
            )

        self.conn.commit()


    def fetch_facturation_synthese(self):
        """Vue synthétique commandes vs factures.

        Partie 1 :
            - Une ligne par commande (table 'commandes')
            - Montants et statut issus des champs déjà calculés de 'commandes'
            - 'derniere_facture' = MAX(date_facture) des factures liées.

        Partie 2 :
            - Une ligne par facture dont le code_mouvement ne correspond à aucune commande
            - Permet d'identifier les factures orphelines.
        """
        cur = self.conn.cursor()
        cur.execute(
            """
            -- Partie 1 : Commandes avec agrégats de factures liées
            SELECT
                c.exercice AS exercice,
                c.num_commande AS num_commande,
                c.fournisseur AS fournisseur,
                c.libelle AS libelle,
                c.marche AS marche,
                c.montant_ttc AS montant_commande,
                c.montant_facture AS montant_facture,
                c.reste_a_facturer AS reste_a_facturer,
                c.statut_facturation AS statut_facturation,
                MAX(f.date_facture) AS derniere_facture
            FROM commandes c
            LEFT JOIN factures f
              ON f.code_mouvement = c.num_commande
            GROUP BY c.id

            UNION ALL

            -- Partie 2 : Factures sans commande correspondante (factures orphelines)
            SELECT
                f.exercice AS exercice,
                NULL AS num_commande,
                f.fournisseur AS fournisseur,
                f.libelle AS libelle,
                f.marche AS marche,
                0.0 AS montant_commande,
                f.montant_service_fait AS montant_facture,
                -f.montant_service_fait AS reste_a_facturer,
                CASE
                    WHEN f.montant_service_fait > 0 THEN 'Facturée sans commande'
                    ELSE 'Non facturée'
                END AS statut_facturation,
                f.date_facture AS derniere_facture
            FROM factures f
            WHERE
                f.code_mouvement IS NULL
                OR f.code_mouvement = ''
                OR f.code_mouvement NOT IN (
                    SELECT num_commande
                    FROM commandes
                    WHERE num_commande IS NOT NULL
                )
            ORDER BY derniere_facture DESC, fournisseur ASC
            """
        )
        return cur.fetchall()

    # -------------- Import Tracking & Incremental Import --------------

    def calculate_file_hash(self, filepath: str) -> str:
        """Calcule le hash MD5 d'un fichier."""
        import hashlib
        hash_md5 = hashlib.md5()
        try:
            with open(filepath, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception as e:
            print(f"[ERREUR] Impossible de calculer le hash de {filepath}: {e}")
            return ""

    def get_import_record(self, filename: str):
        """Récupère l'enregistrement d'import pour un fichier."""
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM import_tracking WHERE filename = ?", (filename,))
        return cur.fetchone()

    def record_import(self, filename: str, filepath: str, file_hash: str, file_size: int,
                     import_type: str, records_imported: int, status: str, error_message: str = None):
        """Enregistre ou met à jour un import dans la table de tracking."""
        import os
        from datetime import datetime

        last_modified = datetime.fromtimestamp(os.path.getmtime(filepath)).isoformat()
        import_date = datetime.now().isoformat()

        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO import_tracking
            (filename, file_path, file_hash, file_size, last_modified_date, import_date,
             import_type, records_imported, status, error_message)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(filename) DO UPDATE SET
                file_path = excluded.file_path,
                file_hash = excluded.file_hash,
                file_size = excluded.file_size,
                last_modified_date = excluded.last_modified_date,
                import_date = excluded.import_date,
                import_type = excluded.import_type,
                records_imported = excluded.records_imported,
                status = excluded.status,
                error_message = excluded.error_message
            """,
            (filename, filepath, file_hash, file_size, last_modified, import_date,
             import_type, records_imported, status, error_message)
        )
        self.conn.commit()

    def should_import_file(self, filepath: str) -> tuple:
        """
        Vérifie si un fichier doit être importé.
        Retourne (should_import: bool, reason: str)
        """
        import os

        filename = os.path.basename(filepath)
        file_hash = self.calculate_file_hash(filepath)

        if not file_hash:
            return False, "Impossible de calculer le hash"

        record = self.get_import_record(filename)

        if record is None:
            return True, "Nouveau fichier"

        # Si l'import précédent a échoué, réessayer
        if record["status"] == "error":
            return True, "Import précédent en erreur - réessai"

        if record["file_hash"] != file_hash:
            return True, "Fichier modifié (hash différent)"

        return False, "Fichier déjà importé et inchangé"

    def scan_excel_directory(self, directory: str, pattern: str = "*.xlsx") -> dict:
        """
        Scanne un dossier pour trouver les fichiers Excel à importer.
        Retourne un dict avec 'commandes' et 'factures' files.
        """
        import os
        import glob

        if not os.path.exists(directory):
            print(f"[ERREUR] Le dossier {directory} n'existe pas")
            return {"commandes": [], "factures": []}

        # Chercher à la fois les .xls et .xlsx
        all_files = []
        for ext in ["*.xls", "*.xlsx"]:
            search_pattern = os.path.join(directory, ext)
            all_files.extend(glob.glob(search_pattern))

        result = {
            "commandes": [],
            "factures": []
        }

        for filepath in all_files:
            filename = os.path.basename(filepath).lower()

            # Déterminer le type de fichier basé sur le nom
            if "commande" in filename:
                result["commandes"].append(filepath)
            elif "facture" in filename:
                result["factures"].append(filepath)
            else:
                # Fichier Excel non reconnu
                print(f"[WARNING] Fichier ignoré (nom non reconnu): {filename}")

        return result


# ------------------ MODELES TABLE ------------------


COMMANDES_COLUMNS = [
    ("exercice", "Exercice"),
    ("num_commande", "N°\nCommande"),
    ("fournisseur", "Fournisseur"),
    ("libelle", "Libellé"),
    ("date_commande", "Date de\nla commande"),
    ("marche", "Marché"),
    ("service_emetteur", "Service\németteur"),
    ("montant_ttc", "Montant\nTTC"),
    ("section", "Section"),
    ("article_fonction", "Article\nfonction"),
    ("article_nature", "Article\nnature"),
    ("statut", "Statut"),
    ("date_envoi", "Date\nd'envoi"),
    ("prochaine_date_rappel", "Prochain\nrappel"),
    ("montant_facture", "Montant\nfacturé"),
    ("reste_a_facturer", "Reste à\nfacturer"),
    ("statut_facturation", "Statut\nfacturation"),
]


class CommandesTableModel(QAbstractTableModel):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.rows = []
        self.refresh()

    def refresh(self):
        self.beginResetModel()
        self.rows = list(self.db.fetch_all_commandes())
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self.rows)

    def columnCount(self, parent=QModelIndex()):
        return len(COMMANDES_COLUMNS)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        row = self.rows[index.row()]
        key, _ = COMMANDES_COLUMNS[index.column()]

        if role == Qt.DisplayRole:
            value = row[key]
            if key in ("date_commande", "prochaine_date_rappel", "date_envoi") and value:
                try:
                    # Supporte date seule ou date+heure pour les rappels.
                    if len(value) > 10:
                        d = datetime.strptime(value, "%Y-%m-%d %H:%M")
                        return d.strftime("%d/%m/%Y %H:%M")
                    else:
                        d = datetime.strptime(value, "%Y-%m-%d")
                        return d.strftime("%d/%m/%Y")
                except Exception:
                    return value
            if key in ("montant_ttc", "montant_facture", "reste_a_facturer") and value is not None:
                return f"{float(value):,.2f}".replace(",", " ").replace(".", ",")
            if key == "libelle" and value:
                return smart_word_wrap(str(value), 50)
            if key == "fournisseur" and value:
                return smart_word_wrap(str(value), 30)
            return value

        if role == Qt.TextAlignmentRole:
            if key in ("montant_ttc", "montant_facture", "reste_a_facturer", "date_commande", "prochaine_date_rappel", "date_envoi"):
                return Qt.AlignRight | Qt.AlignVCenter
            return Qt.AlignLeft | Qt.AlignVCenter

        if role == Qt.BackgroundRole:
            statut = row["statut"]
            if statut == "Envoyée":
                return QBrush(QColor("#d0ffd0"))
            if statut == "A suivre":
                pr = row["prochaine_date_rappel"]
                if pr:
                    try:
                        # Rappel en retard si la date/heure est passée.
                        if len(pr) > 10:
                            d = datetime.strptime(pr, "%Y-%m-%d %H:%M")
                            if d < datetime.now():
                                return QBrush(QColor("#ffd0d0"))
                        else:
                            d = datetime.strptime(pr, "%Y-%m-%d").date()
                            if d < date.today():
                                return QBrush(QColor("#ffd0d0"))
                    except Exception:
                        pass
            # Coloration facturation
            stat_fact = row["statut_facturation"]
            if stat_fact == "Totalement facturée":
                return QBrush(QColor("#e0ffe0"))
            if stat_fact == "Partiellement facturée":
                return QBrush(QColor("#fff3cd"))
            if stat_fact == "Non facturée":
                return QBrush(QColor("#ffe0e0"))

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return COMMANDES_COLUMNS[section][1]
        return section + 1

    def get_row_id(self, row_index):
        if 0 <= row_index < len(self.rows):
            return self.rows[row_index]["id"]
        return None


class CommandesProxy(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.filter_status = "Tous"
        self.filter_fournisseur = ""
        self.filter_facturation = "Tous"
        self.filter_marche = "Tous"
        # Nouveaux filtres multiples
        self.filter_article_fonction = []
        self.filter_article_nature = []
        self.filter_service_emetteur = []
        # Filtre de recherche par N° de commande
        self.filter_num_commande = ""

    def setStatusFilter(self, status):
        self.filter_status = status
        self.invalidateFilter()

    def setFournisseurFilter(self, text):
        self.filter_fournisseur = text or ""
        self.invalidateFilter()

    def setFacturationFilter(self, text):
        self.filter_facturation = text or ""
        self.invalidateFilter()

    def setMarcheFilter(self, text):
        self.filter_marche = text or "Tous"
        self.invalidateFilter()

    def setNumCommandeFilter(self, text):
        """Filtre par numéro de commande."""
        self.filter_num_commande = text or ""
        self.invalidateFilter()
    
    def setArticleFonctionFilter(self, values):
        """Filtre par article fonction (liste de valeurs)."""
        self.filter_article_fonction = values if values else []
        self.invalidateFilter()
    
    def setArticleNatureFilter(self, values):
        """Filtre par article nature (liste de valeurs)."""
        self.filter_article_nature = values if values else []
        self.invalidateFilter()
    
    def setServiceEmetteurFilter(self, values):
        """Filtre par service émetteur (liste de valeurs)."""
        self.filter_service_emetteur = values if values else []
        self.invalidateFilter()

    def lessThan(self, left, right):
        src = self.sourceModel()
        col = left.column()
        key, _ = COMMANDES_COLUMNS[col]

        lv = src.rows[left.row()][key]
        rv = src.rows[right.row()][key]

        if key in ("date_commande", "prochaine_date_rappel", "date_envoi"):
            try:
                ld = datetime.strptime(lv, "%Y-%m-%d").date() if lv else date.min
            except Exception:
                ld = date.min
            try:
                rd = datetime.strptime(rv, "%Y-%m-%d").date() if rv else date.min
            except Exception:
                rd = date.min
            return ld < rd

        if key in ("montant_ttc", "montant_facture", "reste_a_facturer"):
            try:
                lf = float(lv) if lv is not None else 0.0
            except Exception:
                lf = 0.0
            try:
                rf = float(rv) if rv is not None else 0.0
            except Exception:
                rf = 0.0
            return lf < rf

        return str(lv) < str(rv)

    def filterAcceptsRow(self, source_row, source_parent):
        model = self.sourceModel()
        row = model.rows[source_row]

        if self.filter_status != "Tous":
            if row["statut"] != self.filter_status:
                return False

        if self.filter_fournisseur:
            fournisseur = row["fournisseur"] or ""
            if self.filter_fournisseur.lower() not in fournisseur.lower():
                return False

        if self.filter_facturation != "Tous":
            if row["statut_facturation"] != self.filter_facturation:
                return False

        if self.filter_marche != "Tous":
            marche = row["marche"] or ""
            if self.filter_marche != marche:
                return False
        
        # Nouveaux filtres multiples
        if self.filter_article_fonction:
            article_fonction = row["article_fonction"] or ""
            if article_fonction not in self.filter_article_fonction:
                return False

        if self.filter_article_nature:
            article_nature = row["article_nature"] or ""
            if article_nature not in self.filter_article_nature:
                return False

        if self.filter_service_emetteur:
            service_emetteur = row["service_emetteur"] or ""
            if service_emetteur not in self.filter_service_emetteur:
                return False

        # Filtre par numéro de commande
        if self.filter_num_commande:
            num_commande = str(row["num_commande"] or "")
            if self.filter_num_commande.lower() not in num_commande.lower():
                return False

        return True


# ============== NOUVEAU: MODELE FACTURES ==============

FACTURES_COLUMNS = [
    ("exercice", "Exercice"),
    ("num_facture", "N° facture"),
    ("code_mouvement", "Code\nmouvement"),
    ("fournisseur", "Fournisseur"),
    ("libelle", "Libellé"),
    ("date_facture", "Date service\nfait"),
    ("montant_service_fait", "Montant\nservice fait"),
    ("montant_ttc", "Montant\nTTC"),
    ("statut_facture", "Statut"),
    ("marche", "Marché"),
]


class FacturesTableModel(QAbstractTableModel):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.rows = []
        self.refresh()

    def refresh(self):
        self.beginResetModel()
        self.rows = list(self.db.fetch_all_factures())
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self.rows)

    def columnCount(self, parent=QModelIndex()):
        return len(FACTURES_COLUMNS)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        row = self.rows[index.row()]
        key, _ = FACTURES_COLUMNS[index.column()]

        if role == Qt.DisplayRole:
            value = row[key]
            if key == "date_facture" and value:
                try:
                    d = datetime.strptime(value, "%Y-%m-%d").date()
                    return d.strftime("%d/%m/%Y")
                except Exception:
                    return value
            if key in ("montant_service_fait", "montant_ttc") and value is not None:
                return f"{float(value):,.2f}".replace(",", " ").replace(".", ",")
            if key == "libelle" and value:
                return smart_word_wrap(str(value), 50)
            if key == "fournisseur" and value:
                return smart_word_wrap(str(value), 30)
            return value or ""

        if role == Qt.TextAlignmentRole:
            if key in ("montant_service_fait", "montant_ttc", "date_facture"):
                return Qt.AlignRight | Qt.AlignVCenter
            return Qt.AlignLeft | Qt.AlignVCenter

        if role == Qt.BackgroundRole:
            statut = row["statut_facture"] or ""
            if statut == "Facturée":
                return QBrush(QColor("#d0ffd0"))
            elif statut == "Service fait":
                return QBrush(QColor("#fff3cd"))
            elif statut == "En attente de paiement":
                return QBrush(QColor("#ffe0e0"))

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return FACTURES_COLUMNS[section][1]
        return section + 1

    def get_row_id(self, row_index):
        if 0 <= row_index < len(self.rows):
            return self.rows[row_index]["id"]
        return None


class FacturesProxy(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.filter_statut = "Tous"
        self.filter_fournisseur = ""
        self.filter_exercice = "Tous"
        self.filter_marche = "Tous"
        # Filtres de recherche
        self.filter_num_commande = ""
        self.filter_num_facture = ""

    def setStatutFilter(self, text):
        self.filter_statut = text or "Tous"
        self.invalidateFilter()

    def setFournisseurFilter(self, text):
        self.filter_fournisseur = text or ""
        self.invalidateFilter()

    def setExerciceFilter(self, text):
        self.filter_exercice = text or "Tous"
        self.invalidateFilter()

    def setMarcheFilter(self, text):
        self.filter_marche = text or "Tous"
        self.invalidateFilter()

    def setNumCommandeFilter(self, text):
        """Filtre par numéro de commande."""
        self.filter_num_commande = text or ""
        self.invalidateFilter()

    def setNumFactureFilter(self, text):
        """Filtre par numéro de facture."""
        self.filter_num_facture = text or ""
        self.invalidateFilter()

    def lessThan(self, left, right):
        src = self.sourceModel()
        col = left.column()
        key, _ = FACTURES_COLUMNS[col]

        lv = src.rows[left.row()][key]
        rv = src.rows[right.row()][key]

        if key == "date_facture":
            try:
                ld = datetime.strptime(lv, "%Y-%m-%d").date() if lv else date.min
            except Exception:
                ld = date.min
            try:
                rd = datetime.strptime(rv, "%Y-%m-%d").date() if rv else date.min
            except Exception:
                rd = date.min
            return ld < rd

        if key in ("montant_service_fait", "montant_ttc"):
            try:
                lf = float(lv) if lv is not None else 0.0
            except Exception:
                lf = 0.0
            try:
                rf = float(rv) if rv is not None else 0.0
            except Exception:
                rf = 0.0
            return lf < rf

        return str(lv or "") < str(rv or "")

    def filterAcceptsRow(self, source_row, source_parent):
        model = self.sourceModel()
        row = model.rows[source_row]

        if self.filter_statut != "Tous":
            if (row["statut_facture"] or "") != self.filter_statut:
                return False

        if self.filter_fournisseur:
            fournisseur = row["fournisseur"] or ""
            if self.filter_fournisseur.lower() not in fournisseur.lower():
                return False

        if self.filter_exercice != "Tous":
            if (row["exercice"] or "") != self.filter_exercice:
                return False

        if self.filter_marche != "Tous":
            marche = row["marche"] or ""
            if self.filter_marche != marche:
                return False

        # Filtre par numéro de commande (via code_mouvement)
        if self.filter_num_commande:
            code_mouvement = str(row["code_mouvement"] or "")
            if self.filter_num_commande.lower() not in code_mouvement.lower():
                return False

        # Filtre par numéro de facture
        if self.filter_num_facture:
            num_facture = str(row["num_facture"] or "")
            if self.filter_num_facture.lower() not in num_facture.lower():
                return False

        return True


# ============== NOUVEAU: MODELE FACTURATION ==============

FACTURATION_COLUMNS = [
    ("exercice", "Exercice"),
    ("num_commande", "N° Commande"),
    ("fournisseur", "Fournisseur"),
    ("libelle", "Libellé"),
    ("marche", "Marché"),
    ("montant_commande", "Montant\ncommande"),
    ("montant_facture", "Montant\nfacturé"),
    ("reste_a_facturer", "Reste à\nfacturer"),
    ("statut_facturation", "Statut\nfacturation"),
    ("derniere_facture", "Dernière\nfacture"),
]


class FacturationTableModel(QAbstractTableModel):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.rows = []
        self.refresh()

    def refresh(self):
        self.beginResetModel()
        self.rows = list(self.db.fetch_facturation_synthese())
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self.rows)

    def columnCount(self, parent=QModelIndex()):
        return len(FACTURATION_COLUMNS)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        row = self.rows[index.row()]
        key, _ = FACTURATION_COLUMNS[index.column()]

        if role == Qt.DisplayRole:
            value = row[key]
            if key == "derniere_facture" and value:
                try:
                    d = datetime.strptime(value, "%Y-%m-%d").date()
                    return d.strftime("%d/%m/%Y")
                except Exception:
                    return value
            if key in ("montant_commande", "montant_facture", "reste_a_facturer") and value is not None:
                return f"{float(value):,.2f}".replace(",", " ").replace(".", ",")
            if key == "fournisseur" and value:
                return smart_word_wrap(str(value), 30)
            if key == "libelle" and value:
                return smart_word_wrap(str(value), 50)
            return value or ""

        if role == Qt.TextAlignmentRole:
            if key in ("montant_commande", "montant_facture", "reste_a_facturer"):
                return Qt.AlignRight | Qt.AlignVCenter
            return Qt.AlignLeft | Qt.AlignVCenter

        if role == Qt.BackgroundRole:
            statut = row["statut_facturation"] or ""
            if statut == "Totalement facturée":
                return QBrush(QColor("#e0ffe0"))
            elif statut == "Partiellement facturée":
                return QBrush(QColor("#fff3cd"))
            elif statut == "Non facturée":
                return QBrush(QColor("#ffe0e0"))

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return FACTURATION_COLUMNS[section][1]
        return section + 1


class FacturationProxy(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.filter_statut = "Tous"
        self.filter_fournisseur = ""
        self.filter_marche = "Tous"

    def setStatutFilter(self, text):
        self.filter_statut = text or "Tous"
        self.invalidateFilter()

    def setFournisseurFilter(self, text):
        self.filter_fournisseur = text or ""
        self.invalidateFilter()

    def setMarcheFilter(self, text):
        self.filter_marche = text or "Tous"
        self.invalidateFilter()

    def lessThan(self, left, right):
        src = self.sourceModel()
        col = left.column()
        key, _ = FACTURATION_COLUMNS[col]

        lv = src.rows[left.row()][key]
        rv = src.rows[right.row()][key]

        if key == "derniere_facture":
            try:
                ld = datetime.strptime(lv, "%Y-%m-%d").date() if lv else date.min
            except Exception:
                ld = date.min
            try:
                rd = datetime.strptime(rv, "%Y-%m-%d").date() if rv else date.min
            except Exception:
                rd = date.min
            return ld < rd

        if key in ("montant_commande", "montant_facture", "reste_a_facturer"):
            try:
                lf = float(lv) if lv is not None else 0.0
            except Exception:
                lf = 0.0
            try:
                rf = float(rv) if rv is not None else 0.0
            except Exception:
                rf = 0.0
            return lf < rf

        return str(lv or "") < str(rv or "")

    def filterAcceptsRow(self, source_row, source_parent):
        model = self.sourceModel()
        row = model.rows[source_row]

        if self.filter_statut != "Tous":
            if (row["statut_facturation"] or "") != self.filter_statut:
                return False

        if self.filter_fournisseur:
            fournisseur = row["fournisseur"] or ""
            if self.filter_fournisseur.lower() not in fournisseur.lower():
                return False

        if self.filter_marche != "Tous":
            marche = row["marche"] or ""
            if self.filter_marche != marche:
                return False

        return True


# ------------------ DIALOGUE CONFIG ------------------


class ConfigDialog(QDialog):
    def __init__(self, db: Database, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Configuration globale")
        self.resize(600, 500)

        # Créer un QTabWidget pour les différentes sections
        tabs = QTabWidget(self)
        
        # ========== ONGLET GÉNÉRAL ==========
        tab_general = QWidget()
        form_general = QFormLayout(tab_general)

        # Dossier source des fichiers Excel
        self.source_dir_edit = QLineEdit(self)
        self.source_dir_edit.setText(db.get_config("excel_source_directory", "") or "")
        btn_source_dir = QPushButton("Parcourir…", self)
        btn_source_dir.clicked.connect(self.browse_source_dir)
        w_source_dir = QWidget(self)
        l_source_dir = QHBoxLayout(w_source_dir)
        l_source_dir.setContentsMargins(0, 0, 0, 0)
        l_source_dir.addWidget(self.source_dir_edit)
        l_source_dir.addWidget(btn_source_dir)
        form_general.addRow("Dossier des fichiers Excel:", w_source_dir)

        # Intervalle rappels (minutes)
        self.interval_spin = QSpinBox(self)
        self.interval_spin.setRange(1, 1440)
        val = db.get_config("reminder_interval_minutes", "5")
        try:
            val = int(val)
        except ValueError:
            val = 5
        self.interval_spin.setValue(max(1, val))
        self.interval_spin.setSuffix(" min")
        form_general.addRow("Intervalle de vérification des rappels:", self.interval_spin)

        # Délai rappels commandes (jours)
        # Délai rappels commandes (jours)
        self.rem_cmd_spin = QSpinBox(self)
        self.rem_cmd_spin.setRange(0, 365)
        val = db.get_config("global_reminder_days", "7")
        try:
            val = int(val)
        except ValueError:
            val = 7
        self.rem_cmd_spin.setValue(max(0, val))
        self.rem_cmd_spin.setSuffix(" jours")
        form_general.addRow("Délai rappel commandes (jours):", self.rem_cmd_spin)

        # Heure des rappels commandes
        self.rem_cmd_time = QTimeEdit(self)
        self.rem_cmd_time.setDisplayFormat("HH:mm")
        t_val = db.get_config("global_reminder_time", "09:00") or "09:00"
        try:
            hh, mm = map(int, t_val.split(":"))
        except ValueError:
            hh, mm = (9, 0)
        self.rem_cmd_time.setTime(QTime(hh, mm))
        form_general.addRow("Heure des rappels commandes:", self.rem_cmd_time)

        # Rappels par mail
        self.email_reminders_check = QCheckBox("Envoyer aussi les rappels par mail (si Outlook est disponible)", self)
        enabled_val = db.get_config("email_reminders_enabled", "0")
        self.email_reminders_check.setChecked(enabled_val == "1")
        form_general.addRow(self.email_reminders_check)

        self.email_reminders_to = QLineEdit(self)
        self.email_reminders_to.setText(db.get_config("email_reminders_to", ""))
        form_general.addRow("Destinataire des rappels mail:", self.email_reminders_to)

        tabs.addTab(tab_general, "Général")
        
        # ========== ONGLET EXPORTS ==========
        tab_exports = QWidget()
        form_exports = QFormLayout(tab_exports)
        
        # Charger la config exports
        config_exp = db.get_config_exports()
        
        # Logo entreprise
        self.logo_path_edit = QLineEdit(self)
        self.logo_path_edit.setText(config_exp.get("logo_path", ""))
        btn_logo = QPushButton("Parcourir…", self)
        btn_logo.clicked.connect(self.browse_logo)
        w_logo = QWidget(self)
        l_logo = QHBoxLayout(w_logo)
        l_logo.setContentsMargins(0, 0, 0, 0)
        l_logo.addWidget(self.logo_path_edit)
        l_logo.addWidget(btn_logo)
        form_exports.addRow("Logo entreprise (PNG/JPG, 200x80px):", w_logo)
        
        # Nom entreprise
        self.nom_entreprise_edit = QLineEdit(self)
        self.nom_entreprise_edit.setText(config_exp.get("nom_entreprise", ""))
        form_exports.addRow("Nom entreprise:", self.nom_entreprise_edit)
        
        # Adresse ligne 1
        self.adresse_1_edit = QLineEdit(self)
        self.adresse_1_edit.setText(config_exp.get("adresse_1", ""))
        form_exports.addRow("Adresse ligne 1:", self.adresse_1_edit)
        
        # Adresse ligne 2
        self.adresse_2_edit = QLineEdit(self)
        self.adresse_2_edit.setText(config_exp.get("adresse_2", ""))
        form_exports.addRow("Adresse ligne 2:", self.adresse_2_edit)
        
        # Code postal
        self.code_postal_edit = QLineEdit(self)
        self.code_postal_edit.setText(config_exp.get("code_postal", ""))
        self.code_postal_edit.setMaxLength(10)
        form_exports.addRow("Code postal:", self.code_postal_edit)
        
        # Ville
        self.ville_edit = QLineEdit(self)
        self.ville_edit.setText(config_exp.get("ville", ""))
        form_exports.addRow("Ville:", self.ville_edit)
        
        # Séparateur
        line = QLabel("")
        line.setStyleSheet("border-top: 1px solid #ccc; margin: 10px 0;")
        form_exports.addRow(line)
        
        # Options d'export
        self.inclure_couleurs_check = QCheckBox("Inclure les couleurs dans les exports")
        self.inclure_couleurs_check.setChecked(config_exp.get("inclure_couleurs", True))
        form_exports.addRow(self.inclure_couleurs_check)
        
        self.lignes_filtrees_check = QCheckBox("Exporter uniquement les lignes visibles (filtrées)")
        self.lignes_filtrees_check.setChecked(config_exp.get("lignes_filtrees_uniquement", True))
        form_exports.addRow(self.lignes_filtrees_check)
        
        tabs.addTab(tab_exports, "Exports")

        # Boutons OK/Annuler
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addWidget(tabs)
        layout.addWidget(buttons)

    def browse_source_dir(self):
        path = QFileDialog.getExistingDirectory(
            self, "Choisir le dossier contenant les fichiers Excel", ""
        )
        if path:
            self.source_dir_edit.setText(path)
    
    def browse_logo(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Choisir le logo de l'entreprise", "", "Images (*.png *.jpg *.jpeg)"
        )
        if path:
            self.logo_path_edit.setText(path)

    def accept(self):
        # Sauvegarder config générale
        self.db.set_config("excel_source_directory", self.source_dir_edit.text().strip())
        self.db.set_config("reminder_interval_minutes", str(self.interval_spin.value()))
        self.db.set_config("global_reminder_days", str(self.rem_cmd_spin.value()))
        self.db.set_config("global_reminder_time", self.rem_cmd_time.time().toString("HH:mm"))
        self.db.set_config(
            "email_reminders_enabled",
            "1" if self.email_reminders_check.isChecked() else "0"
        )
        self.db.set_config(
            "email_reminders_to",
            self.email_reminders_to.text().strip()
        )
        # Sauvegarder config exports
        config_exp = {
            "logo_path": self.logo_path_edit.text().strip(),
            "nom_entreprise": self.nom_entreprise_edit.text().strip(),
            "adresse_1": self.adresse_1_edit.text().strip(),
            "adresse_2": self.adresse_2_edit.text().strip(),
            "code_postal": self.code_postal_edit.text().strip(),
            "ville": self.ville_edit.text().strip(),
            "inclure_couleurs": self.inclure_couleurs_check.isChecked(),
            "lignes_filtrees_uniquement": self.lignes_filtrees_check.isChecked()
        }
        self.db.save_config_exports(config_exp)
        
        super().accept()


# ------------------ FENETRE PRINCIPALE ------------------


class MainWindow(QMainWindow):
    def __init__(self, db_path):
        super().__init__()
        self.setWindowTitle("Suivi commandes / factures / marchés v5.6")
        self.resize(1400, 800)

        self.db = Database(db_path)
        self.error_log = []  # journal des erreurs pour export

        # Modèles Commandes
        self.cmd_model = CommandesTableModel(self.db)
        self.cmd_proxy = CommandesProxy(self)
        self.cmd_proxy.setSourceModel(self.cmd_model)

        # Modèles Factures
        self.fact_model = FacturesTableModel(self.db)
        self.fact_proxy = FacturesProxy(self)
        self.fact_proxy.setSourceModel(self.fact_model)

        # Modèles Facturation
        self.synth_model = FacturationTableModel(self.db)
        self.synth_proxy = FacturationProxy(self)
        self.synth_proxy.setSourceModel(self.synth_model)

        # Modèles Suivi des Marchés
        self.marches_global_model = MarchesGlobauxTableModel()
        self.marches_global_proxy = MarchesGlobauxProxy(self)
        self.marches_global_proxy.setSourceModel(self.marches_global_model)

        self.marches_tranches_model = MarchesTranchesTableModel()
        self.marches_tranches_proxy = MarchesTranchesProxy(self)
        self.marches_tranches_proxy.setSourceModel(self.marches_tranches_model)

        self.operations_model = OperationsTableModel()
        self.operations_proxy = OperationsProxy(self)
        self.operations_proxy.setSourceModel(self.operations_model)

        self.historique_model = HistoriqueTableModel()
        self.historique_proxy = HistoriqueProxy(self)
        self.historique_proxy.setSourceModel(self.historique_model)

        self.marches_analyzer = None  # Sera initialisé lors du chargement du fichier Excel

        # Vues
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)
        
        # Style moderne pour les onglets avec couleurs identitaires
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                background-color: white;
                top: -2px;
            }
            QTabBar::tab {
                background-color: #f8f9fa;
                color: #495057;
                border: 2px solid transparent;
                border-radius: 8px 8px 0 0;
                padding: 12px 24px;
                margin-right: 4px;
                font-size: 10pt;
                font-weight: normal;
                min-width: 120px;
            }
            QTabBar::tab:hover {
                background-color: #e9ecef;
                transform: translateY(-2px);
            }
            /* Onglet Commandes (index 0) - BLEU */
            QTabBar::tab:selected:first {
                background-color: #0078d4;
                color: white;
                border: 2px solid #005a9e;
                font-weight: bold;
            }
            /* Tous les onglets actifs par défaut - style de base */
            QTabBar::tab:selected {
                background-color: #0078d4;
                color: white;
                border: 2px solid #005a9e;
                font-weight: bold;
            }
        """)
        
        # Note: Les couleurs spécifiques par onglet seront gérées dynamiquement
        self.tabs.currentChanged.connect(self._update_tab_colors)

        # === ONGLET COMMANDES ===
        self.table_cmd = QTableView()
        self.table_cmd.setModel(self.cmd_proxy)
        self.table_cmd.setSortingEnabled(True)
        self.table_cmd.setSelectionBehavior(QTableView.SelectRows)
        self.table_cmd.setSelectionMode(QTableView.ExtendedSelection)
        self.table_cmd.doubleClicked.connect(self.on_cmd_double_clicked)
        self.table_cmd.setWordWrap(True)

        # Police plus petite mais lisible pour voir davantage de colonnes sans scroll
        font_cmd = self.table_cmd.font()
        if font_cmd.pointSize() > 0:
            font_cmd.setPointSize(max(8, font_cmd.pointSize() - 1))
        else:
            font_cmd.setPointSize(9)
        self.table_cmd.setFont(font_cmd)

        header = self.table_cmd.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setStyleSheet(
            "QHeaderView::section { font-weight: bold; background-color: #cfe8ff; padding: 3px; }"
        )
        # Police d’en-tête légèrement réduite aussi
        header_font = header.font()
        if header_font.pointSize() > 0:
            header_font.setPointSize(max(8, header_font.pointSize() - 1))
        else:
            header_font.setPointSize(9)
        header.setFont(header_font)

        # Colonnes auto-ajustées au contenu
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(False)

        self.table_cmd.sortByColumn(4, Qt.DescendingOrder)

        self.tabs.addTab(self.table_cmd, "📋 Commandes")

        # === ONGLET RAPPELS ===
        self.table_rappels = QTableWidget()
        self.table_rappels.setColumnCount(6)
        self.table_rappels.setHorizontalHeaderLabels(
            [
                "Exercice",
                "N° Commande",
                "Fournisseur",
                "Date commande",
                "Prochain rappel",
                "Statut",
            ]
        )
        hdr_rappels = self.table_rappels.horizontalHeader()
        hdr_rappels.setDefaultAlignment(Qt.AlignCenter)
        hdr_rappels.setStyleSheet(
            "QHeaderView::section { font-weight: bold; background-color: #cfe8ff; padding: 4px; }"
        )
        hdr_rappels.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr_rappels.setStretchLastSection(False)
        self.table_rappels.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_rappels.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_rappels.setSelectionMode(QTableWidget.SingleSelection)
        self.table_rappels.setSortingEnabled(True)
        self.tabs.addTab(self.table_rappels, "🔔 Rappels")

        # === ONGLET FACTURES (NOUVEAU AVEC TRI/FILTRES) ===
        self.table_fact = QTableView()
        self.table_fact.setModel(self.fact_proxy)
        self.table_fact.setSortingEnabled(True)
        self.table_fact.setSelectionBehavior(QTableView.SelectRows)
        self.table_fact.setSelectionMode(QTableView.SingleSelection)
        self.table_fact.setWordWrap(True)

        header_fact = self.table_fact.horizontalHeader()
        header_fact.setDefaultAlignment(Qt.AlignCenter)
        header_fact.setStyleSheet(
            "QHeaderView::section { font-weight: bold; background-color: #cfe8ff; padding: 4px; }"
        )
        header_fact.setSectionResizeMode(QHeaderView.ResizeToContents)
        header_fact.setStretchLastSection(False)
        self.table_fact.sortByColumn(4, Qt.DescendingOrder)

        self.tabs.addTab(self.table_fact, "📄 Factures")

        # === ONGLET FACTURATION (NOUVEAU AVEC TRI/FILTRES) ===
        self.table_synth = QTableView()
        self.table_synth.setModel(self.synth_proxy)
        self.table_synth.setSortingEnabled(True)
        self.table_synth.setSelectionBehavior(QTableView.SelectRows)
        self.table_synth.setSelectionMode(QTableView.SingleSelection)
        self.table_synth.setWordWrap(True)

        header_synth = self.table_synth.horizontalHeader()
        header_synth.setDefaultAlignment(Qt.AlignCenter)
        header_synth.setStyleSheet(
            "QHeaderView::section { font-weight: bold; background-color: #cfe8ff; padding: 4px; }"
        )
        header_synth.setSectionResizeMode(QHeaderView.ResizeToContents)
        header_synth.setStretchLastSection(False)
        self.table_synth.sortByColumn(4, Qt.DescendingOrder)

        self.tabs.addTab(self.table_synth, "💰 Facturation")

        # === ONGLET SUIVI DES MARCHÉS ===
        # Créer un widget container avec layout vertical
        marches_widget = QWidget()
        marches_layout = QVBoxLayout(marches_widget)
        marches_layout.setContentsMargins(5, 5, 5, 5)
        marches_layout.setSpacing(10)

        # Label et bouton de rafraîchissement pour la vision globale
        header_global = QWidget()
        header_global_layout = QHBoxLayout(header_global)
        header_global_layout.setContentsMargins(0, 0, 0, 0)

        label_global = QLabel("<b>📊 Vision globale par marché</b>")
        label_global.setStyleSheet("font-size: 11pt; color: #0078d4;")
        header_global_layout.addWidget(label_global)

        btn_refresh_marches = QPushButton("🔄 Actualiser les données")
        btn_refresh_marches.clicked.connect(self.refresh_marches_data)
        btn_refresh_marches.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #005a9e;
            }
        """)
        header_global_layout.addWidget(btn_refresh_marches)

        btn_export_marches = QPushButton("📊 Exporter Excel complet")
        btn_export_marches.clicked.connect(self.export_marches_excel)
        btn_export_marches.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        header_global_layout.addWidget(btn_export_marches)
        header_global_layout.addStretch()

        marches_layout.addWidget(header_global)

        # === FILTRES ===
        filtres_widget = QWidget()
        filtres_layout = QHBoxLayout(filtres_widget)
        filtres_layout.setContentsMargins(0, 5, 0, 5)

        # Filtre par marché
        label_filtre_marche = QLabel("🔍 Marché:")
        filtres_layout.addWidget(label_filtre_marche)

        self.edit_filtre_marche = QLineEdit()
        self.edit_filtre_marche.setPlaceholderText("Filtrer par code marché...")
        self.edit_filtre_marche.setClearButtonEnabled(True)
        self.edit_filtre_marche.setMaximumWidth(200)
        self.edit_filtre_marche.textChanged.connect(
            lambda text: self.marches_global_proxy.setMarcheFilter(text)
        )
        filtres_layout.addWidget(self.edit_filtre_marche)

        # Filtre par fournisseur
        label_filtre_fournisseur = QLabel("🏢 Fournisseur:")
        filtres_layout.addWidget(label_filtre_fournisseur)

        self.edit_filtre_fournisseur = QLineEdit()
        self.edit_filtre_fournisseur.setPlaceholderText("Filtrer par fournisseur...")
        self.edit_filtre_fournisseur.setClearButtonEnabled(True)
        self.edit_filtre_fournisseur.setMaximumWidth(200)
        self.edit_filtre_fournisseur.textChanged.connect(
            lambda text: self.marches_global_proxy.setFournisseurFilter(text)
        )
        filtres_layout.addWidget(self.edit_filtre_fournisseur)

        filtres_layout.addStretch()

        marches_layout.addWidget(filtres_widget)

        # Table vision globale
        self.table_marches_global = QTableView()
        self.table_marches_global.setModel(self.marches_global_proxy)
        self.table_marches_global.setSortingEnabled(True)
        self.table_marches_global.setSelectionBehavior(QTableView.SelectRows)
        self.table_marches_global.setSelectionMode(QTableView.SingleSelection)
        self.table_marches_global.setWordWrap(True)
        self.table_marches_global.setMaximumHeight(350)  # Limiter la hauteur

        # Connecter la sélection pour afficher le détail
        self.table_marches_global.selectionModel().selectionChanged.connect(
            self.on_marche_selection_changed
        )

        # Connecter le double-clic pour éditer
        self.table_marches_global.doubleClicked.connect(self.on_marche_double_clicked)

        header_marches_global = self.table_marches_global.horizontalHeader()
        header_marches_global.setDefaultAlignment(Qt.AlignCenter)
        header_marches_global.setStyleSheet(
            "QHeaderView::section { font-weight: bold; background-color: #cfe8ff; padding: 4px; }"
        )
        header_marches_global.setSectionResizeMode(QHeaderView.ResizeToContents)
        header_marches_global.setStretchLastSection(True)

        marches_layout.addWidget(self.table_marches_global)

        # Séparateur
        separator = QLabel()
        separator.setStyleSheet("background-color: #e0e0e0;")
        separator.setMaximumHeight(2)
        marches_layout.addWidget(separator)

        # Label pour le détail par tranches
        label_detail = QLabel("<b>🔍 Détail par tranche</b>")
        label_detail.setStyleSheet("font-size: 11pt; color: #28a745;")
        marches_layout.addWidget(label_detail)

        # Table vision détaillée (tranches)
        self.table_marches_tranches = QTableView()
        self.table_marches_tranches.setModel(self.marches_tranches_proxy)
        self.table_marches_tranches.setSortingEnabled(True)
        self.table_marches_tranches.setSelectionBehavior(QTableView.SelectRows)
        self.table_marches_tranches.setSelectionMode(QTableView.SingleSelection)
        self.table_marches_tranches.setWordWrap(True)

        header_marches_tranches = self.table_marches_tranches.horizontalHeader()
        header_marches_tranches.setDefaultAlignment(Qt.AlignCenter)
        header_marches_tranches.setStyleSheet(
            "QHeaderView::section { font-weight: bold; background-color: #d4edda; padding: 4px; }"
        )
        header_marches_tranches.setSectionResizeMode(QHeaderView.ResizeToContents)
        header_marches_tranches.setStretchLastSection(True)

        marches_layout.addWidget(self.table_marches_tranches)

        # Ajouter l'onglet
        self.tabs.addTab(marches_widget, "📈 Suivi marchés")

        # === ONGLET OPÉRATIONS ===
        operations_widget = QWidget()
        operations_layout = QVBoxLayout(operations_widget)
        operations_layout.setContentsMargins(5, 5, 5, 5)
        operations_layout.setSpacing(10)

        # En-tête
        header_operations = QWidget()
        header_operations_layout = QHBoxLayout(header_operations)
        header_operations_layout.setContentsMargins(0, 0, 0, 0)

        label_operations = QLabel("<b>📦 Vision par opération (regroupement de lots)</b>")
        label_operations.setStyleSheet("font-size: 11pt; color: #0078d4;")
        header_operations_layout.addWidget(label_operations)
        header_operations_layout.addStretch()

        operations_layout.addWidget(header_operations)

        # Filtre
        filtre_operation_widget = QWidget()
        filtre_operation_layout = QHBoxLayout(filtre_operation_widget)
        filtre_operation_layout.setContentsMargins(0, 5, 0, 5)

        label_filtre_operation = QLabel("🔍 Opération:")
        filtre_operation_layout.addWidget(label_filtre_operation)

        self.edit_filtre_operation = QLineEdit()
        self.edit_filtre_operation.setPlaceholderText("Filtrer par code opération...")
        self.edit_filtre_operation.setClearButtonEnabled(True)
        self.edit_filtre_operation.setMaximumWidth(200)
        self.edit_filtre_operation.textChanged.connect(
            lambda text: self.operations_proxy.setOperationFilter(text)
        )
        filtre_operation_layout.addWidget(self.edit_filtre_operation)

        # Bouton d'export suivi financier
        btn_export_operation = QPushButton("📊 Exporter suivi financier")
        btn_export_operation.setMaximumWidth(200)
        btn_export_operation.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
            QPushButton:pressed {
                background-color: #1e7e34;
            }
        """)
        btn_export_operation.clicked.connect(self.export_suivi_financier_operation)
        filtre_operation_layout.addWidget(btn_export_operation)

        # Bouton d'export spécifique 2020_14G3P
        btn_export_operation_2020 = QPushButton("📊 Export 2020_14G3P")
        btn_export_operation_2020.setMaximumWidth(200)
        btn_export_operation_2020.setStyleSheet("""
            QPushButton {
                background-color: #6f42c1;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a32a3;
            }
            QPushButton:pressed {
                background-color: #4e2b8a;
            }
        """)
        btn_export_operation_2020.clicked.connect(self.export_suivi_financier_2020_14G3P)
        filtre_operation_layout.addWidget(btn_export_operation_2020)

        filtre_operation_layout.addStretch()

        operations_layout.addWidget(filtre_operation_widget)

        # Table opérations
        self.table_operations = QTableView()
        self.table_operations.setModel(self.operations_proxy)
        self.table_operations.setSortingEnabled(True)
        self.table_operations.setSelectionBehavior(QTableView.SelectRows)
        self.table_operations.setSelectionMode(QTableView.SingleSelection)
        self.table_operations.setWordWrap(True)

        # Optimisation de la hauteur des lignes (3 lignes max)
        self.table_operations.verticalHeader().setDefaultSectionSize(60)  # ~3 lignes de texte

        # Connecter le double-clic pour afficher l'historique de l'opération
        self.table_operations.doubleClicked.connect(self.on_operation_double_clicked)

        header_operations_table = self.table_operations.horizontalHeader()
        header_operations_table.setDefaultAlignment(Qt.AlignCenter)
        header_operations_table.setStyleSheet(
            "QHeaderView::section { font-weight: bold; background-color: #fff3cd; padding: 4px; }"
        )

        # Mode de redimensionnement optimisé avec largeurs contrôlées
        header_operations_table.setSectionResizeMode(QHeaderView.Interactive)

        # Définir des largeurs initiales optimales pour chaque colonne
        # Colonnes: Opération, Nb lots, Marchés, Libellé, Fournisseur, Montant initial, Avenants, SF, Payé, Reste réaliser, Reste mandater, %
        self.table_operations.setColumnWidth(0, 120)  # Opération
        self.table_operations.setColumnWidth(1, 60)   # Nb lots
        self.table_operations.setColumnWidth(2, 250)  # Marchés (word wrap sur 3 lignes)
        self.table_operations.setColumnWidth(3, 300)  # Libellé (word wrap sur 3 lignes)
        self.table_operations.setColumnWidth(4, 200)  # Fournisseur (word wrap)
        self.table_operations.setColumnWidth(5, 120)  # Montant initial
        self.table_operations.setColumnWidth(6, 70)   # Avenants
        self.table_operations.setColumnWidth(7, 120)  # Service fait
        self.table_operations.setColumnWidth(8, 120)  # Payé
        self.table_operations.setColumnWidth(9, 120)  # Reste à réaliser
        self.table_operations.setColumnWidth(10, 120) # Reste à mandater
        self.table_operations.setColumnWidth(11, 80)  # % consommé

        # Permettre le redimensionnement manuel par l'utilisateur
        header_operations_table.setStretchLastSection(False)

        operations_layout.addWidget(self.table_operations)

        self.tabs.addTab(operations_widget, "📦 Opérations")

        # === ONGLET HISTORIQUE ===
        historique_widget = QWidget()
        historique_layout = QVBoxLayout(historique_widget)
        historique_layout.setContentsMargins(5, 5, 5, 5)
        historique_layout.setSpacing(10)

        # En-tête
        header_historique = QWidget()
        header_historique_layout = QHBoxLayout(header_historique)
        header_historique_layout.setContentsMargins(0, 0, 0, 0)

        label_historique = QLabel("<b>📜 Historique complet des factures et paiements</b>")
        label_historique.setStyleSheet("font-size: 11pt; color: #0078d4;")
        header_historique_layout.addWidget(label_historique)
        header_historique_layout.addStretch()

        historique_layout.addWidget(header_historique)

        # Filtre
        filtre_historique_widget = QWidget()
        filtre_historique_layout = QHBoxLayout(filtre_historique_widget)
        filtre_historique_layout.setContentsMargins(0, 5, 0, 5)

        label_filtre_historique = QLabel("🔍 Marché:")
        filtre_historique_layout.addWidget(label_filtre_historique)

        self.edit_filtre_historique = QLineEdit()
        self.edit_filtre_historique.setPlaceholderText("Filtrer par code marché...")
        self.edit_filtre_historique.setClearButtonEnabled(True)
        self.edit_filtre_historique.setMaximumWidth(200)
        self.edit_filtre_historique.textChanged.connect(
            lambda text: self.historique_proxy.setMarcheFilter(text)
        )
        filtre_historique_layout.addWidget(self.edit_filtre_historique)
        filtre_historique_layout.addStretch()

        historique_layout.addWidget(filtre_historique_widget)

        # Table historique
        self.table_historique = QTableView()
        self.table_historique.setModel(self.historique_proxy)
        self.table_historique.setSortingEnabled(True)
        self.table_historique.setSelectionBehavior(QTableView.SelectRows)
        self.table_historique.setSelectionMode(QTableView.ExtendedSelection)
        self.table_historique.setWordWrap(True)

        # Optimisation de la hauteur des lignes (3 lignes max)
        self.table_historique.verticalHeader().setDefaultSectionSize(60)  # ~3 lignes de texte

        header_historique_table = self.table_historique.horizontalHeader()
        header_historique_table.setDefaultAlignment(Qt.AlignCenter)
        header_historique_table.setStyleSheet(
            "QHeaderView::section { font-weight: bold; background-color: #d1ecf1; padding: 4px; }"
        )

        # Mode de redimensionnement optimisé avec largeurs contrôlées
        header_historique_table.setSectionResizeMode(QHeaderView.Interactive)

        # Définir des largeurs initiales optimales pour chaque colonne
        # Colonnes: Marché, Fournisseur, Date SF, N° Facture, Libellé, Montant SF, Montant TTC, N° Mandat, Statut
        self.table_historique.setColumnWidth(0, 120)  # Marché
        self.table_historique.setColumnWidth(1, 200)  # Fournisseur (word wrap)
        self.table_historique.setColumnWidth(2, 90)   # Date SF
        self.table_historique.setColumnWidth(3, 100)  # N° Facture
        self.table_historique.setColumnWidth(4, 350)  # Libellé (word wrap sur 3 lignes)
        self.table_historique.setColumnWidth(5, 110)  # Montant SF
        self.table_historique.setColumnWidth(6, 110)  # Montant TTC
        self.table_historique.setColumnWidth(7, 100)  # N° Mandat
        self.table_historique.setColumnWidth(8, 120)  # Statut

        # Permettre le redimensionnement manuel par l'utilisateur
        header_historique_table.setStretchLastSection(False)

        historique_layout.addWidget(self.table_historique)

        self.tabs.addTab(historique_widget, "📜 Historique")

        # Toolbar
        self.status_filter_combo = None
        self.facturation_filter_combo = None
        self.fournisseur_filter_combo = None
        self._init_toolbar()

        # Tray icon
        self._init_tray()

        # Timer rappels
        self.reminder_timer = QTimer(self)
        self.reminder_timer.setInterval(self._get_reminder_interval() * 60 * 1000)
        self.reminder_timer.timeout.connect(self.check_reminders)
        self.reminder_timer.start()

        # Message de bienvenue si première utilisation
        source_dir = self.db.get_config("excel_source_directory", "")
        if not source_dir:
            QMessageBox.information(
                self,
                "Configuration requise",
                "Bienvenue !\n\n"
                "Avant de commencer, configurez le dossier contenant vos fichiers Excel :\n"
                "- Menu 'Configuration...' dans la barre d'outils\n"
                "- Définissez le dossier source des fichiers Excel\n\n"
                "Ensuite, utilisez 'Import incrémental' pour importer les données."
            )

        # resize
        self.resize_all()
        
        # Mettre à jour le badge des rappels
        self._update_rappels_badge()

    def _init_toolbar(self):
        # ========== LIGNE 1 : BOUTONS D'ACTION MODERNISÉS ==========
        tb1 = QToolBar("Actions")
        tb1.setMovable(False)
        tb1.setIconSize(QIcon().actualSize(tb1.iconSize()) * 1.2)
        
        # Style moderne avec boutons colorés, padding généreux et ombres 3D
        tb1.setStyleSheet("""
            QToolBar {
                background-color: #ffffff;
                border-bottom: 2px solid #e0e0e0;
                padding: 6px;
                spacing: 4px;
            }
            QToolButton {
                background-color: #0078d4;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 10px;
                font-weight: bold;
                font-size: 8pt;
                min-height: 24px;
                min-width: 0px;
            }
            QToolButton:hover {
                background-color: #005a9e;
            }
            QToolButton:pressed {
                background-color: #004578;
            }
            QToolButton[objectName="selection_btn"] {
                background-color: #6c757d;
            }
            QToolButton[objectName="selection_btn"]:hover {
                background-color: #545b62;
            }
            QToolButton[objectName="selection_btn"]:pressed {
                background-color: #3d4349;
            }
            QToolButton[objectName="action_envoyee"] {
                background-color: #28a745;
            }
            QToolButton[objectName="action_envoyee"]:hover {
                background-color: #218838;
            }
            QToolButton[objectName="action_envoyee"]:pressed {
                background-color: #1e7e34;
            }
            QToolButton[objectName="action_suivre"] {
                background-color: #fd7e14;
            }
            QToolButton[objectName="action_suivre"]:hover {
                background-color: #e8590c;
            }
            QToolButton[objectName="action_suivre"]:pressed {
                background-color: #bd4b00;
            }
            QToolButton[objectName="action_rappel"] {
                background-color: #17a2b8;
            }
            QToolButton[objectName="action_rappel"]:hover {
                background-color: #138496;
            }
            QToolButton[objectName="action_rappel"]:pressed {
                background-color: #0f6674;
            }
            QToolButton[objectName="config_btn"] {
                background-color: #6c757d;
            }
            QToolButton[objectName="config_btn"]:hover {
                background-color: #545b62;
            }
            QToolButton[objectName="config_btn"]:pressed {
                background-color: #3d4349;
            }
            QToolButton[objectName="export_btn"] {
                background-color: #6610f2;
            }
            QToolButton[objectName="export_btn"]:hover {
                background-color: #5a0cd9;
            }
            QToolButton[objectName="export_btn"]:pressed {
                background-color: #4a09b3;
            }
            QToolButton[objectName="test_rappels_btn"] {
                background-color: #20c997;
            }
            QToolButton[objectName="test_rappels_btn"]:hover {
                background-color: #1ab386;
            }
            QToolButton[objectName="test_rappels_btn"]:pressed {
                background-color: #148f6a;
            }
            QToolButton::menu-indicator {
                image: none;
                width: 0px;
            }
        """)
        self.addToolBar(tb1)

        # Groupe 1: Import de données (BLEU)
        act_incremental = QAction("🔄 Import incrémental", self)
        act_incremental.setToolTip("Import incrémental multi-fichiers : scanne un dossier et importe seulement les nouveaux fichiers ou modifiés")
        act_incremental.triggered.connect(self.import_excel_files_incremental)
        tb1.addAction(act_incremental)

        act_conf = QAction("⚙ Configuration…", self)
        act_conf.setToolTip("Ouvrir la configuration des chemins de fichiers et des paramètres de rappel")
        act_conf.triggered.connect(self.open_config)
        tb1.addAction(act_conf)
        # Appliquer le style config au bouton
        for widget in tb1.children():
            if isinstance(widget, QWidget) and widget.property("text") == "⚙ Configuration…":
                widget.setObjectName("config_btn")

        # Séparateur élégant
        separator1 = tb1.addSeparator()
        
        # Groupe 2: Sélection (GRIS)
        act_sel_all = QAction("☑ Tout sélectionner", self)
        act_sel_all.setToolTip("Sélectionner toutes les lignes du tableau actif")
        act_sel_all.triggered.connect(self.select_all_current_tab)
        tb1.addAction(act_sel_all)

        act_sel_clear = QAction("☐ Tout désélectionner", self)
        act_sel_clear.setToolTip("Désélectionner toutes les lignes du tableau actif")
        act_sel_clear.triggered.connect(self.clear_selection_current_tab)
        tb1.addAction(act_sel_clear)

        # Séparateur élégant
        separator2 = tb1.addSeparator()

        # Groupe 3: Actions sur statut (COLORÉ)
        act_sent = QAction("✉ Marquer envoyée", self)
        act_sent.setToolTip("Marquer les commandes sélectionnées comme envoyées au fournisseur")
        act_sent.triggered.connect(self.mark_selected_sent)
        tb1.addAction(act_sent)

        act_follow = QAction("● Marquer à suivre", self)
        act_follow.setToolTip("Marquer les commandes sélectionnées comme nécessitant un suivi")
        act_follow.triggered.connect(self.mark_selected_follow)
        tb1.addAction(act_follow)

        act_resched = QAction("⏰ Reprogrammer rappel", self)
        act_resched.setToolTip("Reprogrammer le rappel pour les commandes sélectionnées")
        act_resched.triggered.connect(self.reschedule_selected)
        tb1.addAction(act_resched)
        
        act_test_rappels = QAction("🧪 Test Rappels", self)
        act_test_rappels.setToolTip("Tester le système de rappels (mode test sans modification)")
        act_test_rappels.triggered.connect(self.test_reminders_dialog)
        tb1.addAction(act_test_rappels)
        
        # Séparateur élégant
        separator3 = tb1.addSeparator()
        
        # Groupe 4: Exports (avec menu déroulant)
        export_menu = QMenu(self)
        export_menu.setStyleSheet("""
            QMenu {
                background-color: white;
                border: 1px solid #ccc;
            }
            QMenu::item {
                padding: 8px 30px 8px 20px;
            }
            QMenu::item:selected {
                background-color: #0078d4;
                color: white;
            }
        """)
        
        act_export_pdf = QAction("📄 Exporter en PDF", self)
        act_export_pdf.setToolTip("Exporter le tableau actif en PDF")
        act_export_pdf.triggered.connect(self.export_to_pdf)
        export_menu.addAction(act_export_pdf)
        
        act_export_excel = QAction("📊 Exporter en Excel", self)
        act_export_excel.setToolTip("Exporter le tableau actif en Excel")
        act_export_excel.triggered.connect(self.export_to_excel)
        export_menu.addAction(act_export_excel)
        
        act_export = QAction("📥 Exporter", self)
        act_export.setToolTip("Exporter le tableau actif en PDF ou Excel")
        act_export.setMenu(export_menu)
        tb1.addAction(act_export)
        
        # Configurer le bouton pour afficher le menu déroulant
        for widget in tb1.children():
            if hasattr(widget, 'defaultAction'):
                action = widget.defaultAction()
                if action and action.text() == "📥 Exporter":
                    from PyQt5.QtWidgets import QToolButton
                    widget.setPopupMode(QToolButton.InstantPopup)
                    widget.setObjectName("export_btn")
                    break
        
        # Appliquer les styles personnalisés aux boutons via objectName
        btn_index = 0
        for widget in tb1.children():
            if hasattr(widget, 'defaultAction'):
                action = widget.defaultAction()
                if action:
                    text = action.text()
                    if "sélectionner" in text or "désélectionner" in text:
                        widget.setObjectName("selection_btn")
                    elif "envoyée" in text:
                        widget.setObjectName("action_envoyee")
                    elif "suivre" in text:
                        widget.setObjectName("action_suivre")
                    elif "rappel" in text.lower() and "test" not in text.lower():
                        widget.setObjectName("action_rappel")
                    elif "Test Rappels" in text:
                        widget.setObjectName("test_rappels_btn")
                    elif "Configuration" in text:
                        widget.setObjectName("config_btn")
                    widget.setStyle(widget.style())  # Force refresh du style

        # ========== SAUT DE LIGNE ==========
        self.addToolBarBreak(Qt.TopToolBarArea)

        # ========== LIGNE 2 : FILTRES CLASSIQUES ==========
        tb2 = QToolBar("Filtres")
        tb2.setMovable(False)
        # Style moderne avec fond gris clair, coins arrondis et bordure (compact)
        tb2.setStyleSheet("""
            QToolBar {
                background-color: #f5f5f5;
                border: 1px solid #d0d0d0;
                border-radius: 6px;
                padding: 4px;
                spacing: 3px;
            }
            QLabel {
                font-weight: bold;
                padding: 0px 3px;
            }
            QComboBox {
                border: 1px solid #c0c0c0;
                border-radius: 4px;
                padding: 2px 5px;
                background-color: white;
                min-height: 20px;
            }
            QComboBox:hover {
                border: 1px solid #0078d4;
            }
            QComboBox::drop-down {
                border: none;
            }
        """)
        self.addToolBar(Qt.TopToolBarArea, tb2)
        
        # Label principal "Filtres:" en gras
        label_filtres = QLabel("<b>Filtres:</b>", self)
        tb2.addWidget(label_filtres)
        
        # Séparateur visuel
        sep1 = QLabel("||", self)
        sep1.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb2.addWidget(sep1)
        
        # Filtre 1 (Statut)
        label_statut = QLabel("<b>Statut:</b>", self)
        tb2.addWidget(label_statut)
        self.filter1_combo = QComboBox(self)
        self.filter1_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.filter1_combo.setMinimumContentsLength(10)
        self.filter1_combo.currentTextChanged.connect(self.on_filter1_changed)
        tb2.addWidget(self.filter1_combo)

        # Séparateur visuel
        sep2 = QLabel("||", self)
        sep2.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb2.addWidget(sep2)

        # Filtre 2 (Facturation/Exercice)
        self.filter2_label = QLabel("<b>Facturation:</b>", self)
        tb2.addWidget(self.filter2_label)
        self.filter2_combo = QComboBox(self)
        self.filter2_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.filter2_combo.setMinimumContentsLength(10)
        self.filter2_combo.currentTextChanged.connect(self.on_filter2_changed)
        tb2.addWidget(self.filter2_combo)

        # Séparateur visuel
        self.sep3 = QLabel("||", self)
        self.sep3.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb2.addWidget(self.sep3)

        # Filtre 3 (Fournisseur)
        self.label_fournisseur = QLabel("<b>Fournisseur:</b>", self)
        tb2.addWidget(self.label_fournisseur)
        self.fournisseur_filter_combo = QComboBox(self)
        self.fournisseur_filter_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.fournisseur_filter_combo.setMinimumContentsLength(12)
        self.fournisseur_filter_combo.currentTextChanged.connect(self.on_fournisseur_filter_changed)
        tb2.addWidget(self.fournisseur_filter_combo)

        # Séparateur visuel
        self.sep4 = QLabel("||", self)
        self.sep4.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb2.addWidget(self.sep4)

        # Filtre 4 (Marché)
        self.marche_label = QLabel("<b>Marché:</b>", self)
        tb2.addWidget(self.marche_label)
        self.marche_filter_combo = QComboBox(self)
        self.marche_filter_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.marche_filter_combo.setMinimumContentsLength(8)
        self.marche_filter_combo.currentTextChanged.connect(self.on_marche_filter_changed)
        tb2.addWidget(self.marche_filter_combo)

        # ========== FILTRES MULTIPLES (sur la même ligne) ==========
        
        # Séparateur visuel
        self.sep5 = QLabel("||", self)
        self.sep5.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb2.addWidget(self.sep5)
        
        # Filtre Article fonction (choix multiples) avec icône ☑
        self.article_fonction_label_base = "☑ Art. fonction"
        self.article_fonction_label = QLabel(f"<b>{self.article_fonction_label_base}:</b>", self)
        tb2.addWidget(self.article_fonction_label)
        self.article_fonction_filter = CheckableComboBox(self)
        self.article_fonction_filter.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.article_fonction_filter.setMinimumContentsLength(12)
        self.article_fonction_filter.currentIndexChanged.connect(self.on_article_fonction_changed)
        tb2.addWidget(self.article_fonction_filter)
        
        # Séparateur visuel
        self.sep6 = QLabel("||", self)
        self.sep6.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb2.addWidget(self.sep6)
        
        # Filtre Article nature (choix multiples) avec icône ☑
        self.article_nature_label_base = "☑ Art. nature"
        self.article_nature_label = QLabel(f"<b>{self.article_nature_label_base}:</b>", self)
        tb2.addWidget(self.article_nature_label)
        self.article_nature_filter = CheckableComboBox(self)
        self.article_nature_filter.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.article_nature_filter.setMinimumContentsLength(12)
        self.article_nature_filter.currentIndexChanged.connect(self.on_article_nature_changed)
        tb2.addWidget(self.article_nature_filter)
        
        # Séparateur visuel
        self.sep7 = QLabel("||", self)
        self.sep7.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb2.addWidget(self.sep7)
        
        # Filtre Service émetteur (choix multiples) avec icône ☑
        self.service_emetteur_label_base = "☑ Service"
        self.service_emetteur_label = QLabel(f"<b>{self.service_emetteur_label_base}:</b>", self)
        tb2.addWidget(self.service_emetteur_label)
        self.service_emetteur_filter = CheckableComboBox(self)
        self.service_emetteur_filter.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.service_emetteur_filter.setMinimumContentsLength(12)
        self.service_emetteur_filter.currentIndexChanged.connect(self.on_service_emetteur_changed)
        tb2.addWidget(self.service_emetteur_filter)
        
        # Séparateur visuel avant le bouton de réinitialisation
        self.sep8 = QLabel("||", self)
        self.sep8.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb2.addWidget(self.sep8)
        
        # Bouton pour réinitialiser les filtres multiples
        act_clear_multi = QAction("Réinitialiser filtres", self)
        act_clear_multi.triggered.connect(self.clear_multiple_filters)
        tb2.addAction(act_clear_multi)

        # Connecter le changement d'onglet pour adapter les filtres
        self.tabs.currentChanged.connect(self.on_tab_changed)

        # ========== LIGNE 3 : RECHERCHE PAR NUMÉRO ==========
        tb3 = QToolBar("Recherche")
        tb3.setMovable(False)
        tb3.setStyleSheet("""
            QToolBar {
                background-color: #e8f4f8;
                border: 1px solid #b0d0e0;
                border-radius: 6px;
                padding: 4px;
                spacing: 3px;
            }
            QLabel {
                font-weight: bold;
                padding: 0px 3px;
            }
            QLineEdit {
                border: 1px solid #c0c0c0;
                border-radius: 4px;
                padding: 2px 5px;
                background-color: white;
                min-height: 20px;
                min-width: 150px;
            }
            QLineEdit:hover {
                border: 1px solid #0078d4;
            }
            QLineEdit:focus {
                border: 2px solid #0078d4;
            }
        """)
        self.addToolBar(Qt.TopToolBarArea, tb3)

        # Label principal "Recherche:" en gras
        label_recherche = QLabel("<b>🔍 Recherche:</b>", self)
        tb3.addWidget(label_recherche)

        # Séparateur visuel
        sep_rech1 = QLabel("||", self)
        sep_rech1.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb3.addWidget(sep_rech1)

        # Champ de recherche N° Commande (visible pour Commandes et Factures)
        self.label_search_cmd = QLabel("<b>N° Commande:</b>", self)
        tb3.addWidget(self.label_search_cmd)
        self.search_num_commande = QLineEdit(self)
        self.search_num_commande.setPlaceholderText("Rechercher un n° de commande...")
        self.search_num_commande.setClearButtonEnabled(True)
        self.search_num_commande.textChanged.connect(self.on_search_num_commande_changed)
        tb3.addWidget(self.search_num_commande)

        # Séparateur visuel
        self.sep_rech2 = QLabel("||", self)
        self.sep_rech2.setStyleSheet("font-weight: normal; color: #a0a0a0; padding: 0px 2px;")
        tb3.addWidget(self.sep_rech2)

        # Champ de recherche N° Facture (visible seulement pour Factures)
        self.label_search_fact = QLabel("<b>N° Facture:</b>", self)
        tb3.addWidget(self.label_search_fact)
        self.search_num_facture = QLineEdit(self)
        self.search_num_facture.setPlaceholderText("Rechercher un n° de facture...")
        self.search_num_facture.setClearButtonEnabled(True)
        self.search_num_facture.textChanged.connect(self.on_search_num_facture_changed)
        tb3.addWidget(self.search_num_facture)

        # Espace flexible pour pousser les éléments à gauche
        tb3.addWidget(QWidget())

        # Sauvegarder la référence à la toolbar de recherche
        self.search_toolbar = tb3

        # Initialiser les filtres pour le premier onglet (APRÈS création des widgets)
        self.on_tab_changed(0)

    def _init_tray(self):
        if not QSystemTrayIcon.isSystemTrayAvailable():
            self.tray = None
            return
        self.tray = QSystemTrayIcon(self)
        self.tray.setIcon(self.style().standardIcon(QStyle.SP_FileDialogInfoView))

        menu = QMenu()
        act_show = QAction("Afficher la fenêtre", self)
        act_show.triggered.connect(self.showNormal)
        menu.addAction(act_show)
        act_quit = QAction("Quitter", self)
        act_quit.triggered.connect(QApplication.instance().quit)
        menu.addAction(act_quit)

        self.tray.setContextMenu(menu)
        self.tray.show()

    def log_error(self, context, message):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] {context}: {message}"
        self.error_log.append(line)

    def _save_error_log(self):
        """Sauvegarde immédiate des logs d'erreur."""
        if self.error_log:
            try:
                log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "suivi_log.txt")
                with open(log_path, "a", encoding="utf-8") as f:
                    for line in self.error_log:
                        f.write(line + "\n")
                self.error_log.clear()
            except Exception:
                pass

    def _get_reminder_interval(self):
        val = self.db.get_config("reminder_interval_minutes", "5")
        try:
            minutes = int(val)
        except ValueError:
            minutes = 5
        return max(1, minutes)

    # ------------ Gestion des onglets et filtres ------------

    def on_tab_changed(self, index):
        """Adapter les filtres selon l'onglet actif."""
        tab_name = self.tabs.tabText(index)
        
        # Réinitialiser les combos
        self.filter1_combo.blockSignals(True)
        self.filter2_combo.blockSignals(True)
        self.filter1_combo.clear()
        self.filter2_combo.clear()
        
        if "Commandes" in tab_name:
            # Filtre 1: Statut commande
            self.filter1_combo.addItems(["Tous", "A suivre", "Envoyée"])
            # Filtre 2: Statut facturation
            self.filter2_label.setText("<b>Facturation:</b>")
            self.filter2_label.setVisible(True)
            self.filter2_combo.setVisible(True)
            self.filter2_combo.addItems(["Tous", "Non facturée", "Partiellement facturée", "Totalement facturée"])
            # Filtre Marché: visible et avec liste déroulante
            self.marche_label.setVisible(True)
            self.marche_filter_combo.setVisible(True)
            # Filtres multiples: visibles uniquement pour Commandes
            self.sep5.setVisible(True)
            self.article_fonction_label.setVisible(True)
            self.article_fonction_filter.setVisible(True)
            self.sep6.setVisible(True)
            self.article_nature_label.setVisible(True)
            self.article_nature_filter.setVisible(True)
            self.sep7.setVisible(True)
            self.service_emetteur_label.setVisible(True)
            self.service_emetteur_filter.setVisible(True)
            self.sep8.setVisible(True)
            # Recherche: N° Commande visible, N° Facture caché
            self.label_search_cmd.setVisible(True)
            self.search_num_commande.setVisible(True)
            self.sep_rech2.setVisible(False)
            self.label_search_fact.setVisible(False)
            self.search_num_facture.setVisible(False)
            
        elif "Factures" in tab_name and "Facturation" not in tab_name:
            # Filtre 1: Statut facture
            self.filter1_combo.addItems(["Tous", "Facturée", "Service fait", "En attente de paiement", "A vérifier"])
            # Filtre 2: Exercice
            self.filter2_label.setText("<b>Exercice:</b>")
            self.filter2_label.setVisible(True)
            self.filter2_combo.setVisible(True)
            exercices = self._get_exercices_factures()
            self.filter2_combo.addItem("Tous")
            self.filter2_combo.addItems(exercices)
            # Filtre Marché: VISIBLE pour Factures
            self.marche_label.setVisible(True)
            self.marche_filter_combo.setVisible(True)
            # Filtres multiples: cachés pour Factures
            self.sep5.setVisible(False)
            self.article_fonction_label.setVisible(False)
            self.article_fonction_filter.setVisible(False)
            self.sep6.setVisible(False)
            self.article_nature_label.setVisible(False)
            self.article_nature_filter.setVisible(False)
            self.sep7.setVisible(False)
            self.service_emetteur_label.setVisible(False)
            self.service_emetteur_filter.setVisible(False)
            self.sep8.setVisible(False)
            # Recherche: N° Commande et N° Facture visibles
            self.label_search_cmd.setVisible(True)
            self.search_num_commande.setVisible(True)
            self.sep_rech2.setVisible(True)
            self.label_search_fact.setVisible(True)
            self.search_num_facture.setVisible(True)
            
        elif "Facturation" in tab_name:
            # Filtre 1: Statut facturation
            self.filter1_combo.addItems(["Tous", "Non facturée", "Partiellement facturée", "Totalement facturée"])
            # Filtre 2: caché (pas de second filtre pour Facturation)
            self.filter2_label.setVisible(False)
            self.filter2_combo.setVisible(False)
            # Filtre Marché: visible et avec liste déroulante
            self.marche_label.setVisible(True)
            self.marche_filter_combo.setVisible(True)
            # Filtres multiples: cachés pour Facturation
            self.sep5.setVisible(False)
            self.article_fonction_label.setVisible(False)
            self.article_fonction_filter.setVisible(False)
            self.sep6.setVisible(False)
            self.article_nature_label.setVisible(False)
            self.article_nature_filter.setVisible(False)
            self.sep7.setVisible(False)
            self.service_emetteur_label.setVisible(False)
            self.service_emetteur_filter.setVisible(False)
            self.sep8.setVisible(False)
            # Recherche: tous cachés pour Facturation
            self.label_search_cmd.setVisible(False)
            self.search_num_commande.setVisible(False)
            self.sep_rech2.setVisible(False)
            self.label_search_fact.setVisible(False)
            self.search_num_facture.setVisible(False)

        elif "Rappels" in tab_name:
            # Pas de filtres spécifiques pour rappels
            self.filter2_label.setVisible(False)
            self.filter2_combo.setVisible(False)
            self.filter1_combo.addItem("Tous")
            # Filtre Marché: caché pour Rappels
            self.marche_label.setVisible(False)
            self.marche_filter_combo.setVisible(False)
            # Filtres multiples: cachés pour Rappels
            self.sep5.setVisible(False)
            self.article_fonction_label.setVisible(False)
            self.article_fonction_filter.setVisible(False)
            self.sep6.setVisible(False)
            self.article_nature_label.setVisible(False)
            self.article_nature_filter.setVisible(False)
            self.sep7.setVisible(False)
            self.service_emetteur_label.setVisible(False)
            self.service_emetteur_filter.setVisible(False)
            self.sep8.setVisible(False)
            # Recherche: tous cachés pour Rappels
            self.label_search_cmd.setVisible(False)
            self.search_num_commande.setVisible(False)
            self.sep_rech2.setVisible(False)
            self.label_search_fact.setVisible(False)
            self.search_num_facture.setVisible(False)

        self.filter1_combo.blockSignals(False)
        self.filter2_combo.blockSignals(False)
        
        # Rafraîchir la liste des fournisseurs
        self.refresh_fournisseur_filter()
        
        # Rafraîchir la liste des marchés (pour Commandes et Facturation)
        self.refresh_marche_filter()
        
        # Rafraîchir les filtres multiples (uniquement pour Commandes)
        self.refresh_multiple_filters()

    def _get_exercices_factures(self):
        """Récupère la liste des exercices dans les factures."""
        cur = self.db.conn.cursor()
        cur.execute("SELECT DISTINCT exercice FROM factures WHERE exercice IS NOT NULL AND exercice != '' ORDER BY exercice DESC")
        return [r[0] for r in cur.fetchall()]

    def _get_marches(self):
        """Récupère la liste des marchés selon l'onglet actif."""
        cur = self.db.conn.cursor()
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        marches_set = set()
        
        # Pour l'onglet Factures, lire depuis les factures
        if "Factures" in tab_name and "Facturation" not in tab_name:
            cur.execute("SELECT DISTINCT marche FROM factures WHERE marche IS NOT NULL AND marche != '' ORDER BY marche")
            marches_set.update([r[0] for r in cur.fetchall()])
        # Pour Commandes, lire depuis les commandes
        elif "Commandes" in tab_name:
            cur.execute("SELECT DISTINCT marche FROM commandes WHERE marche IS NOT NULL AND marche != '' ORDER BY marche")
            marches_set.update([r[0] for r in cur.fetchall()])
        # Pour Facturation, lire depuis les DEUX tables (commandes ET factures)
        elif "Facturation" in tab_name:
            # Marchés depuis commandes
            cur.execute("SELECT DISTINCT marche FROM commandes WHERE marche IS NOT NULL AND marche != '' ORDER BY marche")
            marches_set.update([r[0] for r in cur.fetchall()])
            # Marchés depuis factures
            cur.execute("SELECT DISTINCT marche FROM factures WHERE marche IS NOT NULL AND marche != '' ORDER BY marche")
            marches_set.update([r[0] for r in cur.fetchall()])
        
        return sorted(list(marches_set))

    def on_filter1_changed(self, text):
        """Filtre 1 change selon l'onglet."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        if "Commandes" in tab_name:
            self.cmd_proxy.setStatusFilter(text)
        elif "Factures" in tab_name and "Facturation" not in tab_name:
            self.fact_proxy.setStatutFilter(text)
        elif "Facturation" in tab_name:
            self.synth_proxy.setStatutFilter(text)

    def on_filter2_changed(self, text):
        """Filtre 2 change selon l'onglet."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        if "Commandes" in tab_name:
            self.cmd_proxy.setFacturationFilter(text)
        elif "Factures" in tab_name and "Facturation" not in tab_name:
            self.fact_proxy.setExerciceFilter(text)

    def on_fournisseur_filter_changed(self, text):
        """Filtre fournisseur - fonctionne sur tous les onglets avec modèles."""
        # Si "Tous" est sélectionné, on passe une chaîne vide pour ne pas filtrer
        filter_text = "" if text == "Tous" else text
        
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        if "Commandes" in tab_name:
            self.cmd_proxy.setFournisseurFilter(filter_text)
        elif "Factures" in tab_name and "Facturation" not in tab_name:
            self.fact_proxy.setFournisseurFilter(filter_text)
        elif "Facturation" in tab_name:
            self.synth_proxy.setFournisseurFilter(filter_text)

    def on_marche_filter_changed(self, text):
        """Filtre marché - fonctionne sur Commandes, Factures et Facturation."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        if "Commandes" in tab_name:
            self.cmd_proxy.setMarcheFilter(text)
        elif "Factures" in tab_name and "Facturation" not in tab_name:
            self.fact_proxy.setMarcheFilter(text)
        elif "Facturation" in tab_name:
            self.synth_proxy.setMarcheFilter(text)
    
    def on_article_fonction_changed(self, index):
        """Filtre article fonction (choix multiples) - Commandes uniquement."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        if "Commandes" in tab_name:
            checked = self.article_fonction_filter.checked_items()
            self.cmd_proxy.setArticleFonctionFilter(checked)
        self.update_multi_filter_labels()

    
    def on_article_nature_changed(self, index):
        """Filtre article nature (choix multiples) - Commandes uniquement."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        if "Commandes" in tab_name:
            checked = self.article_nature_filter.checked_items()
            self.cmd_proxy.setArticleNatureFilter(checked)
        self.update_multi_filter_labels()

    
    def on_service_emetteur_changed(self, index):
        """Filtre service émetteur (choix multiples) - Commandes uniquement."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        if "Commandes" in tab_name:
            checked = self.service_emetteur_filter.checked_items()
            self.cmd_proxy.setServiceEmetteurFilter(checked)
        self.update_multi_filter_labels()

    
    def update_multi_filter_labels(self):
        """Met à jour les libellés des filtres multiples selon qu'ils sont actifs ou non."""
        # Les filtres peuvent ne pas encore être initialisés au moment de l'appel.
        try:
            af_checked = self.article_fonction_filter.checked_items()
            an_checked = self.article_nature_filter.checked_items()
            se_checked = self.service_emetteur_filter.checked_items()
        except AttributeError:
            return

        def apply(label, base_text, is_active):
            if is_active:
                # Texte avec mention "(actif)" et couleur mise en avant
                label.setText(f"<b>{base_text} (actif):</b>")
                label.setStyleSheet("color: #0078d4;")
            else:
                # Texte et style par défaut
                label.setText(f"<b>{base_text}:</b>")
                label.setStyleSheet("")

        apply(self.article_fonction_label, self.article_fonction_label_base, bool(af_checked))
        apply(self.article_nature_label, self.article_nature_label_base, bool(an_checked))
        apply(self.service_emetteur_label, self.service_emetteur_label_base, bool(se_checked))

    def clear_multiple_filters(self):
        """Réinitialise tous les filtres à choix multiples."""
        self.article_fonction_filter.clear_selection()
        self.article_nature_filter.clear_selection()
        self.service_emetteur_filter.clear_selection()
        
        # Réappliquer les filtres vides
        self.cmd_proxy.setArticleFonctionFilter([])
        self.cmd_proxy.setArticleNatureFilter([])
        self.cmd_proxy.setServiceEmetteurFilter([])
        self.update_multi_filter_labels()

    def on_search_num_commande_changed(self, text):
        """Filtre par numéro de commande - fonctionne sur Commandes et Factures."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)

        if "Commandes" in tab_name:
            self.cmd_proxy.setNumCommandeFilter(text)
        elif "Factures" in tab_name and "Facturation" not in tab_name:
            self.fact_proxy.setNumCommandeFilter(text)

    def on_search_num_facture_changed(self, text):
        """Filtre par numéro de facture - fonctionne sur Factures uniquement."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)

        if "Factures" in tab_name and "Facturation" not in tab_name:
            self.fact_proxy.setNumFactureFilter(text)

    def select_all_current_tab(self):
        """Sélectionner tout dans l'onglet actif."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        if tab_name == "Commandes":
            self.table_cmd.selectAll()
        elif tab_name == "Factures":
            self.table_fact.selectAll()
        elif tab_name == "Facturation":
            self.table_synth.selectAll()

    def clear_selection_current_tab(self):
        """Désélectionner tout dans l'onglet actif."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        if tab_name == "Commandes":
            self.table_cmd.clearSelection()
        elif tab_name == "Factures":
            self.table_fact.clearSelection()
        elif tab_name == "Facturation":
            self.table_synth.clearSelection()

    # ------------ Import incrémental multi-fichiers ------------

    def import_excel_files_incremental(self):
        """
        Import incrémental : scanne un dossier et importe seulement les fichiers
        nouveaux ou modifiés.
        """
        # Récupérer le dossier source depuis la config
        source_dir = self.db.get_config("excel_source_directory", "")

        if not source_dir or not os.path.isdir(source_dir):
            # Demander à l'utilisateur de choisir le dossier
            source_dir = QFileDialog.getExistingDirectory(
                self,
                "Choisir le dossier contenant les fichiers Excel",
                "",
                QFileDialog.ShowDirsOnly
            )
            if not source_dir:
                return

            # Sauvegarder dans la config
            self.db.set_config("excel_source_directory", source_dir)

        # Scanner le dossier
        files = self.db.scan_excel_directory(source_dir)

        total_files = len(files["commandes"]) + len(files["factures"])
        if total_files == 0:
            QMessageBox.information(
                self,
                "Aucun fichier trouvé",
                f"Aucun fichier Excel trouvé dans :\n{source_dir}\n\n"
                "Les fichiers doivent contenir 'commande' ou 'facture' dans leur nom."
            )
            return

        # Vérifier quels fichiers doivent être importés
        to_import = {"commandes": [], "factures": []}
        skipped = {"commandes": [], "factures": []}

        for file_type in ["commandes", "factures"]:
            for filepath in files[file_type]:
                should_import, reason = self.db.should_import_file(filepath)
                if should_import:
                    to_import[file_type].append((filepath, reason))
                else:
                    skipped[file_type].append((filepath, reason))

        nb_to_import = len(to_import["commandes"]) + len(to_import["factures"])
        nb_skipped = len(skipped["commandes"]) + len(skipped["factures"])

        # Afficher un résumé et demander confirmation
        summary = f"📁 Dossier : {source_dir}\n\n"
        summary += f"📊 Analyse :\n"
        summary += f"  • {total_files} fichiers Excel trouvés\n"
        summary += f"  • {nb_to_import} à importer (nouveaux/modifiés)\n"
        summary += f"  • {nb_skipped} déjà à jour\n\n"

        if nb_to_import > 0:
            summary += "📥 Fichiers à importer :\n"
            for filepath, reason in to_import["commandes"]:
                summary += f"  • {os.path.basename(filepath)} ({reason})\n"
            for filepath, reason in to_import["factures"]:
                summary += f"  • {os.path.basename(filepath)} ({reason})\n"

        if nb_to_import == 0:
            QMessageBox.information(self, "Aucune mise à jour nécessaire", summary)
            return

        reply = QMessageBox.question(
            self,
            "Import incrémental",
            summary + "\nContinuer l'import ?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        # Importer les fichiers
        results = []
        errors = []

        for filepath, reason in to_import["commandes"]:
            try:
                count = self.import_commandes_from_file(filepath)
                results.append(f"✅ {os.path.basename(filepath)}: {count} commandes")
            except Exception as e:
                errors.append(f"❌ {os.path.basename(filepath)}: {str(e)}")

        for filepath, reason in to_import["factures"]:
            try:
                count = self.import_factures_from_file(filepath)
                results.append(f"✅ {os.path.basename(filepath)}: {count} factures")
            except Exception as e:
                errors.append(f"❌ {os.path.basename(filepath)}: {str(e)}")

        # Recalculer la facturation
        self.db.recompute_facturation()

        # Rafraîchir les vues
        self.cmd_model.refresh()
        self.fact_model.refresh()
        self.synth_model.refresh()
        self.refresh_fournisseur_filter()
        self.refresh_rappels_tab()
        self.resize_all()

        # Afficher le résultat
        result_msg = "Import incrémental terminé !\n\n"
        result_msg += "\n".join(results)
        if errors:
            result_msg += "\n\nErreurs :\n" + "\n".join(errors)

        if errors:
            QMessageBox.warning(self, "Import terminé avec erreurs", result_msg)
        else:
            QMessageBox.information(self, "Import réussi", result_msg)

    def import_commandes_from_file(self, filepath: str) -> int:
        """
        Importe les commandes d'un fichier spécifique.
        Retourne le nombre de commandes importées.
        """
        import os

        filename = os.path.basename(filepath)
        file_hash = self.db.calculate_file_hash(filepath)
        file_size = os.path.getsize(filepath)

        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            self.db.record_import(filename, filepath, file_hash, file_size,
                                 "commandes", 0, "error", str(e))
            raise

        expected = [
            "Exercice", "N° Commande", "Fournisseur", "Libellé",
            "Date de la commande", "Marché", "Service émetteur", "Montant TTC",
            "Section", "Article par fonction", "Article par nature"
        ]
        missing = [c for c in expected if c not in df.columns]
        if missing:
            error_msg = "Colonnes manquantes: " + ", ".join(missing)
            self.db.record_import(filename, filepath, file_hash, file_size,
                                 "commandes", 0, "error", error_msg)
            raise ValueError(error_msg)

        # Grouper par (Exercice, N° Commande)
        grouped = df.groupby(["Exercice", "N° Commande"], as_index=False).agg({
            "Fournisseur": "first",
            "Libellé": "first",
            "Date de la commande": "first",
            "Marché": "first",
            "Service émetteur": "first",
            "Montant TTC": "sum",
            "Section": "first",
            "Article par fonction": "first",
            "Article par nature": "first"
        })

        for _, row in grouped.iterrows():
            data = {
                "exercice": str(row["Exercice"]),
                "num_commande": str(row["N° Commande"]),
                "fournisseur": "" if pd.isna(row["Fournisseur"]) else str(row["Fournisseur"]),
                "libelle": "" if pd.isna(row["Libellé"]) else str(row["Libellé"]),
                "date_commande": parse_date_safe(row["Date de la commande"]),
                "marche": "" if pd.isna(row["Marché"]) else str(row["Marché"]),
                "service_emetteur": "" if pd.isna(row["Service émetteur"]) else str(row["Service émetteur"]),
                "montant_ttc": None if pd.isna(row["Montant TTC"]) else float(row["Montant TTC"]),
                "section": "" if pd.isna(row["Section"]) else str(row["Section"]),
                "article_fonction": "" if pd.isna(row["Article par fonction"]) else str(row["Article par fonction"]),
                "article_nature": "" if pd.isna(row["Article par nature"]) else str(row["Article par nature"]),
                "source_file": filename
            }
            self.db.upsert_commande(data)

        # Enregistrer l'import
        count = len(grouped)
        self.db.record_import(filename, filepath, file_hash, file_size,
                             "commandes", count, "success")

        return count

    def import_factures_from_file(self, filepath: str) -> int:
        """
        Importe les factures d'un fichier spécifique.
        Retourne le nombre de factures importées.
        """
        import os
        from datetime import datetime

        filename = os.path.basename(filepath)
        file_hash = self.db.calculate_file_hash(filepath)
        file_size = os.path.getsize(filepath)

        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            self.db.record_import(filename, filepath, file_hash, file_size,
                                 "factures", 0, "error", str(e))
            raise

        # DEBUG: Afficher toutes les colonnes trouvées
        print(f"[DEBUG] Colonnes trouvées dans {filename}:")
        print(f"  Total: {len(df.columns)} colonnes")
        for i, col in enumerate(df.columns):
            print(f"  [{i}] '{col}'")

        required_cols = {
            "Exercice", "Code mouvement", "Nom tiers",
            "Libellé mouvement", "Date service fait", "Montant service fait"
        }
        missing = required_cols - set(df.columns)
        if missing:
            error_msg = "Colonnes manquantes: " + ", ".join(missing)
            self.db.record_import(filename, filepath, file_hash, file_size,
                                 "factures", 0, "error", error_msg)
            raise ValueError(error_msg)

        # Trouver la colonne N° de facture (optionnelle)
        col_facture = None
        for col in ["N° pièce", "Facture", "N° facture"]:
            if col in df.columns:
                col_facture = col
                print(f"[DEBUG] Colonne facture trouvée: '{col}'")
                break

        # Trouver la colonne Montant TTC
        col_montant_ttc = None
        for col in ["Montant TTC", "Montant TTC mouvement"]:
            if col in df.columns:
                col_montant_ttc = col
                print(f"[DEBUG] Colonne montant TTC trouvée: '{col}'")
                break

        # Trouver la colonne Marché
        col_marche = None
        for col in ["Marché", "March\u00e9"]:
            if col in df.columns:
                col_marche = col
                print(f"[DEBUG] Colonne marché trouvée: '{col}'")
                break

        if col_marche is None:
            print(f"[WARNING] Aucune colonne Marché trouvée dans {filename}")

        # Trouver les colonnes additionnelles pour l'analyse des marchés
        col_tranche = "Tranche" if "Tranche" in df.columns else None
        col_mandat = "Mandat" if "Mandat" in df.columns else None
        col_montant_initial = "Montant initial" if "Montant initial" in df.columns else None

        now = datetime.now().isoformat(timespec="seconds")
        cur = self.db.conn.cursor()
        count = 0

        for _, row in df.iterrows():
            exercice = "" if pd.isna(row["Exercice"]) else str(int(row["Exercice"]))
            num_facture = "" if col_facture is None or pd.isna(row[col_facture]) else str(row[col_facture])
            code_mouvement = "" if pd.isna(row["Code mouvement"]) else str(row["Code mouvement"])
            fournisseur = "" if pd.isna(row["Nom tiers"]) else str(row["Nom tiers"])
            libelle = "" if pd.isna(row["Libellé mouvement"]) else str(row["Libellé mouvement"])
            date_facture = parse_date_safe(row["Date service fait"])
            montant_ttc = None if col_montant_ttc is None or pd.isna(row[col_montant_ttc]) else float(row[col_montant_ttc])
            montant_service_fait = 0.0 if pd.isna(row["Montant service fait"]) else float(row["Montant service fait"])
            marche = "" if col_marche is None or pd.isna(row[col_marche]) else str(row[col_marche])

            # Nouvelles colonnes pour l'analyse des marchés
            tranche = "" if col_tranche is None or pd.isna(row[col_tranche]) else str(row[col_tranche])
            # IMPORTANT: Le N° de commande est dans "Code mouvement", pas dans "Commande" (qui est vide)
            commande = code_mouvement
            num_mandat = "" if col_mandat is None or pd.isna(row[col_mandat]) else str(row[col_mandat])
            montant_initial = None if col_montant_initial is None or pd.isna(row[col_montant_initial]) else float(row[col_montant_initial])

            statut = compute_facture_status(num_facture, montant_service_fait, date_facture)

            cur.execute(
                """
                INSERT INTO factures (
                    exercice, num_facture, code_mouvement,
                    fournisseur, libelle, date_facture,
                    montant_ttc, montant_service_fait, marche,
                    statut_facture,
                    rappel_facture_actif, frequence_rappel_facture_jours,
                    prochaine_date_rappel_facture, notes, last_update, source_file,
                    tranche, commande, num_mandat, montant_initial
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    exercice, num_facture, code_mouvement,
                    fournisseur, libelle, date_facture,
                    montant_ttc, montant_service_fait, marche,
                    statut,
                    0, None, None, "", now, filename,
                    tranche, commande, num_mandat, montant_initial
                )
            )
            count += 1

        self.db.conn.commit()

        # Enregistrer l'import
        self.db.record_import(filename, filepath, file_hash, file_size,
                             "factures", count, "success")

        return count

    # ------------ Actions commandes ------------

    def selected_cmd_ids(self):
        idxs = self.table_cmd.selectionModel().selectedRows()
        if not idxs:
            return []
        ids = []
        for proxy_index in idxs:
            src_row = self.cmd_proxy.mapToSource(proxy_index).row()
            cid = self.cmd_model.get_row_id(src_row)
            if cid is not None:
                ids.append(cid)
        return ids

    def mark_selected_sent(self):
        ids = self.selected_cmd_ids()
        if not ids:
            QMessageBox.information(self, "Sélection vide", "Aucune ligne sélectionnée.")
            return
        self.db.update_statut_for_ids(ids, "Envoyée", disable_rappel=True)
        self.cmd_model.refresh()
        self.synth_model.refresh()
        self.refresh_rappels_tab()
        self.resize_all()

    def mark_selected_follow(self):
        ids = self.selected_cmd_ids()
        if not ids:
            QMessageBox.information(self, "Sélection vide", "Aucune ligne sélectionnée.")
            return
        # Changer le statut vers "A suivre"
        self.db.update_statut_for_ids(ids, "A suivre", disable_rappel=False)
        # NOUVEAU : Activer automatiquement le rappel
        self.db.reschedule_rappel_for_ids(ids)
        self.cmd_model.refresh()
        self.synth_model.refresh()
        self.refresh_rappels_tab()
        self.resize_all()

    def reschedule_selected(self):
        ids = self.selected_cmd_ids()
        if not ids:
            QMessageBox.information(self, "Sélection vide", "Aucune ligne sélectionnée.")
            return
        self.db.reschedule_rappel_for_ids(ids)
        self.cmd_model.refresh()
        self.refresh_rappels_tab()
        self.resize_all()

    def on_cmd_double_clicked(self, index):
        src_row = self.cmd_proxy.mapToSource(index).row()
        cid = self.cmd_model.get_row_id(src_row)
        if cid is None:
            return
        reply = QMessageBox.question(
            self,
            "Confirmer",
            "Marquer cette commande comme envoyée au fournisseur ?\n\n"
            "Cette action désactivera les rappels pour cette commande.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.db.update_statut_for_ids([cid], "Envoyée", disable_rappel=True)
        self.cmd_model.refresh()
        self.synth_model.refresh()
        self.refresh_rappels_tab()
        self.resize_all()

    # ------------ Rappels ------------

    def check_reminders(self):
        due = self.db.due_reminders()
        if not due:
            return

        # Debug minimal en console pour vérification
        try:
            print("[DEBUG] due_reminders:", [(row["id"], row["num_commande"], row["prochaine_date_rappel"]) for row in due])
        except Exception:
            pass

        lines = []
        for row in due:
            dc = row["date_commande"]
            try:
                dc_fmt = datetime.strptime(dc, "%Y-%m-%d").strftime("%d/%m/%Y") if dc else ""
            except Exception:
                dc_fmt = dc or ""
            lines.append(f"{row['num_commande']} - {row['fournisseur']} ({dc_fmt})")

        text = "\n".join(lines[:10])
        if len(lines) > 10:
            text += f"\n… et {len(lines) - 10} autres."

        title = f"Rappels commandes ({len(due)})"

        # Popup visuel
        if self.tray is not None and self.tray.isVisible():
            self.tray.showMessage(title, text, QSystemTrayIcon.Information, 15000)
        else:
            QMessageBox.information(self, title, text)

        # Envoi mail éventuel
        self._send_email_reminders(title, text, due)

        # Replanifier tous les rappels traités
        for row in due:
            self.db.reschedule_rappel_for_ids([row["id"]])
        self.cmd_model.refresh()
        self.refresh_rappels_tab()
        self.resize_all()


    def _send_email_reminders(self, subject, body, due_rows):
        """Envoie le rappel par mail via Outlook (si disponible et activé).

        Implémentation:
        - Lit 'email_reminders_enabled' et 'email_reminders_to' dans la config.
        - Utilise COM Outlook si possible (environnement Windows avec Outlook installé).
        - En cas d'échec (librairie manquante, Outlook absent, etc.), ne bloque pas l'application.
        """
        try:
            enabled = self.db.get_config("email_reminders_enabled", "0") == "1"
            to_addr = (self.db.get_config("email_reminders_to", "") or "").strip()
        except Exception:
            return

        if not enabled or not to_addr:
            return

        # Tentative via Outlook COM (environnement Windows classique avec Office 365)
        try:
            import win32com.client  # type: ignore

            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To = to_addr
            mail.Subject = subject
            mail.Body = body
            mail.Send()
        except Exception as e:
            # En cas d'erreur, on loggue seulement en console sans interrompre le flux.
            try:
                print("[WARN] Echec envoi rappel mail:", e)
            except Exception:
                pass
    def test_reminders_dialog(self):
        """
        Fenêtre de test du système de rappels.
        Affiche tous les rappels actifs et permet de simuler un popup sans modifier les dates.
        """
        # Récupérer TOUS les rappels actifs (même futurs)
        all_rappels = self.db.all_active_reminders()
        
        if not all_rappels:
            QMessageBox.information(
                self, 
                "🧪 Test Rappels", 
                "Aucun rappel actif trouvé.\n\n"
                "Pour tester les rappels :\n"
                "1. Assurez-vous d'avoir des commandes avec statut 'À suivre' ou 'Envoyée'\n"
                "2. Vérifiez que le rappel est activé (case cochée)\n"
                "3. Une date de prochain rappel doit être définie"
            )
            return
        
        # Récupérer aussi les rappels dus AUJOURD'HUI
        due_rappels = self.db.due_reminders()
        
        # Créer la fenêtre de dialogue
        dialog = QDialog(self)
        dialog.setWindowTitle("🧪 Test du système de rappels")
        dialog.setMinimumWidth(900)
        dialog.setMinimumHeight(500)
        
        layout = QVBoxLayout()
        
        # === EN-TÊTE INFORMATIF ===
        info_label = QLabel(
            f"<b>📊 Statistiques des rappels :</b><br>"
            f"• Rappels actifs totaux : <b>{len(all_rappels)}</b><br>"
            f"• Rappels dus aujourd'hui : <b style='color: #dc3545;'>{len(due_rappels)}</b><br>"
            f"• Intervalle de vérification : <b>{self._get_reminder_interval()} minutes</b><br><br>"
            f"<i>Mode test : aucune modification ne sera apportée aux dates de rappel</i>"
        )
        info_label.setStyleSheet("padding: 10px; background-color: #e7f3ff; border-radius: 5px;")
        layout.addWidget(info_label)
        
        # === TABLEAU DES RAPPELS ===
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels([
            "N° Commande", 
            "Fournisseur", 
            "Date commande",
            "Prochain rappel",
            "Statut",
            "État"
        ])
        table.setRowCount(len(all_rappels))
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setAlternatingRowColors(True)
        
        due_ids = {r["id"] for r in due_rappels}
        today = datetime.now().date()
        
        for i, row in enumerate(all_rappels):
            # N° Commande
            table.setItem(i, 0, QTableWidgetItem(row["num_commande"] or ""))
            
            # Fournisseur
            table.setItem(i, 1, QTableWidgetItem(row["fournisseur"] or ""))
            
            # Date commande
            dc = row["date_commande"]
            try:
                dc_fmt = datetime.strptime(dc, "%Y-%m-%d").strftime("%d/%m/%Y") if dc else ""
            except Exception:
                dc_fmt = dc or ""
            table.setItem(i, 2, QTableWidgetItem(dc_fmt))
            
            # Prochain rappel
            pr = row["prochaine_date_rappel"]
            try:
                pr_date = datetime.strptime(pr, "%Y-%m-%d").date() if pr else None
                pr_fmt = pr_date.strftime("%d/%m/%Y") if pr_date else ""
            except Exception:
                pr_date = None
                pr_fmt = pr or ""
            table.setItem(i, 3, QTableWidgetItem(pr_fmt))
            
            # Statut
            table.setItem(i, 4, QTableWidgetItem(row["statut"] or ""))
            
            # État (Dû / Futur)
            is_due = row["id"] in due_ids
            etat_item = QTableWidgetItem("🔴 DÛ" if is_due else "🟢 Futur")
            if is_due:
                etat_item.setBackground(QColor("#ffe6e6"))
                font = etat_item.font()
                font.setBold(True)
                etat_item.setFont(font)
            table.setItem(i, 5, etat_item)
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        # === BOUTONS D'ACTION ===
        btn_layout = QHBoxLayout()
        
        # Bouton : Simuler popup (rappels dus)
        btn_simulate_due = QPushButton("🔔 Simuler popup (rappels dus)")
        btn_simulate_due.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 10pt;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        btn_simulate_due.setEnabled(len(due_rappels) > 0)
        btn_simulate_due.clicked.connect(lambda: self._simulate_popup(due_rappels))
        btn_layout.addWidget(btn_simulate_due)
        
        # Bouton : Simuler popup (tous les rappels)
        btn_simulate_all = QPushButton("🔔 Simuler popup (tous)")
        btn_simulate_all.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 10pt;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        btn_simulate_all.clicked.connect(lambda: self._simulate_popup(all_rappels))
        btn_layout.addWidget(btn_simulate_all)
        
        btn_layout.addStretch()
        
        # Bouton : Fermer
        btn_close = QPushButton("Fermer")
        btn_close.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 10pt;
            }
            QPushButton:hover {
                background-color: #545b62;
            }
        """)
        btn_close.clicked.connect(dialog.close)
        btn_layout.addWidget(btn_close)
        
        layout.addLayout(btn_layout)
        
        dialog.setLayout(layout)
        dialog.exec_()
    
    def _simulate_popup(self, rappels):
        """
        Simule l'affichage d'un popup de rappel sans modifier les dates.
        """
        if not rappels:
            QMessageBox.information(self, "Test Rappels", "Aucun rappel à afficher.")
            return
        
        lines = []
        for row in rappels:
            dc = row["date_commande"]
            try:
                dc_fmt = datetime.strptime(dc, "%Y-%m-%d").strftime("%d/%m/%Y") if dc else ""
            except Exception:
                dc_fmt = dc or ""
            
            # sqlite3.Row supporte l'accès par [] mais pas .get()
            pr = row["prochaine_date_rappel"] or ""
            try:
                pr_fmt = datetime.strptime(pr, "%Y-%m-%d").strftime("%d/%m/%Y") if pr else ""
            except Exception:
                pr_fmt = pr or ""
            
            lines.append(
                f"{row['num_commande']} - {row['fournisseur']}\n"
                f"   Commande: {dc_fmt} | Rappel: {pr_fmt}"
            )
        
        text = "\n\n".join(lines[:10])
        if len(lines) > 10:
            text += f"\n\n… et {len(lines) - 10} autres rappels."
        
        text = f"🧪 MODE TEST - Simulation de rappel\n\n{text}\n\n(Aucune modification des dates de rappel)"
        
        QMessageBox.information(self, "🔔 Rappels commandes (TEST)", text)

    # ------------ Rafraîchissement vues ------------

    def refresh_fournisseur_filter(self):
        """Mise à jour de la liste des fournisseurs dans le filtre."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        cur = self.db.conn.cursor()
        fournisseurs_set = set()
        
        if "Commandes" in tab_name:
            cur.execute(
                "SELECT DISTINCT fournisseur FROM commandes WHERE fournisseur IS NOT NULL AND fournisseur != '' ORDER BY fournisseur"
            )
            fournisseurs_set.update([r[0] for r in cur.fetchall()])
        elif "Factures" in tab_name and "Facturation" not in tab_name:
            cur.execute(
                "SELECT DISTINCT fournisseur FROM factures WHERE fournisseur IS NOT NULL AND fournisseur != '' ORDER BY fournisseur"
            )
            fournisseurs_set.update([r[0] for r in cur.fetchall()])
        elif "Facturation" in tab_name:
            # Fournisseurs depuis commandes
            cur.execute(
                "SELECT DISTINCT fournisseur FROM commandes WHERE fournisseur IS NOT NULL AND fournisseur != '' ORDER BY fournisseur"
            )
            fournisseurs_set.update([r[0] for r in cur.fetchall()])
            # Fournisseurs depuis factures
            cur.execute(
                "SELECT DISTINCT fournisseur FROM factures WHERE fournisseur IS NOT NULL AND fournisseur != '' ORDER BY fournisseur"
            )
            fournisseurs_set.update([r[0] for r in cur.fetchall()])
        else:
            return
        
        fournisseurs = sorted(list(fournisseurs_set))
        self.fournisseur_filter_combo.blockSignals(True)
        self.fournisseur_filter_combo.clear()
        self.fournisseur_filter_combo.addItem("Tous")
        for f in fournisseurs:
            self.fournisseur_filter_combo.addItem(f)
        self.fournisseur_filter_combo.blockSignals(False)

    def refresh_marche_filter(self):
        """Mise à jour de la liste des marchés dans le filtre."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        # Le filtre marché est pertinent pour Commandes, Factures et Facturation
        if "Commandes" not in tab_name and "Factures" not in tab_name and "Facturation" not in tab_name:
            return
        
        marches = self._get_marches()
        self.marche_filter_combo.blockSignals(True)
        self.marche_filter_combo.clear()
        self.marche_filter_combo.addItem("Tous")
        for m in marches:
            self.marche_filter_combo.addItem(m)
        self.marche_filter_combo.blockSignals(False)
    
    def refresh_multiple_filters(self):
        """Mise à jour des filtres à choix multiples."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        # Les filtres multiples sont seulement pour Commandes
        if "Commandes" not in tab_name:
            return
        
        cur = self.db.conn.cursor()
        
        # Article fonction
        cur.execute(
            "SELECT DISTINCT article_fonction FROM commandes WHERE article_fonction IS NOT NULL AND article_fonction != '' ORDER BY article_fonction"
        )
        article_fonctions = [r[0] for r in cur.fetchall()]
        self.article_fonction_filter.blockSignals(True)
        self.article_fonction_filter.clear()
        for af in article_fonctions:
            self.article_fonction_filter.addItem(af)
        self.article_fonction_filter.blockSignals(False)
        
        # Article nature
        cur.execute(
            "SELECT DISTINCT article_nature FROM commandes WHERE article_nature IS NOT NULL AND article_nature != '' ORDER BY article_nature"
        )
        article_natures = [r[0] for r in cur.fetchall()]
        self.article_nature_filter.blockSignals(True)
        self.article_nature_filter.clear()
        for an in article_natures:
            self.article_nature_filter.addItem(an)
        self.article_nature_filter.blockSignals(False)
        
        # Service émetteur
        cur.execute(
            "SELECT DISTINCT service_emetteur FROM commandes WHERE service_emetteur IS NOT NULL AND service_emetteur != '' ORDER BY service_emetteur"
        )
        service_emetteurs = [r[0] for r in cur.fetchall()]
        self.service_emetteur_filter.blockSignals(True)
        self.service_emetteur_filter.clear()
        for se in service_emetteurs:
            self.service_emetteur_filter.addItem(se)
        self.service_emetteur_filter.blockSignals(False)
        # Mettre à jour l'état visuel des labels (actif/inactif)
        self.update_multi_filter_labels()


    def refresh_rappels_tab(self):
        rows = self.db.all_active_reminders()
        self.table_rappels.setRowCount(len(rows))

        def fmt(val):
            if not val:
                return ""
            try:
                # Supporte soit une date seule, soit une date avec heure.
                if len(val) > 10:
                    d = datetime.strptime(val, "%Y-%m-%d %H:%M")
                    return d.strftime("%d/%m/%Y %H:%M")
                else:
                    d = datetime.strptime(val, "%Y-%m-%d")
                    return d.strftime("%d/%m/%Y")
            except Exception:
                return val

        for i, row in enumerate(rows):
            vals = [
                row["exercice"] or "",
                row["num_commande"] or "",
                row["fournisseur"] or "",
                fmt(row["date_commande"]),
                fmt(row["prochaine_date_rappel"]),
                row["statut"] or "",
            ]
            for j, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                if j in (3, 4):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                self.table_rappels.setItem(i, j, item)

        self.table_rappels.resizeColumnsToContents()
        self.table_rappels.resizeRowsToContents()
        
        # Mettre à jour le badge de l'onglet Rappels
        self._update_rappels_badge()

    def resize_all(self):
        """Ajuste colonnes et lignes de tous les onglets sur le contenu."""
        # Commandes
        self.table_cmd.resizeColumnsToContents()
        self.table_cmd.resizeRowsToContents()

        # Rappels
        self.table_rappels.resizeColumnsToContents()
        self.table_rappels.resizeRowsToContents()

        # Factures
        self.table_fact.resizeColumnsToContents()
        self.table_fact.resizeRowsToContents()

        # Facturation
        self.table_synth.resizeColumnsToContents()
        self.table_synth.resizeRowsToContents()

    # ------------ Config ------------

    
    def _update_tab_colors(self, index):
        """Met à jour dynamiquement les couleurs des onglets selon l'onglet actif."""
        # Couleurs selon l'onglet
        colors = {
            0: {"bg": "#0078d4", "border": "#005a9e", "shadow": "rgba(0, 120, 212, 0.3)"},  # Commandes - Bleu
            1: {"bg": "#fd7e14", "border": "#e8590c", "shadow": "rgba(253, 126, 20, 0.3)"},  # Rappels - Orange
            2: {"bg": "#28a745", "border": "#218838", "shadow": "rgba(40, 167, 69, 0.3)"},   # Factures - Vert
            3: {"bg": "#6f42c1", "border": "#5a32a3", "shadow": "rgba(111, 66, 193, 0.3)"},  # Facturation - Violet
        }
        
        if index in colors:
            color = colors[index]
            # Mettre à jour le style avec la couleur de l'onglet actif
            self.tabs.setStyleSheet(f"""
                QTabWidget::pane {{
                    border: 2px solid #e0e0e0;
                    border-radius: 8px;
                    background-color: white;
                    top: -2px;
                }}
                QTabBar::tab {{
                    background-color: #f8f9fa;
                    color: #495057;
                    border: 2px solid transparent;
                    border-radius: 8px 8px 0 0;
                    padding: 12px 24px;
                    margin-right: 4px;
                    font-size: 10pt;
                    font-weight: normal;
                    min-width: 120px;
                }}
                QTabBar::tab:hover {{
                    background-color: #e9ecef;
                }}
                QTabBar::tab:selected {{
                    background-color: {color['bg']};
                    color: white;
                    border: 2px solid {color['border']};
                    font-weight: bold;
                }}
            """)
    
    def _update_rappels_badge(self):
        """Met à jour le badge de l'onglet Rappels avec le nombre de rappels actifs."""
        try:
            # Compter le nombre de rappels actifs
            rows = self.db.all_active_reminders()
            count = len(rows)
            
            # Trouver l'index de l'onglet Rappels
            for i in range(self.tabs.count()):
                tab_text = self.tabs.tabText(i)
                if "Rappels" in tab_text:
                    # Mettre à jour le texte avec le badge
                    if count > 0:
                        new_text = f"🔔 Rappels ({count})"
                    else:
                        new_text = "🔔 Rappels"
                    self.tabs.setTabText(i, new_text)
                    break
        except Exception:
            pass  # En cas d'erreur, on ne fait rien

    # ==================== FONCTIONS D'EXPORT ====================
    
    def _get_active_table_data(self):
        """Récupère les données du tableau actif avec métadonnées."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        # Déterminer le tableau, le modèle et les colonnes
        if "Commandes" in tab_name:
            table = self.table_cmd
            proxy = self.cmd_proxy
            source_model = self.cmd_model
            columns = COMMANDES_COLUMNS
            titre = "Liste des commandes"
        elif "Factures" in tab_name and "Facturation" not in tab_name:
            table = self.table_fact
            proxy = self.fact_proxy
            source_model = self.fact_model
            columns = FACTURES_COLUMNS
            titre = "Liste des factures"
        elif "Facturation" in tab_name:
            table = self.table_synth
            proxy = self.synth_proxy
            source_model = self.synth_model
            columns = FACTURATION_COLUMNS
            titre = "Synthèse de facturation"
        elif "Rappels" in tab_name:
            # Pour les rappels, on utilise le QTableWidget
            titre = "Rappels de paiement en cours"
            headers = []
            for col in range(self.table_rappels.columnCount()):
                header_item = self.table_rappels.horizontalHeaderItem(col)
                headers.append(header_item.text() if header_item else f"Col{col}")
            
            data = []
            for row in range(self.table_rappels.rowCount()):
                row_data = []
                for col in range(self.table_rappels.columnCount()):
                    item = self.table_rappels.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            
            return {
                "titre": titre,
                "headers": headers,
                "data": data,
                "tab_name": tab_name
            }
        else:
            return None
        
        # Récupérer les en-têtes
        headers = [col[1] for col in columns]
        
        # Récupérer les données filtrées
        data = []
        config = self.db.get_config_exports()
        
        if config.get("lignes_filtrees_uniquement", True):
            # Exporter seulement les lignes visibles (filtrées)
            for row in range(proxy.rowCount()):
                row_data = []
                for col in range(len(columns)):
                    index = proxy.index(row, col)
                    value = proxy.data(index, Qt.DisplayRole)
                    row_data.append(str(value) if value is not None else "")
                data.append(row_data)
        else:
            # Exporter toutes les lignes (non filtré)
            for row in range(source_model.rowCount()):
                row_data = []
                for col in range(len(columns)):
                    index = source_model.index(row, col)
                    value = source_model.data(index, Qt.DisplayRole)
                    row_data.append(str(value) if value is not None else "")
                data.append(row_data)
        
        return {
            "titre": titre,
            "headers": headers,
            "data": data,
            "tab_name": tab_name,
            "columns": columns,
            "proxy": proxy,
            "source_model": source_model
        }
    
    def _get_active_filters_description(self):
        """Retourne une description textuelle des filtres actifs."""
        tab_index = self.tabs.currentIndex()
        tab_name = self.tabs.tabText(tab_index)
        
        filters = []
        
        # Filtre 1
        filter1_text = self.filter1_combo.currentText()
        if filter1_text and filter1_text != "Tous":
            filters.append(f"Statut: {filter1_text}")
        
        # Filtre 2
        if self.filter2_combo.isVisible():
            filter2_text = self.filter2_combo.currentText()
            if filter2_text and filter2_text != "Tous":
                filters.append(f"Exercice/Facturation: {filter2_text}")
        
        # Filtre fournisseur
        fournisseur_text = self.fournisseur_filter_combo.currentText()
        if fournisseur_text and fournisseur_text != "Tous":
            filters.append(f"Fournisseur: {fournisseur_text}")
        
        # Filtre marché
        if self.marche_filter_combo.isVisible():
            marche_text = self.marche_filter_combo.currentText()
            if marche_text and marche_text != "Tous":
                filters.append(f"Marché: {marche_text}")
        
        # Filtres multiples (Commandes uniquement)
        if "Commandes" in tab_name:
            checked_af = self.article_fonction_filter.checked_items()
            if checked_af:
                filters.append(f"Art. fonction: {', '.join(checked_af)}")
            
            checked_an = self.article_nature_filter.checked_items()
            if checked_an:
                filters.append(f"Art. nature: {', '.join(checked_an)}")
            
            checked_se = self.service_emetteur_filter.checked_items()
            if checked_se:
                filters.append(f"Service: {', '.join(checked_se)}")
        
        if not filters:
            return "Aucun filtre appliqué"
        
        return " | ".join(filters)
    
    def export_to_pdf(self):
        """Exporte le tableau actif en PDF avec mise en page professionnelle."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            
            # Récupérer les données
            table_data = self._get_active_table_data()
            if not table_data:
                QMessageBox.warning(self, "Export PDF", "Aucune donnée à exporter.")
                return
            
            # Récupérer la configuration
            config = self.db.get_config_exports()
            
            # Dialogue de sauvegarde
            today = date.today().strftime("%Y-%m-%d")
            default_name = f"{table_data['titre'].replace(' ', '_')}_{today}.pdf"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Enregistrer le PDF",
                default_name,
                "PDF (*.pdf)"
            )
            
            if not file_path:
                return
            
            # Déterminer l'orientation selon le nombre de colonnes
            num_cols = len(table_data['headers'])
            pagesize = landscape(A4) if num_cols > 6 else A4
            
            # Créer le document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=pagesize,
                leftMargin=1.5*cm,
                rightMargin=1.5*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )
            
            # Styles
            styles = getSampleStyleSheet()
            story = []
            
            # === EN-TÊTE ===
            # Logo (si configuré)
            logo_path = config.get("logo_path", "")
            if logo_path and os.path.exists(logo_path):
                try:
                    logo = Image(logo_path, width=4*cm, height=1.6*cm)
                    story.append(logo)
                    story.append(Spacer(1, 0.3*cm))
                except Exception:
                    pass  # Si le logo ne peut pas être chargé, on continue
            
            # Informations entreprise
            nom_entreprise = config.get("nom_entreprise", "")
            if nom_entreprise:
                style_entreprise = ParagraphStyle(
                    'Entreprise',
                    parent=styles['Normal'],
                    fontSize=12,
                    alignment=TA_CENTER,
                    fontName='Helvetica-Bold'
                )
                story.append(Paragraph(nom_entreprise, style_entreprise))
                
                adresse_1 = config.get("adresse_1", "")
                adresse_2 = config.get("adresse_2", "")
                code_postal = config.get("code_postal", "")
                ville = config.get("ville", "")
                
                style_adresse = ParagraphStyle(
                    'Adresse',
                    parent=styles['Normal'],
                    fontSize=9,
                    alignment=TA_CENTER
                )
                
                if adresse_1:
                    story.append(Paragraph(adresse_1, style_adresse))
                if adresse_2:
                    story.append(Paragraph(adresse_2, style_adresse))
                if code_postal or ville:
                    story.append(Paragraph(f"{code_postal} {ville}", style_adresse))
                
                story.append(Spacer(1, 0.5*cm))
            
            # Ligne de séparation
            from reportlab.platypus import HRFlowable
            story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#0078d4")))
            story.append(Spacer(1, 0.3*cm))
            
            # Titre du document
            style_titre = ParagraphStyle(
                'Titre',
                parent=styles['Heading1'],
                fontSize=16,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor("#0078d4")
            )
            story.append(Paragraph(table_data['titre'], style_titre))
            
            # Date de génération
            now = datetime.now().strftime("%d/%m/%Y à %H:%M")
            style_date = ParagraphStyle(
                'Date',
                parent=styles['Normal'],
                fontSize=9,
                alignment=TA_CENTER,
                textColor=colors.grey
            )
            story.append(Paragraph(f"Généré le : {now}", style_date))
            story.append(Spacer(1, 0.2*cm))
            
            # Ligne de séparation
            story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#0078d4")))
            story.append(Spacer(1, 0.2*cm))
            
            # Filtres appliqués
            filters_desc = self._get_active_filters_description()
            style_filtres = ParagraphStyle(
                'Filtres',
                parent=styles['Normal'],
                fontSize=8,
                alignment=TA_LEFT,
                textColor=colors.grey
            )
            story.append(Paragraph(f"<b>Filtres appliqués :</b> {filters_desc}", style_filtres))
            story.append(Spacer(1, 0.5*cm))
            
            # === TABLEAU DE DONNÉES ===
            # Préparer les données avec word wrap pour certaines colonnes
            
            # Style pour les cellules avec word wrap
            style_cell = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=6,
                leading=7,
                alignment=TA_LEFT
            )
            
            # Convertir les données en Paragraphs pour word wrap sur colonnes 2 (Fournisseur) et 3 (Libellé)
            pdf_data = []
            
            # En-têtes (première ligne)
            pdf_data.append(table_data['headers'])
            
            # Données avec word wrap
            for row_data in table_data['data']:
                new_row = []
                for col_idx, cell_value in enumerate(row_data):
                    # Colonnes 2 (Fournisseur) et 3 (Libellé) : word wrap
                    if col_idx == 2:  # Fournisseur - word wrap 2 lignes max
                        # Limiter à environ 35 caractères pour forcer 2 lignes
                        text = str(cell_value)[:70] if cell_value else ""
                        new_row.append(Paragraph(text, style_cell))
                    elif col_idx == 3:  # Libellé - word wrap
                        text = str(cell_value) if cell_value else ""
                        new_row.append(Paragraph(text, style_cell))
                    elif col_idx == 8:  # Section - raccourcir les termes
                        text = str(cell_value) if cell_value else ""
                        text = text.replace("Fonctionnement", "Fonct").replace("Investissement", "Invest")
                        new_row.append(text)
                    elif col_idx == 16:  # Statut facturation - ajouter retour ligne
                        text = str(cell_value) if cell_value else ""
                        # Ajouter retour ligne après le premier mot (ex: "Totalement\nfacturée")
                        if " " in text:
                            parts = text.split(" ", 1)
                            text = parts[0] + "\n" + parts[1]
                        new_row.append(text)
                    else:
                        # Autres colonnes : texte simple
                        new_row.append(str(cell_value) if cell_value else "")
                pdf_data.append(new_row)
            
            # Définir les largeurs de colonnes en cm (optimisées pour A4 paysage)
            num_cols = len(table_data['headers'])
            
            if num_cols == 17:  # Commandes
                col_widths = [
                    1.0*cm,   # Exercice
                    1.3*cm,   # N° Commande
                    3.0*cm,   # Fournisseur (word wrap)
                    4.5*cm,   # Libellé (word wrap) ← PLUS LARGE
                    1.5*cm,   # Date commande
                    1.5*cm,   # Marché
                    1.3*cm,   # Service émetteur
                    1.5*cm,   # Montant TTC
                    1.5*cm,   # Section ← AU LIEU DE 0.9
                    1.2*cm,   # Article fonction ← AU LIEU DE 0.7
                    1.2*cm,   # Article nature ← AU LIEU DE 0.7
                    1.3*cm,   # Statut
                    1.5*cm,   # Date envoi
                    1.5*cm,   # Prochain rappel
                    1.5*cm,   # Montant facturé
                    1.5*cm,   # Reste à facturer
                    1.7*cm,   # Statut facturation
                ]
            else:
                # Largeurs automatiques pour autres onglets
                col_widths = None
            
            # Créer le tableau avec largeurs de colonnes
            pdf_table = Table(pdf_data, colWidths=col_widths, repeatRows=1)
            
            # Style du tableau (optimisé pour densité)
            table_style = [
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0078d4")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 6),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
                ('TOPPADDING', (0, 0), (-1, 0), 3),
                ('LEFTPADDING', (0, 0), (-1, 0), 2),
                ('RIGHTPADDING', (0, 0), (-1, 0), 2),
                # Données
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 1), (-1, -1), 2),
                ('RIGHTPADDING', (0, 1), (-1, -1), 2),
                ('TOPPADDING', (0, 1), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")])
            ]
            
            # Alignement à droite pour les colonnes de montants (Commandes)
            if num_cols == 17:
                # Colonnes : Montant TTC (7), Montant facturé (14), Reste à facturer (15)
                table_style.append(('ALIGN', (7, 1), (7, -1), 'RIGHT'))   # Montant TTC
                table_style.append(('ALIGN', (14, 1), (14, -1), 'RIGHT')) # Montant facturé
                table_style.append(('ALIGN', (15, 1), (15, -1), 'RIGHT')) # Reste à facturer
            
            # Appliquer les couleurs de fond selon le statut (si demandé)
            if config.get("inclure_couleurs", True) and "Facturation" in table_data.get("tab_name", ""):
                # Pour la facturation, colorier selon le statut
                for row_idx, row_data in enumerate(table_data['data'], start=1):
                    # Le statut est généralement en avant-dernière colonne
                    if len(row_data) >= 2:
                        statut = row_data[-2] if len(row_data) > 1 else ""
                        if "Totalement" in str(statut):
                            table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor("#e0ffe0")))
                        elif "Partiellement" in str(statut):
                            table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor("#fff3cd")))
                        elif "Non facturée" in str(statut):
                            table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor("#ffe0e0")))
            
            pdf_table.setStyle(TableStyle(table_style))
            story.append(pdf_table)
            
            # Pied de page (sera sur toutes les pages)
            story.append(Spacer(1, 1*cm))
            style_footer = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=7,
                alignment=TA_CENTER,
                textColor=colors.grey
            )
            story.append(Paragraph(f"Document généré par Suivi Commandes/Factures/Marchés", style_footer))
            
            # Générer le PDF
            doc.build(story)
            
            # Message de succès
            reply = QMessageBox.question(
                self,
                "Export réussi",
                f"Le fichier PDF a été créé avec succès :\n{file_path}\n\nVoulez-vous l'ouvrir ?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                import subprocess
                if sys.platform == "win32":
                    os.startfile(file_path)
                elif sys.platform == "darwin":
                    subprocess.call(["open", file_path])
                else:
                    subprocess.call(["xdg-open", file_path])
        
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export PDF", f"Une erreur est survenue :\n{str(e)}")
    
    def export_to_excel(self):
        """Exporte le tableau actif en Excel avec mise en page professionnelle."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as OpenpyxlImage
            
            # Récupérer les données
            table_data = self._get_active_table_data()
            if not table_data:
                QMessageBox.warning(self, "Export Excel", "Aucune donnée à exporter.")
                return
            
            # Récupérer la configuration
            config = self.db.get_config_exports()
            
            # Dialogue de sauvegarde
            today = date.today().strftime("%Y-%m-%d")
            default_name = f"{table_data['titre'].replace(' ', '_')}_{today}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Enregistrer le fichier Excel",
                default_name,
                "Excel (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Créer le classeur
            wb = Workbook()
            
            # === FEUILLE DE GARDE ===
            ws_garde = wb.active
            ws_garde.title = "Informations"
            
            row = 1
            
            # Logo (si configuré)
            logo_path = config.get("logo_path", "")
            if logo_path and os.path.exists(logo_path):
                try:
                    img = OpenpyxlImage(logo_path)
                    # Redimensionner si nécessaire
                    img.width = 200
                    img.height = 80
                    ws_garde.add_image(img, 'A1')
                    row = 6
                except Exception:
                    pass
            
            # Informations entreprise
            nom_entreprise = config.get("nom_entreprise", "")
            if nom_entreprise:
                ws_garde[f'A{row}'] = nom_entreprise
                ws_garde[f'A{row}'].font = Font(size=14, bold=True)
                row += 1
                
                adresse_1 = config.get("adresse_1", "")
                adresse_2 = config.get("adresse_2", "")
                code_postal = config.get("code_postal", "")
                ville = config.get("ville", "")
                
                if adresse_1:
                    ws_garde[f'A{row}'] = adresse_1
                    row += 1
                if adresse_2:
                    ws_garde[f'A{row}'] = adresse_2
                    row += 1
                if code_postal or ville:
                    ws_garde[f'A{row}'] = f"{code_postal} {ville}"
                    row += 1
            
            row += 2
            
            # Titre du document
            ws_garde[f'A{row}'] = table_data['titre']
            ws_garde[f'A{row}'].font = Font(size=16, bold=True, color="0078D4")
            row += 1
            
            # Date de génération
            now = datetime.now().strftime("%d/%m/%Y à %H:%M")
            ws_garde[f'A{row}'] = f"Généré le : {now}"
            ws_garde[f'A{row}'].font = Font(size=10, italic=True)
            row += 2
            
            # Filtres appliqués
            filters_desc = self._get_active_filters_description()
            ws_garde[f'A{row}'] = "Filtres appliqués :"
            ws_garde[f'A{row}'].font = Font(bold=True)
            row += 1
            ws_garde[f'A{row}'] = filters_desc
            ws_garde[f'A{row}'].font = Font(size=9, color="666666")
            
            # Largeur de colonne
            ws_garde.column_dimensions['A'].width = 60
            
            # === FEUILLE DE DONNÉES ===
            ws_data = wb.create_sheet(title="Données")
            
            # En-têtes
            for col_idx, header in enumerate(table_data['headers'], start=1):
                cell = ws_data.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Données
            for row_idx, row_data in enumerate(table_data['data'], start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin', color="CCCCCC"),
                        right=Side(style='thin', color="CCCCCC"),
                        top=Side(style='thin', color="CCCCCC"),
                        bottom=Side(style='thin', color="CCCCCC")
                    )
                    
                    # Alternance de couleurs
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            # Appliquer les couleurs selon le statut (si demandé)
            if config.get("inclure_couleurs", True) and "Facturation" in table_data.get("tab_name", ""):
                for row_idx, row_data in enumerate(table_data['data'], start=2):
                    if len(row_data) >= 2:
                        statut = row_data[-2] if len(row_data) > 1 else ""
                        color = None
                        if "Totalement" in str(statut):
                            color = "E0FFE0"
                        elif "Partiellement" in str(statut):
                            color = "FFF3CD"
                        elif "Non facturée" in str(statut):
                            color = "FFE0E0"
                        
                        if color:
                            for col_idx in range(1, len(row_data) + 1):
                                ws_data.cell(row=row_idx, column=col_idx).fill = PatternFill(
                                    start_color=color, end_color=color, fill_type="solid"
                                )
            
            # Figer la première ligne
            ws_data.freeze_panes = "A2"
            
            # Activer les filtres automatiques
            ws_data.auto_filter.ref = ws_data.dimensions
            
            # Ajuster les largeurs de colonnes
            for col_idx, header in enumerate(table_data['headers'], start=1):
                # Calcul de la largeur optimale
                max_length = len(str(header))
                for row_data in table_data['data'][:100]:  # Échantillon des 100 premières lignes
                    if col_idx <= len(row_data):
                        cell_value = str(row_data[col_idx - 1])
                        max_length = max(max_length, len(cell_value))
                
                # Limiter la largeur entre 10 et 50
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws_data.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # === FEUILLE STATISTIQUES ===
            ws_stats = wb.create_sheet(title="Statistiques")
            
            # Titre principal
            row = 1
            ws_stats[f'A{row}'] = "Statistiques détaillées"
            ws_stats[f'A{row}'].font = Font(size=14, bold=True, color="0078D4")
            ws_stats.merge_cells(f'A{row}:D{row}')
            row += 2
            
            # Analyser les données pour extraire les colonnes nécessaires
            headers = table_data['headers']
            data = table_data['data']
            
            # Trouver les indices des colonnes (peuvent varier selon l'onglet)
            col_indices = {}
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if "montant" in header_lower and "ttc" in header_lower:
                    col_indices['montant_ttc'] = idx
                elif "montant" in header_lower and "factur" in header_lower:
                    col_indices['montant_facture'] = idx
                elif "reste" in header_lower and "facturer" in header_lower:
                    col_indices['reste_a_facturer'] = idx
                elif "fournisseur" in header_lower:
                    col_indices['fournisseur'] = idx
                elif "statut" in header_lower and "facturation" in header_lower:
                    col_indices['statut_facturation'] = idx
                elif "statut" in header_lower and "facturation" not in header_lower:
                    col_indices['statut'] = idx
                elif "section" in header_lower:
                    col_indices['section'] = idx
                elif "exercice" in header_lower:
                    col_indices['exercice'] = idx
                elif "marché" in header_lower or "marche" in header_lower:
                    col_indices['marche'] = idx
            
            # Fonction helper pour extraire montant
            def parse_montant(value):
                if not value:
                    return 0.0
                try:
                    clean = str(value).replace(' ', '').replace(',', '.').replace('€', '').strip()
                    return float(clean)
                except:
                    return 0.0
            
            # ==========================================
            # BLOC 1 : SYNTHÈSE FINANCIÈRE
            # ==========================================
            ws_stats[f'A{row}'] = "📊 SYNTHÈSE FINANCIÈRE"
            ws_stats[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
            ws_stats[f'A{row}'].fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            ws_stats.merge_cells(f'A{row}:B{row}')
            row += 1
            
            total_commandes = 0.0
            total_facture = 0.0
            total_reste = 0.0
            nb_commandes = len(data)
            
            for row_data in data:
                if 'montant_ttc' in col_indices and col_indices['montant_ttc'] < len(row_data):
                    total_commandes += parse_montant(row_data[col_indices['montant_ttc']])
                if 'montant_facture' in col_indices and col_indices['montant_facture'] < len(row_data):
                    total_facture += parse_montant(row_data[col_indices['montant_facture']])
                if 'reste_a_facturer' in col_indices and col_indices['reste_a_facturer'] < len(row_data):
                    total_reste += parse_montant(row_data[col_indices['reste_a_facturer']])
            
            taux_facturation = (total_facture / total_commandes * 100) if total_commandes > 0 else 0
            montant_moyen = total_commandes / nb_commandes if nb_commandes > 0 else 0
            
            ws_stats[f'A{row}'] = "Montant total des commandes (TTC)"
            ws_stats[f'B{row}'] = total_commandes
            ws_stats[f'B{row}'].number_format = '#,##0.00 €'
            ws_stats[f'B{row}'].font = Font(bold=True)
            row += 1
            
            ws_stats[f'A{row}'] = "Montant total facturé"
            ws_stats[f'B{row}'] = total_facture
            ws_stats[f'B{row}'].number_format = '#,##0.00 €'
            ws_stats[f'B{row}'].font = Font(color="28A745")
            row += 1
            
            ws_stats[f'A{row}'] = "Reste à facturer"
            ws_stats[f'B{row}'] = total_reste
            ws_stats[f'B{row}'].number_format = '#,##0.00 €'
            ws_stats[f'B{row}'].font = Font(color="DC3545", bold=True)
            row += 1
            
            ws_stats[f'A{row}'] = "Taux de facturation"
            ws_stats[f'B{row}'] = taux_facturation / 100
            ws_stats[f'B{row}'].number_format = '0.00%'
            ws_stats[f'B{row}'].font = Font(bold=True)
            row += 1
            
            ws_stats[f'A{row}'] = "Montant moyen par commande"
            ws_stats[f'B{row}'] = montant_moyen
            ws_stats[f'B{row}'].number_format = '#,##0.00 €'
            row += 2
            
            # ==========================================
            # BLOC 2 : STATUT DE FACTURATION
            # ==========================================
            if 'statut_facturation' in col_indices:
                ws_stats[f'A{row}'] = "💰 RÉPARTITION PAR STATUT DE FACTURATION"
                ws_stats[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
                ws_stats[f'A{row}'].fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
                ws_stats.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # En-têtes tableau
                ws_stats[f'A{row}'] = "Statut facturation"
                ws_stats[f'B{row}'] = "Nombre"
                ws_stats[f'C{row}'] = "Montant"
                ws_stats[f'D{row}'] = "% du total"
                for col in ['A', 'B', 'C', 'D']:
                    ws_stats[f'{col}{row}'].font = Font(bold=True)
                    ws_stats[f'{col}{row}'].fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                row += 1
                
                # Calculer les stats
                stats_fact = {}
                for row_data in data:
                    if col_indices['statut_facturation'] < len(row_data):
                        statut = str(row_data[col_indices['statut_facturation']])
                        if statut not in stats_fact:
                            stats_fact[statut] = {'count': 0, 'montant': 0.0}
                        stats_fact[statut]['count'] += 1
                        if 'montant_ttc' in col_indices and col_indices['montant_ttc'] < len(row_data):
                            stats_fact[statut]['montant'] += parse_montant(row_data[col_indices['montant_ttc']])
                
                # Ordre souhaité
                ordre = ["Non facturée", "Partiellement facturée", "Totalement facturée"]
                for statut in ordre:
                    if statut in stats_fact:
                        count = stats_fact[statut]['count']
                        montant = stats_fact[statut]['montant']
                        pct = (montant / total_commandes * 100) if total_commandes > 0 else 0
                        
                        ws_stats[f'A{row}'] = statut
                        ws_stats[f'B{row}'] = count
                        ws_stats[f'C{row}'] = montant
                        ws_stats[f'C{row}'].number_format = '#,##0.00 €'
                        ws_stats[f'D{row}'] = pct / 100
                        ws_stats[f'D{row}'].number_format = '0.00%'
                        
                        # Couleur selon statut
                        color = None
                        if "Totalement" in statut:
                            color = "E0FFE0"
                        elif "Partiellement" in statut:
                            color = "FFF3CD"
                        elif "Non facturée" in statut:
                            color = "FFE0E0"
                        
                        if color:
                            for col in ['A', 'B', 'C', 'D']:
                                ws_stats[f'{col}{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        
                        row += 1
                row += 1
            
            # ==========================================
            # BLOC 3 : TOP 10 FOURNISSEURS
            # ==========================================
            if 'fournisseur' in col_indices:
                ws_stats[f'A{row}'] = "🏢 TOP 10 FOURNISSEURS"
                ws_stats[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
                ws_stats[f'A{row}'].fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
                ws_stats.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # En-têtes
                ws_stats[f'A{row}'] = "Fournisseur"
                ws_stats[f'B{row}'] = "Nb commandes"
                ws_stats[f'C{row}'] = "Montant total"
                ws_stats[f'D{row}'] = "% du total"
                for col in ['A', 'B', 'C', 'D']:
                    ws_stats[f'{col}{row}'].font = Font(bold=True)
                    ws_stats[f'{col}{row}'].fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                row += 1
                
                # Calculer les stats
                stats_fourn = {}
                for row_data in data:
                    if col_indices['fournisseur'] < len(row_data):
                        fourn = str(row_data[col_indices['fournisseur']]).strip()
                        if not fourn:
                            fourn = "(Vide)"
                        if fourn not in stats_fourn:
                            stats_fourn[fourn] = {'count': 0, 'montant': 0.0}
                        stats_fourn[fourn]['count'] += 1
                        if 'montant_ttc' in col_indices and col_indices['montant_ttc'] < len(row_data):
                            stats_fourn[fourn]['montant'] += parse_montant(row_data[col_indices['montant_ttc']])
                
                # Trier par montant décroissant et prendre top 10
                top_fourn = sorted(stats_fourn.items(), key=lambda x: x[1]['montant'], reverse=True)[:10]
                
                for fourn, stats in top_fourn:
                    count = stats['count']
                    montant = stats['montant']
                    pct = (montant / total_commandes * 100) if total_commandes > 0 else 0
                    
                    ws_stats[f'A{row}'] = fourn
                    ws_stats[f'B{row}'] = count
                    ws_stats[f'C{row}'] = montant
                    ws_stats[f'C{row}'].number_format = '#,##0.00 €'
                    ws_stats[f'D{row}'] = pct / 100
                    ws_stats[f'D{row}'].number_format = '0.00%'
                    
                    row += 1
                row += 1
            
            # ==========================================
            # BLOC 4 : SECTION FONCTIONNEMENT/INVESTISSEMENT
            # ==========================================
            if 'section' in col_indices:
                ws_stats[f'A{row}'] = "🎯 RÉPARTITION PAR SECTION"
                ws_stats[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
                ws_stats[f'A{row}'].fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
                ws_stats.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # En-têtes
                ws_stats[f'A{row}'] = "Section"
                ws_stats[f'B{row}'] = "Nombre"
                ws_stats[f'C{row}'] = "Montant"
                ws_stats[f'D{row}'] = "% du total"
                for col in ['A', 'B', 'C', 'D']:
                    ws_stats[f'{col}{row}'].font = Font(bold=True)
                    ws_stats[f'{col}{row}'].fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                row += 1
                
                # Calculer les stats
                stats_section = {}
                for row_data in data:
                    if col_indices['section'] < len(row_data):
                        section = str(row_data[col_indices['section']]).strip()
                        if not section:
                            section = "(Non renseigné)"
                        if section not in stats_section:
                            stats_section[section] = {'count': 0, 'montant': 0.0}
                        stats_section[section]['count'] += 1
                        if 'montant_ttc' in col_indices and col_indices['montant_ttc'] < len(row_data):
                            stats_section[section]['montant'] += parse_montant(row_data[col_indices['montant_ttc']])
                
                # Trier par montant décroissant
                sorted_sections = sorted(stats_section.items(), key=lambda x: x[1]['montant'], reverse=True)
                
                for section, stats in sorted_sections:
                    count = stats['count']
                    montant = stats['montant']
                    pct = (montant / total_commandes * 100) if total_commandes > 0 else 0
                    
                    ws_stats[f'A{row}'] = section
                    ws_stats[f'B{row}'] = count
                    ws_stats[f'C{row}'] = montant
                    ws_stats[f'C{row}'].number_format = '#,##0.00 €'
                    ws_stats[f'D{row}'] = pct / 100
                    ws_stats[f'D{row}'].number_format = '0.00%'
                    
                    row += 1
                row += 1
            
            # ==========================================
            # BLOC 5 : RÉPARTITION PAR EXERCICE
            # ==========================================
            if 'exercice' in col_indices:
                ws_stats[f'A{row}'] = "📅 RÉPARTITION PAR EXERCICE"
                ws_stats[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
                ws_stats[f'A{row}'].fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
                ws_stats.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # En-têtes
                ws_stats[f'A{row}'] = "Exercice"
                ws_stats[f'B{row}'] = "Nombre"
                ws_stats[f'C{row}'] = "Montant"
                ws_stats[f'D{row}'] = "% du total"
                for col in ['A', 'B', 'C', 'D']:
                    ws_stats[f'{col}{row}'].font = Font(bold=True)
                    ws_stats[f'{col}{row}'].fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                row += 1
                
                # Calculer les stats
                stats_exercice = {}
                for row_data in data:
                    if col_indices['exercice'] < len(row_data):
                        exercice = str(row_data[col_indices['exercice']]).strip()
                        if not exercice:
                            exercice = "(Non renseigné)"
                        if exercice not in stats_exercice:
                            stats_exercice[exercice] = {'count': 0, 'montant': 0.0}
                        stats_exercice[exercice]['count'] += 1
                        if 'montant_ttc' in col_indices and col_indices['montant_ttc'] < len(row_data):
                            stats_exercice[exercice]['montant'] += parse_montant(row_data[col_indices['montant_ttc']])
                
                # Trier par exercice décroissant
                sorted_exercices = sorted(stats_exercice.items(), reverse=True)
                
                for exercice, stats in sorted_exercices:
                    count = stats['count']
                    montant = stats['montant']
                    pct = (montant / total_commandes * 100) if total_commandes > 0 else 0
                    
                    ws_stats[f'A{row}'] = exercice
                    ws_stats[f'B{row}'] = count
                    ws_stats[f'C{row}'] = montant
                    ws_stats[f'C{row}'].number_format = '#,##0.00 €'
                    ws_stats[f'D{row}'] = pct / 100
                    ws_stats[f'D{row}'].number_format = '0.00%'
                    
                    row += 1
                row += 1
            
            # ==========================================
            # BLOC 6 : RÉPARTITION PAR STATUT
            # ==========================================
            if 'statut' in col_indices:
                ws_stats[f'A{row}'] = "📊 RÉPARTITION PAR STATUT"
                ws_stats[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
                ws_stats[f'A{row}'].fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
                ws_stats.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # En-têtes
                ws_stats[f'A{row}'] = "Statut"
                ws_stats[f'B{row}'] = "Nombre"
                ws_stats[f'C{row}'] = "Montant"
                ws_stats[f'D{row}'] = "% du total"
                for col in ['A', 'B', 'C', 'D']:
                    ws_stats[f'{col}{row}'].font = Font(bold=True)
                    ws_stats[f'{col}{row}'].fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                row += 1
                
                # Calculer les stats
                stats_statut = {}
                for row_data in data:
                    if col_indices['statut'] < len(row_data):
                        statut = str(row_data[col_indices['statut']]).strip()
                        if not statut:
                            statut = "(Non renseigné)"
                        if statut not in stats_statut:
                            stats_statut[statut] = {'count': 0, 'montant': 0.0}
                        stats_statut[statut]['count'] += 1
                        if 'montant_ttc' in col_indices and col_indices['montant_ttc'] < len(row_data):
                            stats_statut[statut]['montant'] += parse_montant(row_data[col_indices['montant_ttc']])
                
                # Ordre souhaité
                ordre_statut = ["A suivre", "Envoyée"]
                for statut in ordre_statut:
                    if statut in stats_statut:
                        count = stats_statut[statut]['count']
                        montant = stats_statut[statut]['montant']
                        pct = (montant / total_commandes * 100) if total_commandes > 0 else 0
                        
                        ws_stats[f'A{row}'] = statut
                        ws_stats[f'B{row}'] = count
                        ws_stats[f'C{row}'] = montant
                        ws_stats[f'C{row}'].number_format = '#,##0.00 €'
                        ws_stats[f'D{row}'] = pct / 100
                        ws_stats[f'D{row}'].number_format = '0.00%'
                        
                        # Couleur selon statut
                        if statut == "Envoyée":
                            color = "E0FFE0"
                        elif statut == "A suivre":
                            color = "FFF3CD"
                        else:
                            color = None
                        
                        if color:
                            for col in ['A', 'B', 'C', 'D']:
                                ws_stats[f'{col}{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        
                        row += 1
                
                # Autres statuts (si présents)
                for statut, stats in stats_statut.items():
                    if statut not in ordre_statut:
                        count = stats['count']
                        montant = stats['montant']
                        pct = (montant / total_commandes * 100) if total_commandes > 0 else 0
                        
                        ws_stats[f'A{row}'] = statut
                        ws_stats[f'B{row}'] = count
                        ws_stats[f'C{row}'] = montant
                        ws_stats[f'C{row}'].number_format = '#,##0.00 €'
                        ws_stats[f'D{row}'] = pct / 100
                        ws_stats[f'D{row}'].number_format = '0.00%'
                        
                        row += 1
            
            # Ajuster les largeurs de colonnes
            ws_stats.column_dimensions['A'].width = 35
            ws_stats.column_dimensions['B'].width = 15
            ws_stats.column_dimensions['C'].width = 18
            ws_stats.column_dimensions['D'].width = 12
            
            # Sauvegarder le fichier
            wb.save(file_path)
            
            # Message de succès
            reply = QMessageBox.question(
                self,
                "Export réussi",
                f"Le fichier Excel a été créé avec succès :\n{file_path}\n\nVoulez-vous l'ouvrir ?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                import subprocess
                if sys.platform == "win32":
                    os.startfile(file_path)
                elif sys.platform == "darwin":
                    subprocess.call(["open", file_path])
                else:
                    subprocess.call(["xdg-open", file_path])
        
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export Excel", f"Une erreur est survenue :\n{str(e)}")

    # ============== MÉTHODES POUR LE SUIVI DES MARCHÉS ==============

    def sync_marches_cache_from_database(self):
        """
        Synchronise le cache MarchesSync depuis la base de données principale.
        Retourne le nombre de factures synchronisées.
        """
        from marches_sync import MarchesSync

        # D'abord, vérifier combien de factures au total
        cur = self.db.conn.cursor()
        cur.execute("SELECT COUNT(*) as total FROM factures")
        total_factures = cur.fetchone()['total']
        print(f"[DEBUG] Total factures dans la base: {total_factures}")

        # Vérifier combien ont un marché rempli
        cur.execute("SELECT COUNT(*) as avec_marche FROM factures WHERE marche IS NOT NULL AND marche != ''")
        avec_marche = cur.fetchone()['avec_marche']
        print(f"[DEBUG] Factures avec marché: {avec_marche}")

        # Si aucune facture avec marché, afficher quelques exemples
        if avec_marche == 0:
            cur.execute("SELECT exercice, num_facture, fournisseur, marche FROM factures LIMIT 5")
            exemples = cur.fetchall()
            print(f"[DEBUG] Exemples de factures (5 premières):")
            for ex in exemples:
                print(f"  - {ex['exercice']}/{ex['num_facture']}: marché='{ex['marche']}'")

        # Charger toutes les factures avec un marché non vide depuis la base
        cur.execute("""
            SELECT
                marche,
                fournisseur,
                libelle,
                date_facture as date_sf,
                num_facture,
                montant_ttc,
                montant_service_fait as montant_sf,
                montant_initial,
                num_mandat,
                tranche,
                commande
            FROM factures
            WHERE marche IS NOT NULL AND marche != ''
        """)

        rows = cur.fetchall()

        if not rows:
            print("[WARNING] Aucune facture avec un marché n'a été trouvée dans la base")
            return 0

        # Convertir en DataFrame avec noms de colonnes explicites
        # Les colonnes doivent correspondre exactement à l'ordre du SELECT
        column_names = [
            'marche', 'fournisseur', 'libelle', 'date_sf', 'num_facture',
            'montant_ttc', 'montant_sf', 'montant_initial', 'num_mandat',
            'tranche', 'commande'
        ]

        # Convertir les rows en liste de tuples pour créer le DataFrame
        data = [tuple(row) for row in rows]
        df = pd.DataFrame(data, columns=column_names)

        print(f"[DEBUG] DataFrame créé avec {len(df)} lignes et colonnes: {list(df.columns)}")

        # Si montant_initial est vide/null, utiliser montant_ttc comme fallback
        try:
            if 'montant_initial' in df.columns and 'montant_ttc' in df.columns:
                # Remplacer les valeurs null de montant_initial par montant_ttc
                df.loc[df['montant_initial'].isna(), 'montant_initial'] = df.loc[df['montant_initial'].isna(), 'montant_ttc']
            elif 'montant_ttc' in df.columns:
                # Si pas de colonne montant_initial, la créer
                df['montant_initial'] = df['montant_ttc']
        except Exception as e:
            print(f"[WARNING] Erreur lors du traitement de montant_initial: {e}")
            # Fallback: créer une colonne vide
            if 'montant_initial' not in df.columns:
                df['montant_initial'] = 0.0

        # Synchroniser vers le cache MarchesSync
        try:
            sync = MarchesSync()
            stats = sync.sync_from_excel("database_sync", df, force=True)
            print(f"[SYNC CACHE] {stats['nb_inserted']} insérées, {stats['nb_deleted']} supprimées")
        except Exception as e:
            # Si l'erreur est due au fichier inexistant, on peut l'ignorer
            # car on utilise un chemin factice pour la base de données
            print(f"[INFO] Sync avec chemin factice: {e}")

        return len(df)

    def refresh_marches_data(self):
        """Rafraîchit les données de l'onglet Suivi des marchés."""
        # Vérifier si des données existent dans la base
        cur = self.db.conn.cursor()
        cur.execute("SELECT COUNT(*) as count FROM factures")
        nb_factures = cur.fetchone()['count']

        if nb_factures == 0:
            QMessageBox.warning(
                self,
                "Aucune donnée",
                "Aucune facture n'a été importée dans la base de données.\n\n"
                "Veuillez d'abord importer des données en utilisant :\n"
                "• '🔄 Import incrémental' pour importer les fichiers requis."
            )
            return

        try:
            # Synchroniser le cache MarchesSync depuis notre base de données
            nb_synced = self.sync_marches_cache_from_database()
            print(f"[INFO] {nb_synced} factures synchronisées vers le cache marchés")

            # Utiliser un chemin factice pour MarchesAnalyzer
            # (il va charger depuis son cache SQLite qui a été peuplé depuis notre DB)
            fact_path = "database_sync"

            # Initialiser l'analyseur avec la base de données
            self.marches_analyzer = MarchesAnalyzer(fact_path, database=self.db, use_cache=True)

            # Charger les données depuis le cache
            if not self.marches_analyzer.load_data():
                QMessageBox.critical(
                    self,
                    "Erreur de chargement",
                    "Impossible de charger les données.\n\n"
                    "Les données sont dans la base mais l'analyseur de marchés n'a pas pu les charger."
                )
                return

            # Calculer la vision globale
            vision_globale = self.marches_analyzer.get_vision_globale()
            self.marches_global_model.set_data(vision_globale)

            # Calculer la vision détaillée (toutes les tranches)
            vision_detaillee = self.marches_analyzer.get_vision_detaillee()
            self.marches_tranches_model.set_data(vision_detaillee)

            # Calculer la vision opérations
            vision_operations = self.marches_analyzer.get_vision_operations()
            self.operations_model.set_data(vision_operations)

            # Calculer l'historique complet des factures
            historique_factures = self.marches_analyzer.get_historique_factures()
            self.historique_model.set_data(historique_factures)

            # Message de confirmation
            nb_marches = len(vision_globale)
            nb_tranches = len(vision_detaillee)
            nb_operations = len(vision_operations)
            nb_lignes_historique = len(historique_factures)
            QMessageBox.information(
                self,
                "Données actualisées",
                f"✅ Données chargées avec succès !\n\n"
                f"• {nb_marches} marchés analysés\n"
                f"• {nb_tranches} tranches identifiées\n"
                f"• {nb_operations} opérations détectées\n"
                f"• {nb_lignes_historique} lignes d'historique"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Une erreur est survenue lors du chargement des données :\n\n{str(e)}"
            )

    def on_marche_selection_changed(self, selected, deselected):
        """Appelée quand la sélection dans la table globale change."""
        # Récupérer le marché sélectionné
        indexes = self.table_marches_global.selectionModel().selectedRows()

        if not indexes:
            # Aucune sélection : afficher toutes les tranches
            self.marches_tranches_proxy.setMarcheFilter("")
            return

        # Récupérer le marché depuis le modèle source
        source_index = self.marches_global_proxy.mapToSource(indexes[0])
        if source_index.row() < len(self.marches_global_model.rows):
            marche = self.marches_global_model.rows[source_index.row()].get('marche', '')

            # Filtrer les tranches pour ce marché
            self.marches_tranches_proxy.setMarcheFilter(marche)

    def on_marche_double_clicked(self, index):
        """Appelée lors d'un double-clic sur un marché."""
        # Récupérer le marché depuis le modèle source
        source_index = self.marches_global_proxy.mapToSource(index)
        if source_index.row() < len(self.marches_global_model.rows):
            marche_data = self.marches_global_model.rows[source_index.row()]
            code_marche = marche_data.get('marche', '')

            if not code_marche:
                return

            # Ouvrir le dialogue d'édition
            dialog = EditMarcheDialog(self.db, code_marche, marche_data, self)
            if dialog.exec_() == QDialog.Accepted:
                # Rafraîchir les données après modification
                self.refresh_marches_data()

    def on_operation_double_clicked(self, index):
        """Appelée lors d'un double-clic sur une opération."""
        # Récupérer l'opération depuis le modèle source
        source_index = self.operations_proxy.mapToSource(index)
        if source_index.row() < len(self.operations_model.rows):
            operation_data = self.operations_model.rows[source_index.row()]
            code_operation = operation_data.get('operation', '')
            marches = operation_data.get('marches', [])

            if not code_operation:
                return

            # Basculer vers l'onglet Historique
            # Trouver l'index de l'onglet Historique
            for i in range(self.tabs.count()):
                if self.tabs.tabText(i) == "📜 Historique":
                    self.tabs.setCurrentIndex(i)
                    break

            # Si l'opération contient un seul marché, filtrer par ce marché
            # Sinon, filtrer par le code de l'opération (tous les marchés commençant par ce code)
            if isinstance(marches, list) and len(marches) == 1:
                self.edit_filtre_historique.setText(marches[0])
            else:
                self.edit_filtre_historique.setText(code_operation)

    def export_suivi_financier_operation(self):
        """Exporte le suivi financier de l'opération sélectionnée."""
        if not self.marches_analyzer:
            QMessageBox.warning(
                self,
                "Données non chargées",
                "Veuillez d'abord charger les données en cliquant sur 'Actualiser les données'."
            )
            return

        # Récupérer l'opération sélectionnée
        selected_indexes = self.table_operations.selectionModel().selectedRows()
        if not selected_indexes:
            QMessageBox.warning(
                self,
                "Aucune sélection",
                "Veuillez sélectionner une opération à exporter."
            )
            return

        # Récupérer le code de l'opération
        source_index = self.operations_proxy.mapToSource(selected_indexes[0])
        if source_index.row() >= len(self.operations_model.rows):
            return

        operation_data = self.operations_model.rows[source_index.row()]
        code_operation = operation_data.get('operation', '')

        if not code_operation:
            QMessageBox.warning(
                self,
                "Erreur",
                "Impossible de récupérer le code de l'opération."
            )
            return

        exercice_choisi = None

        # Demander le chemin du fichier
        from PyQt5.QtWidgets import QFileDialog
        default_filename = f"suivi_financier_{code_operation}.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter suivi financier opération",
            default_filename,
            "Excel Files (*.xlsx)"
        )

        if not filepath:
            return

        # Exporter
        try:
            success = self.marches_analyzer.export_suivi_financier_operation(
                code_operation,
                filepath,
                exercice_filter=exercice_choisi,
                special_export=False
            )

            if success:
                QMessageBox.information(
                    self,
                    "Export réussi",
                    f"✅ Suivi financier opération {code_operation} exporté !\n\n"
                    f"Fichier : {filepath}\n\n"
                    f"Le fichier contient 2 feuilles :\n"
                    f"• FINANCIER : Détail facture par facture\n"
                    f"• A jour : Vue synthétique par tranche"
                )
            else:
                QMessageBox.critical(
                    self,
                    "Erreur",
                    "Une erreur est survenue lors de l'export.\n\n"
                    "Consultez la console pour plus de détails."
                )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Une erreur est survenue lors de l'export :\n\n{str(e)}"
            )

    def export_suivi_financier_2020_14G3P(self):
        """Export spécifique pour l'opération 2020_14G3P."""
        if not self.marches_analyzer:
            QMessageBox.warning(
                self,
                "Données non chargées",
                "Veuillez d'abord charger les données en cliquant sur 'Actualiser les données'."
            )
            return

        code_operation = "2020_14G3P"
        exercices = self.marches_analyzer.get_exercices_for_operation(code_operation)

        from PyQt5.QtWidgets import QInputDialog, QFileDialog
        exercice_choisi, ok = QInputDialog.getItem(
            self,
            "Choisir l'exercice",
            "Sélectionnez l'exercice à exporter :",
            exercices,
            0,
            False
        )
        if not ok:
            return

        default_filename = f"suivi_financier_{code_operation}.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter suivi financier 2020_14G3P",
            default_filename,
            "Excel Files (*.xlsx)"
        )

        if not filepath:
            return

        try:
            success = self.marches_analyzer.export_suivi_financier_operation(
                code_operation,
                filepath,
                exercice_filter=exercice_choisi,
                special_export=True
            )

            if success:
                QMessageBox.information(
                    self,
                    "Export réussi",
                    f"✅ Suivi financier opération {code_operation} exporté !\n\n"
                    f"Fichier : {filepath}\n\n"
                    f"Le fichier contient 2 feuilles :\n"
                    f"• FINANCIER : Détail facture par facture\n"
                    f"• A jour : Vue synthétique par tranche"
                )
            else:
                QMessageBox.critical(
                    self,
                    "Erreur",
                    "Une erreur est survenue lors de l'export.\n\n"
                    "Consultez la console pour plus de détails."
                )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Une erreur est survenue lors de l'export :\n\n{str(e)}"
            )

    def export_marches_excel(self):
        """Exporte les données des marchés dans un fichier Excel avec 5 feuilles."""
        if not self.marches_analyzer:
            QMessageBox.warning(
                self,
                "Données non chargées",
                "Veuillez d'abord charger les données en cliquant sur 'Actualiser les données'."
            )
            return

        # Demander le chemin du fichier
        from PyQt5.QtWidgets import QFileDialog
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter les données des marchés",
            "export_marches.xlsx",
            "Fichiers Excel (*.xlsx)"
        )

        if not filepath:
            return

        # Exporter
        try:
            success = self.marches_analyzer.export_to_excel(filepath)
            if success:
                QMessageBox.information(
                    self,
                    "Export réussi",
                    f"✅ Les données ont été exportées avec succès !\n\n"
                    f"Fichier : {filepath}\n\n"
                    f"Le fichier contient 5 feuilles :\n"
                    f"• Opérations\n"
                    f"• Marchés\n"
                    f"• Tranches\n"
                    f"• Avenants\n"
                    f"• Historique"
                )
            else:
                QMessageBox.critical(
                    self,
                    "Erreur d'export",
                    "Une erreur est survenue lors de l'export."
                )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur d'export",
                f"Une erreur est survenue lors de l'export :\n\n{str(e)}"
            )

    def open_config(self):
        dlg = ConfigDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            # relance le timer avec le nouvel intervalle
            self.reminder_timer.setInterval(self._get_reminder_interval() * 60 * 1000)

    def closeEvent(self, event):
        if self.error_log:
            try:
                log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "suivi_log.txt")
                with open(log_path, "a", encoding="utf-8") as f:
                    for line in self.error_log:
                        f.write(line + "\n")
                self.error_log.clear()
            except Exception:
                pass
        super().closeEvent(event)


def main():
    app = QApplication(sys.argv)
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), DB_NAME)
    win = MainWindow(db_path)
    # Ouverture directement en plein écran (fenêtre maximisée)
    win.showMaximized()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
