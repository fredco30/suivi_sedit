"""
Module de suivi financier des marchés publics
Calcule la vision globale par marché et le détail par tranche
Avec synchronisation SQLite pour optimiser les performances
"""

import os

import pandas as pd
import numpy as np
from typing import List, Dict, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from marches_sync import MarchesSync
from suivi_financier_agg import (
    Ecriture,
    ResultatSuivi,
    STATUT_ENGAGEMENT,
    STATUT_FACTURE,
    agreger_ecritures,
)


class MarchesAnalyzer:
    """
    Analyseur des données de marchés publics à partir du fichier Excel.
    Utilise un cache SQLite pour améliorer les performances.

    Colonnes Excel utilisées (index 0-based):
    - O (14): Montant initial
    - AD (29): Montant TTC
    - AH (33): Montant service fait
    - AI (34): Date service fait
    - AM (38): Commande
    - E (4): Code mouvement (n° d'engagement, repli du n° de commande)
    - AN (39): Marché
    - AO (40): Tranche
    - AT (45): Mandat
    """

    # Index des colonnes dans le fichier Excel
    COL_MONTANT_INITIAL = 14   # O
    COL_MONTANT_TTC = 29       # AD
    COL_MONTANT_SF = 33        # AH
    COL_DATE_SF = 34           # AI
    COL_COMMANDE = 38          # AM
    COL_MARCHE = 39            # AN
    COL_TRANCHE = 40           # AO
    COL_MANDAT = 45            # AT
    COL_FOURNISSEUR = 8        # I - Nom tiers
    COL_LIBELLE = 13           # N - Libellé
    COL_FACTURE = 36           # AL - Facture
    COL_CODE_MOUVEMENT = 4     # E - Code mouvement (n° d'engagement)

    # Provenance de l'enveloppe initiale affichée dans l'en-tête de l'export
    ENVELOPPE_BASE = "base"
    ENVELOPPE_SEDIT = "sedit"
    ENVELOPPE_ABSENTE = "absente"
    # Du plus fiable au moins fiable : la provenance la plus faible d'un
    # groupe l'emporte sur l'en-tête.
    RANG_PROVENANCE = {ENVELOPPE_BASE: 0, ENVELOPPE_SEDIT: 1, ENVELOPPE_ABSENTE: 2}

    SOURCE_DATABASE = "database_sync"
    EXTENSIONS_EXCEL = (".xls", ".xlsx", ".xlsm")

    def __init__(self, excel_path, database=None, use_cache: bool = True,
                 cache_path: Optional[str] = None):
        """
        Initialise l'analyseur avec la ou les sources de factures.

        Args:
            excel_path: Source des factures. Accepte un chemin de fichier, un
                motif (`data_sources/factures*.xls`), un répertoire, une liste
                de ces formes, ou le mot-clé "database_sync" pour ne lire que
                le cache. Les exports SEDIT étant annuels, le suivi d'une
                opération pluriannuelle se lit sur leur réunion.
            database: Instance de Database pour accéder aux ajustements manuels (optionnel)
            use_cache: Utiliser le cache SQLite (True par défaut)
            cache_path: Fichier de cache SQLite à utiliser. Par défaut celui de
                l'application ; un traitement en lot a intérêt à en nommer un
                autre pour ne pas modifier le cache de travail.
        """
        self.excel_path = excel_path
        self.df = None
        self.df_marches = None
        self.db = database
        self.use_cache = use_cache
        self.cache_path = cache_path
        self.sync = MarchesSync(cache_path) if use_cache and cache_path else (
            MarchesSync() if use_cache else None
        )
        self.sync_stats = None
        self._vision_operations = None

    @classmethod
    def resoudre_sources(cls, excel_path) -> List[str]:
        """Développe une source en liste de fichiers Excel, sans doublon.

        Un répertoire est développé en ses fichiers Excel, un motif en ses
        correspondances ; l'ordre est stable pour que deux chargements
        successifs produisent le même cache.
        """
        import glob as _glob

        entrees = [excel_path] if isinstance(excel_path, str) else list(excel_path or [])
        fichiers: List[str] = []

        for entree in entrees:
            entree = str(entree)
            if entree == cls.SOURCE_DATABASE:
                continue
            if os.path.isdir(entree):
                trouves = [
                    chemin for chemin in sorted(_glob.glob(os.path.join(entree, "*")))
                    if chemin.lower().endswith(cls.EXTENSIONS_EXCEL)
                ]
            elif any(joker in entree for joker in "*?["):
                trouves = sorted(_glob.glob(entree))
            else:
                trouves = [entree]

            for chemin in trouves:
                if chemin not in fichiers:
                    fichiers.append(chemin)

        return fichiers

    @staticmethod
    def extract_exercice_from_bdc(num_commande) -> str:
        """Déduit l'exercice à partir du n° de BDC (2 premiers chiffres)."""
        if not num_commande:
            return "Inconnu"
        num_str = str(num_commande).strip()
        if len(num_str) < 2 or not num_str[:2].isdigit():
            return "Inconnu"
        return f"20{num_str[:2]}"

    def get_exercices_for_operation(self, code_operation: str) -> List[str]:
        """Retourne la liste des exercices disponibles pour une opération."""
        operations_data = self.get_vision_operations()
        operation_info = next((op for op in operations_data if op['operation'] == code_operation), None)
        if not operation_info:
            return ["Tous"]

        marches_operation = operation_info['marches']
        exercices = set()
        for marche in marches_operation:
            df_marche = self.df_marches[self.df_marches.iloc[:, self.COL_MARCHE] == marche]
            for _, row in df_marche.iterrows():
                num_commande = row.iloc[self.COL_COMMANDE] if not pd.isna(row.iloc[self.COL_COMMANDE]) else ""
                exercices.add(self.extract_exercice_from_bdc(num_commande))

        exercices_list = sorted(exercices)
        return ["Tous"] + exercices_list

    def _resoudre_num_bdc(self, df: pd.DataFrame) -> pd.Series:
        """N° de bon de commande de chaque ligne SEDIT.

        La colonne « Commande » (AM) n'est renseignée que lorsque la ligne a été
        rattachée à un bon de commande — elle est vide sur près de 40 % des
        lignes, qui se retrouvaient alors sans BDC et hors de toute agrégation.
        La colonne « Code mouvement » (E) porte le n° d'engagement sur toutes
        les lignes et concorde avec « Commande » partout où les deux sont
        renseignées : elle sert donc de repli.
        """
        commande = df.iloc[:, self.COL_COMMANDE]
        mouvement = df.iloc[:, self.COL_CODE_MOUVEMENT]
        vide = commande.isna() | (commande.astype(str).str.strip() == '')
        return commande.mask(vide, mouvement)

    def _preparer_lignes_a_synchroniser(self, df_excel: pd.DataFrame) -> pd.DataFrame:
        """Extrait d'un export SEDIT les colonnes conservées dans le cache."""
        df_to_sync = df_excel[
            df_excel.iloc[:, self.COL_MARCHE].notna() &
            (df_excel.iloc[:, self.COL_MARCHE] != '')
        ].copy()

        return pd.DataFrame({
            'marche': df_to_sync.iloc[:, self.COL_MARCHE],
            'fournisseur': df_to_sync.iloc[:, self.COL_FOURNISSEUR],
            'libelle': df_to_sync.iloc[:, self.COL_LIBELLE],
            'date_sf': df_to_sync.iloc[:, self.COL_DATE_SF],
            'num_facture': df_to_sync.iloc[:, self.COL_FACTURE],
            'montant_initial': df_to_sync.iloc[:, self.COL_MONTANT_INITIAL],
            'montant_sf': df_to_sync.iloc[:, self.COL_MONTANT_SF],
            'montant_ttc': df_to_sync.iloc[:, self.COL_MONTANT_TTC],
            'num_mandat': df_to_sync.iloc[:, self.COL_MANDAT],
            'tranche': df_to_sync.iloc[:, self.COL_TRANCHE],
            'commande': self._resoudre_num_bdc(df_to_sync),
        })

    def _synchroniser_sources(self, sources: List[str], force_reload: bool) -> Dict:
        """Synchronise chaque fichier source vers le cache et cumule les stats.

        Chaque fichier est traité indépendamment : une ligne présente dans
        plusieurs exports annuels n'est stockée qu'une fois, et la
        synchronisation de l'un ne retire jamais les lignes des autres.
        """
        cumul = {
            'nb_inserted': 0, 'nb_updated': 0, 'nb_unchanged': 0, 'nb_deleted': 0,
            'duration': 0.0, 'status': 'success', 'sources': [], 'nb_sources': len(sources),
        }

        for source in sources:
            needs_sync, reason = self.sync.file_needs_sync(source)
            if not (needs_sync or force_reload):
                print(f"[CACHE] {source} : {reason}")
                cumul['sources'].append({'fichier': source, 'status': 'cached', 'message': reason})
                continue

            print(f"[SYNC] {source} : {reason}")
            try:
                df_excel = pd.read_excel(source)
            except Exception as e:
                print(f"[ERREUR] Lecture impossible de {source} : {e}")
                cumul['status'] = 'partial'
                cumul['sources'].append({'fichier': source, 'status': 'error', 'message': str(e)})
                continue

            stats = self.sync.sync_from_excel(
                source, self._preparer_lignes_a_synchroniser(df_excel), force=force_reload
            )
            stats['fichier'] = source
            cumul['sources'].append(stats)

            if stats.get('status') == 'error':
                cumul['status'] = 'partial'
                continue

            for cle in ('nb_inserted', 'nb_updated', 'nb_unchanged', 'nb_deleted', 'duration'):
                cumul[cle] += stats.get(cle, 0)

        print(f"[OK] Sync terminee sur {len(sources)} fichier(s) : "
              f"{cumul['nb_inserted']} inserees, {cumul['nb_deleted']} supprimees, "
              f"{cumul['nb_unchanged']} inchangees (duree: {cumul['duration']:.2f}s)")
        return cumul

    def load_data(self, force_reload: bool = False):
        """
        Charge les données depuis le ou les fichiers Excel, via le cache SQLite.

        Args:
            force_reload: Force le rechargement depuis Excel même si le cache est valide

        Returns:
            True si succès, False sinon
        """
        # Les données changent : la vision mémorisée n'est plus valable.
        self._vision_operations = None

        try:
            if self.use_cache and self.sync:
                # Si le chemin est "database_sync", charger uniquement depuis le cache
                # (les données ont déjà été synchronisées depuis la base de données)
                if self.excel_path == self.SOURCE_DATABASE:
                    print(f"[CACHE] Chargement depuis le cache (source: database)")
                    self.sync_stats = {'status': 'cached', 'message': 'Données depuis database'}
                else:
                    sources = self.resoudre_sources(self.excel_path)
                    if not sources:
                        print(f"[ERREUR] Aucun fichier source trouvé pour {self.excel_path!r}")
                        return False

                    # Un fichier retiré de la sélection ne doit plus peser sur
                    # le suivi : ses lignes sortent du cache.
                    nb_elaguees = self.sync.prune_sources(sources)
                    if nb_elaguees:
                        print(f"[SYNC] {nb_elaguees} lignes retirées (fichiers hors sélection)")

                    self.sync_stats = self._synchroniser_sources(sources, force_reload)

                # Charger depuis SQLite (beaucoup plus rapide)
                df_from_cache = self.sync.load_to_dataframe()

                # Reconstituer un DataFrame avec la structure Excel (colonnes par index)
                # Créer un DataFrame vide avec 50 colonnes (suffisant pour couvrir tous les index)
                nb_cols = max([
                    self.COL_MONTANT_INITIAL, self.COL_MONTANT_TTC, self.COL_MONTANT_SF,
                    self.COL_DATE_SF, self.COL_COMMANDE, self.COL_MARCHE, self.COL_TRANCHE,
                    self.COL_MANDAT, self.COL_FOURNISSEUR, self.COL_LIBELLE, self.COL_FACTURE
                ]) + 1

                self.df = pd.DataFrame(index=df_from_cache.index, columns=range(nb_cols))

                # Mapper les colonnes du cache vers les index Excel
                self.df.iloc[:, self.COL_MARCHE] = df_from_cache['marche']
                self.df.iloc[:, self.COL_FOURNISSEUR] = df_from_cache['fournisseur']
                self.df.iloc[:, self.COL_LIBELLE] = df_from_cache['libelle']
                self.df.iloc[:, self.COL_DATE_SF] = df_from_cache['date_sf']
                self.df.iloc[:, self.COL_FACTURE] = df_from_cache['num_facture']
                self.df.iloc[:, self.COL_MONTANT_INITIAL] = df_from_cache['montant_initial']
                self.df.iloc[:, self.COL_MONTANT_SF] = df_from_cache['montant_sf']
                self.df.iloc[:, self.COL_MONTANT_TTC] = df_from_cache['montant_ttc']
                self.df.iloc[:, self.COL_MANDAT] = df_from_cache['num_mandat']
                self.df.iloc[:, self.COL_TRANCHE] = df_from_cache['tranche']
                self.df.iloc[:, self.COL_COMMANDE] = df_from_cache['commande']

                # df_marches pointe déjà vers le DataFrame filtré
                self.df_marches = self.df.copy()

            else:
                # Mode sans cache (chargement direct depuis Excel)
                print("📂 Chargement direct depuis Excel (cache désactivé)")
                sources = self.resoudre_sources(self.excel_path)
                if not sources:
                    print(f"[ERREUR] Aucun fichier source trouvé pour {self.excel_path!r}")
                    return False
                self.df = pd.concat(
                    [pd.read_excel(source) for source in sources], ignore_index=True
                ).drop_duplicates()
                self.df_marches = self.df[
                    self.df.iloc[:, self.COL_MARCHE].notna() &
                    (self.df.iloc[:, self.COL_MARCHE] != '')
                ].copy()
                self.df_marches.iloc[:, self.COL_COMMANDE] = \
                    self._resoudre_num_bdc(self.df_marches)

            return True

        except Exception as e:
            print(f"[ERREUR] Erreur lors du chargement: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _get_col_value(self, row, col_idx, default=None):
        """Récupère une valeur de colonne de manière sécurisée."""
        try:
            value = row.iloc[col_idx]
            if pd.isna(value) or value == '':
                return default
            return value
        except:
            return default

    def calculate_montant_initial_tranche(self, marche: str, tranche) -> float:
        """
        Calcule le montant initial d'une tranche.

        Priorité :
        1. Pour la TF : Utiliser montant_initial_manuel du marché
        2. Pour les TOs : Utiliser les tranches définies dans la base de données (table tranches)
        3. Sinon, calculer depuis Excel (ancien comportement)

        Règle métier (calcul Excel) :
        - Prendre Montant initial (O)
        - Ignorer les 0,00
        - Dédupliquer les valeurs positives (souvent répétées)
        - Additionner ces valeurs positives distinctes
        """
        # Essayer d'abord de charger depuis la base de données
        if self.db:
            # Normaliser la tranche en valeur numérique
            tranche_num = None
            if not pd.isna(tranche):
                try:
                    tranche_num = float(tranche)
                except (ValueError, TypeError):
                    pass

            # Convertir le numéro de tranche en code (0=TF, 1=TO1, 2=TO2, etc.)
            if pd.isna(tranche) or tranche_num == 0:
                code_tranche = "TF"
                # Pour la TF, utiliser montant_initial_manuel du marché
                marche_record = self.db.get_marche(marche)
                if marche_record and marche_record["montant_initial_manuel"]:
                    montant = float(marche_record["montant_initial_manuel"])
                    print(f"[TRANCHE] {marche} TF: {montant} € (depuis montant_initial_manuel)")
                    return montant
            else:
                # Pour les TOs (TO1, TO2, etc.)
                code_tranche = f"TO{int(tranche_num)}" if tranche_num else str(tranche)

                # Pour les TOs, chercher dans la table tranches
                cur = self.db.conn.cursor()
                cur.execute(
                    "SELECT montant FROM tranches WHERE code_marche = ? AND code_tranche = ?",
                    (marche, code_tranche)
                )
                row = cur.fetchone()
                if row and row["montant"]:
                    print(f"[TRANCHE] {marche} {code_tranche}: {row['montant']} € (depuis DB)")
                    return float(row["montant"])

        # Fallback : calculer depuis Excel (ancien comportement)
        # Filtrer les lignes pour ce marché et cette tranche
        mask = (self.df_marches.iloc[:, self.COL_MARCHE] == marche)

        # Gérer le cas où tranche peut être NaN, None ou une valeur
        if pd.isna(tranche):
            mask &= self.df_marches.iloc[:, self.COL_TRANCHE].isna()
        else:
            mask &= (self.df_marches.iloc[:, self.COL_TRANCHE] == tranche)

        df_tranche = self.df_marches[mask]

        if len(df_tranche) == 0:
            return 0.0

        # Récupérer les montants initiaux
        montants = df_tranche.iloc[:, self.COL_MONTANT_INITIAL].copy()

        # Convertir en float et ignorer les NaN
        montants = pd.to_numeric(montants, errors='coerce').fillna(0.0)

        # Ignorer les 0,00
        montants = montants[montants > 0.01]

        # Dédupliquer les valeurs
        montants_uniques = montants.unique()

        # Additionner
        montant_calcule = float(montants_uniques.sum())
        if montant_calcule > 0:
            print(f"[TRANCHE] {marche} tranche {tranche}: {montant_calcule} € (calculé depuis Excel)")
        return montant_calcule

    def calculate_service_fait_tranche(self, marche: str, tranche) -> float:
        """
        Calcule le service fait cumulé pour une tranche.

        Règle : somme de AH pour ce couple (marché, tranche)
        """
        mask = (self.df_marches.iloc[:, self.COL_MARCHE] == marche)

        if pd.isna(tranche):
            mask &= self.df_marches.iloc[:, self.COL_TRANCHE].isna()
        else:
            mask &= (self.df_marches.iloc[:, self.COL_TRANCHE] == tranche)

        df_tranche = self.df_marches[mask]

        if len(df_tranche) == 0:
            return 0.0

        sf = df_tranche.iloc[:, self.COL_MONTANT_SF].copy()
        sf = pd.to_numeric(sf, errors='coerce').fillna(0.0)

        return float(sf.sum())

    def calculate_paye_tranche(self, marche: str, tranche) -> float:
        """
        Calcule le payé cumulé pour une tranche.

        Règle : somme de AD pour ce couple (marché, tranche) avec AT (mandat) non vide
        """
        mask = (self.df_marches.iloc[:, self.COL_MARCHE] == marche)

        if pd.isna(tranche):
            mask &= self.df_marches.iloc[:, self.COL_TRANCHE].isna()
        else:
            mask &= (self.df_marches.iloc[:, self.COL_TRANCHE] == tranche)

        # Ajouter le filtre sur le mandat (doit être renseigné)
        mask &= self.df_marches.iloc[:, self.COL_MANDAT].notna()
        mask &= (self.df_marches.iloc[:, self.COL_MANDAT] != '')

        df_tranche = self.df_marches[mask]

        if len(df_tranche) == 0:
            return 0.0

        ttc = df_tranche.iloc[:, self.COL_MONTANT_TTC].copy()
        ttc = pd.to_numeric(ttc, errors='coerce').fillna(0.0)

        return float(ttc.sum())

    def get_vision_detaillee(self) -> List[Dict]:
        """
        Retourne la vision détaillée : 1 ligne par couple (marché, tranche).

        Colonnes retournées :
        - marche
        - tranche (affiché comme TF, TO1, TO2... ou la valeur brute)
        - montant_initial_tranche
        - service_fait_tranche
        - paye_tranche
        - pourcent_consomme_tranche
        """
        if self.df_marches is None or len(self.df_marches) == 0:
            return []

        results = []

        # Grouper par (marché, tranche)
        grouped = self.df_marches.groupby(
            [self.df_marches.iloc[:, self.COL_MARCHE],
             self.df_marches.iloc[:, self.COL_TRANCHE]]
        )

        for (marche, tranche), _ in grouped:
            montant_initial = self.calculate_montant_initial_tranche(marche, tranche)
            service_fait = self.calculate_service_fait_tranche(marche, tranche)
            paye = self.calculate_paye_tranche(marche, tranche)

            # Calculer le pourcentage
            pourcent = 0.0
            if montant_initial > 0:
                pourcent = (service_fait / montant_initial) * 100

            # Formater le libellé de la tranche
            if pd.isna(tranche):
                tranche_libelle = "Sans tranche"
            else:
                # Si c'est un nombre, formater en TF, TO1, TO2...
                try:
                    tranche_num = int(float(tranche))
                    if tranche_num == 0:
                        tranche_libelle = "TF"
                    else:
                        tranche_libelle = f"TO{tranche_num}"
                except:
                    tranche_libelle = str(tranche)

            results.append({
                'marche': marche,
                'tranche': tranche,
                'tranche_libelle': tranche_libelle,
                'montant_initial_tranche': montant_initial,
                'service_fait_tranche': service_fait,
                'paye_tranche': paye,
                'pourcent_consomme_tranche': pourcent
            })

        # Trier par marché puis tranche
        results.sort(key=lambda x: (x['marche'], x['tranche'] if not pd.isna(x['tranche']) else -1))

        return results

    def get_vision_globale(self) -> List[Dict]:
        """
        Retourne la vision globale : 1 ligne par marché.

        Colonnes retournées :
        - marche
        - libelle_marche
        - fournisseur
        - montant_initial_marche
        - service_fait_cumule
        - paye_cumule
        - reste_a_realiser
        - reste_a_mandater
        - pourcent_consomme
        """
        if self.df_marches is None or len(self.df_marches) == 0:
            return []

        results = []

        # Grouper par marché
        marches_uniques = self.df_marches.iloc[:, self.COL_MARCHE].unique()

        for marche in marches_uniques:
            df_marche = self.df_marches[self.df_marches.iloc[:, self.COL_MARCHE] == marche]

            # Récupérer le fournisseur (prendre le premier non vide)
            fournisseurs = df_marche.iloc[:, self.COL_FOURNISSEUR].dropna()
            fournisseur = fournisseurs.iloc[0] if len(fournisseurs) > 0 else ""

            # Récupérer le libellé (prendre le premier non vide)
            libelles = df_marche.iloc[:, self.COL_LIBELLE].dropna()
            libelle = libelles.iloc[0] if len(libelles) > 0 else ""

            # Calculer le montant initial du marché
            # Priorité 1: Montant manuel + avenants depuis la BD
            # Priorité 2: Calcul automatique depuis Excel (somme des tranches)
            montant_initial_marche = 0.0
            montant_excel = 0.0
            nb_avenants = 0

            if self.db:
                # Essayer de récupérer le montant depuis la base de données
                montant_bd = self.db.get_montant_total_marche(marche)
                if montant_bd > 0:
                    montant_initial_marche = montant_bd

                    # Récupérer le nombre d'avenants pour affichage
                    avenants = self.db.get_avenants(marche)
                    nb_avenants = len(avenants) if avenants else 0

            # Si pas de montant manuel en BD, calculer depuis Excel
            if montant_initial_marche == 0:
                tranches = df_marche.iloc[:, self.COL_TRANCHE].unique()
                montant_excel = sum(
                    self.calculate_montant_initial_tranche(marche, t)
                    for t in tranches
                )
                montant_initial_marche = montant_excel

            # Service fait cumulé (somme de AH pour tout le marché)
            sf = df_marche.iloc[:, self.COL_MONTANT_SF].copy()
            sf = pd.to_numeric(sf, errors='coerce').fillna(0.0)
            service_fait_cumule = float(sf.sum())

            # Payé cumulé (somme de AD avec mandat non vide)
            df_mandates = df_marche[
                df_marche.iloc[:, self.COL_MANDAT].notna() &
                (df_marche.iloc[:, self.COL_MANDAT] != '')
            ]
            ttc = df_mandates.iloc[:, self.COL_MONTANT_TTC].copy()
            ttc = pd.to_numeric(ttc, errors='coerce').fillna(0.0)
            paye_cumule = float(ttc.sum())

            # Calculs dérivés
            reste_a_realiser = montant_initial_marche - service_fait_cumule
            reste_a_mandater = service_fait_cumule - paye_cumule

            pourcent_consomme = 0.0
            if montant_initial_marche > 0:
                pourcent_consomme = (service_fait_cumule / montant_initial_marche) * 100

            # Extraire le code opération
            code_operation = self.extract_operation(marche)

            results.append({
                'marche': marche,
                'operation': code_operation,
                'libelle_marche': libelle,
                'fournisseur': fournisseur,
                'montant_initial_marche': montant_initial_marche,
                'service_fait_cumule': service_fait_cumule,
                'paye_cumule': paye_cumule,
                'reste_a_realiser': reste_a_realiser,
                'reste_a_mandater': reste_a_mandater,
                'pourcent_consomme': pourcent_consomme,
                'nb_avenants': nb_avenants,
                'montant_excel': montant_excel
            })

        # Trier par marché
        results.sort(key=lambda x: x['marche'])

        return results

    def get_tranches_for_marche(self, marche: str) -> List[Dict]:
        """
        Retourne le détail des tranches pour un marché donné.
        """
        vision_detaillee = self.get_vision_detaillee()
        return [t for t in vision_detaillee if t['marche'] == marche]

    @staticmethod
    def extract_operation(marche: str) -> str:
        """
        Extrait le code opération d'un code marché.

        Règles :
        - Si >= 2 séparateurs ET dernier segment = petit numéro → c'est un LOT
          Exemples: 2024_17_1 → 2024_17 | 2024_1_3 → 2024_1
        - Si 1 seul séparateur → c'est une OPÉRATION INDÉPENDANTE (pas de regroupement)
          Exemples: 2025_12 → 2025_12 | 2023_17 → 2023_17
        - Cas spéciaux restent inchangés
          Exemples: 2020_14G3P → 2020_14G3P
        """
        import re

        if not marche:
            return ""

        normalized = str(marche).strip()

        # Compter les séparateurs (_ et -)
        nb_underscores = normalized.count('_')
        nb_dashes = normalized.count('-')
        total_separators = nb_underscores + nb_dashes

        # Si moins de 2 séparateurs → le marché EST l'opération (pas de regroupement)
        if total_separators < 2:
            return normalized

        # Si >= 2 séparateurs, vérifier si le dernier segment est un numéro de lot
        last_underscore = normalized.rfind('_')
        last_dash = normalized.rfind('-')
        last_sep = max(last_underscore, last_dash)

        if last_sep > 0:
            suffix = normalized[last_sep+1:]

            # Vérifier si c'est un petit numéro (1-2 chiffres) = probable numéro de lot
            if suffix.isdigit() and len(suffix) <= 2:
                # C'est un lot, extraire l'opération
                return normalized[:last_sep]

        # Sinon, le marché est l'opération complète
        return normalized

    def invalider_vision(self):
        """Oublie la vision par opération mémorisée.

        À appeler après une écriture en base qui change les montants des
        marchés (enveloppes, avenants, tranches).
        """
        self._vision_operations = None

    def get_vision_operations(self, force_refresh: bool = False) -> List[Dict]:
        """
        Retourne la vision par opération (regroupement de marchés/lots).

        Le résultat est mémorisé : il est recalculé une fois par opération lors
        d'un traitement en lot, ce qui coûtait 27 s pour 77 opérations.
        `invalider_vision()` ou un `load_data()` le remet à zéro.

        Retourne :
        - operation : code opération
        - nb_lots : nombre de marchés/lots
        - marchés : liste des codes marchés
        - montant_initial_total
        - service_fait_total
        - paye_total
        - reste_a_realiser
        - reste_a_mandater
        - pourcent_consomme
        - nb_avenants_total
        """
        if not force_refresh and self._vision_operations is not None:
            return self._vision_operations

        vision_marches = self.get_vision_globale()

        if not vision_marches:
            return []

        # Regrouper par opération
        operations_dict = {}

        for marche_data in vision_marches:
            operation = self.extract_operation(marche_data['marche'])

            if operation not in operations_dict:
                operations_dict[operation] = {
                    'operation': operation,
                    'marches': [],
                    'libelles': [],
                    'fournisseurs': [],
                    'montant_initial_total': 0.0,
                    'service_fait_total': 0.0,
                    'paye_total': 0.0,
                    'nb_avenants_total': 0
                }

            op = operations_dict[operation]
            op['marches'].append(marche_data['marche'])

            # Ajouter le libellé s'il n'est pas déjà présent
            libelle = marche_data.get('libelle_marche', '')
            if libelle and libelle not in op['libelles']:
                op['libelles'].append(libelle)

            # Ajouter le fournisseur s'il n'est pas déjà présent
            fournisseur = marche_data.get('fournisseur', '')
            if fournisseur and fournisseur not in op['fournisseurs']:
                op['fournisseurs'].append(fournisseur)

            op['montant_initial_total'] += marche_data['montant_initial_marche']
            op['service_fait_total'] += marche_data['service_fait_cumule']
            op['paye_total'] += marche_data['paye_cumule']
            op['nb_avenants_total'] += marche_data['nb_avenants']

        # Calculer les résultats finaux
        results = []
        for operation, data in operations_dict.items():
            nb_lots = len(data['marches'])
            montant_initial = data['montant_initial_total']
            sf_total = data['service_fait_total']
            paye_total = data['paye_total']

            reste_a_realiser = montant_initial - sf_total
            reste_a_mandater = sf_total - paye_total

            pourcent = 0.0
            if montant_initial > 0:
                pourcent = (sf_total / montant_initial) * 100

            # Combiner les libellés et fournisseurs
            libelle_combined = " | ".join(data['libelles']) if data['libelles'] else ""
            fournisseur_combined = " | ".join(data['fournisseurs']) if data['fournisseurs'] else ""

            results.append({
                'operation': operation,
                'nb_lots': nb_lots,
                'marches': data['marches'],
                'libelle': libelle_combined,
                'fournisseur': fournisseur_combined,
                'montant_initial_total': montant_initial,
                'service_fait_total': sf_total,
                'paye_total': paye_total,
                'reste_a_realiser': reste_a_realiser,
                'reste_a_mandater': reste_a_mandater,
                'pourcent_consomme': pourcent,
                'nb_avenants_total': data['nb_avenants_total']
            })

        # Trier par opération
        results.sort(key=lambda x: x['operation'])

        self._vision_operations = results
        return results

    def get_historique_factures(self, marche: str = None) -> List[Dict]:
        """
        Retourne l'historique détaillé des factures/paiements.

        Args:
            marche: Si spécifié, retourne uniquement l'historique de ce marché.
                   Si None, retourne l'historique complet de tous les marchés.

        Retourne une liste de dictionnaires avec :
        - marche
        - date_sf
        - num_facture
        - libelle
        - montant_sf
        - montant_ttc
        - num_mandat
        - statut (Payé / Service fait / Facturée / En attente)
        """
        if self.df_marches is None or len(self.df_marches) == 0:
            return []

        # Filtrer par marché si spécifié
        if marche:
            df_filtered = self.df_marches[self.df_marches.iloc[:, self.COL_MARCHE] == marche]
        else:
            df_filtered = self.df_marches

        results = []

        for idx, row in df_filtered.iterrows():
            marche_code = row.iloc[self.COL_MARCHE]
            fournisseur = row.iloc[self.COL_FOURNISSEUR]
            date_sf = row.iloc[self.COL_DATE_SF]
            num_facture = row.iloc[self.COL_FACTURE]
            libelle = row.iloc[self.COL_LIBELLE]
            montant_sf = row.iloc[self.COL_MONTANT_SF]
            montant_ttc = row.iloc[self.COL_MONTANT_TTC]
            num_mandat = row.iloc[self.COL_MANDAT]

            # Convertir les valeurs
            try:
                montant_sf = float(montant_sf) if pd.notna(montant_sf) else 0.0
            except:
                montant_sf = 0.0

            try:
                montant_ttc = float(montant_ttc) if pd.notna(montant_ttc) else 0.0
            except:
                montant_ttc = 0.0

            # Formater la date
            date_sf_str = ""
            if pd.notna(date_sf):
                try:
                    if isinstance(date_sf, str):
                        date_sf_str = date_sf
                    else:
                        date_sf_str = date_sf.strftime("%d/%m/%Y")
                except:
                    date_sf_str = str(date_sf)

            # Déterminer le statut
            has_mandat = pd.notna(num_mandat) and str(num_mandat).strip() != ''
            has_facture = pd.notna(num_facture) and str(num_facture).strip() != ''
            has_sf = montant_sf > 0.01

            if has_mandat:
                statut = "✅ Payé"
            elif has_sf:
                statut = "⏳ Service fait"
            elif has_facture:
                statut = "📋 Facturée"
            else:
                statut = "⚠️ En attente"

            results.append({
                'marche': marche_code,
                'fournisseur': str(fournisseur) if pd.notna(fournisseur) else "",
                'date_sf': date_sf_str,
                'num_facture': str(num_facture) if pd.notna(num_facture) else "",
                'libelle': str(libelle) if pd.notna(libelle) else "",
                'montant_sf': montant_sf,
                'montant_ttc': montant_ttc,
                'num_mandat': str(num_mandat) if pd.notna(num_mandat) else "",
                'statut': statut
            })

        # Trier par marché puis par date
        results.sort(key=lambda x: (x['marche'], x['date_sf']), reverse=True)

        return results

    def export_to_excel(self, filepath: str) -> bool:
        """
        Exporte toutes les données dans un fichier Excel avec 5 feuilles :
        1. Opérations - Vision par opération
        2. Marchés - Vision globale par marché
        3. Tranches - Détail par tranche
        4. Avenants - Liste de tous les avenants
        5. Historique - Historique complet des factures/paiements

        Returns:
            bool: True si succès, False sinon
        """
        try:
            wb = Workbook()

            # Styles communs
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            border_thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # ===============================
            # FEUILLE 1 : OPÉRATIONS
            # ===============================
            ws_operations = wb.active
            ws_operations.title = "Opérations"

            vision_operations = self.get_vision_operations()

            # En-têtes
            headers_operations = [
                "Opération", "Nb lots", "Marchés", "Libellé", "Fournisseur",
                "Montant initial total", "Avenants", "Service fait total", "Payé total",
                "Reste à réaliser", "Reste à mandater", "% consommé"
            ]

            for col_idx, header in enumerate(headers_operations, 1):
                cell = ws_operations.cell(1, col_idx, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            # Données
            for row_idx, op in enumerate(vision_operations, 2):
                ws_operations.cell(row_idx, 1, op.get('operation', ''))
                ws_operations.cell(row_idx, 2, op.get('nb_lots', 0))
                marches_str = ", ".join(op.get('marches', [])) if isinstance(op.get('marches'), list) else str(op.get('marches', ''))
                ws_operations.cell(row_idx, 3, marches_str)
                ws_operations.cell(row_idx, 4, op.get('libelle', ''))
                ws_operations.cell(row_idx, 5, op.get('fournisseur', ''))
                ws_operations.cell(row_idx, 6, op.get('montant_initial_total', 0))
                ws_operations.cell(row_idx, 7, op.get('nb_avenants_total', 0))
                ws_operations.cell(row_idx, 8, op.get('service_fait_total', 0))
                ws_operations.cell(row_idx, 9, op.get('paye_total', 0))
                ws_operations.cell(row_idx, 10, op.get('reste_a_realiser', 0))
                ws_operations.cell(row_idx, 11, op.get('reste_a_mandater', 0))
                ws_operations.cell(row_idx, 12, op.get('pourcent_consomme', 0) / 100)

                # Format numérique
                for col in [6, 8, 9, 10, 11]:
                    ws_operations.cell(row_idx, col).number_format = '#,##0.00 €'
                ws_operations.cell(row_idx, 12).number_format = '0.00%'

            # Ajuster les largeurs
            ws_operations.column_dimensions['A'].width = 20
            ws_operations.column_dimensions['B'].width = 10
            ws_operations.column_dimensions['C'].width = 40
            ws_operations.column_dimensions['D'].width = 40  # Libellé
            ws_operations.column_dimensions['E'].width = 30  # Fournisseur
            for col in ['F', 'G', 'H', 'I', 'J', 'K']:
                ws_operations.column_dimensions[col].width = 18
            ws_operations.column_dimensions['L'].width = 12

            # ===============================
            # FEUILLE 2 : MARCHÉS
            # ===============================
            ws_marches = wb.create_sheet("Marchés")

            vision_globale = self.get_vision_globale()

            # En-têtes
            headers_marches = [
                "Marché", "Opération", "Libellé", "Fournisseur", "Montant initial",
                "Avenants", "Service fait cumulé", "Payé cumulé",
                "Reste à réaliser", "Reste à mandater", "% consommé"
            ]

            for col_idx, header in enumerate(headers_marches, 1):
                cell = ws_marches.cell(1, col_idx, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            # Données
            for row_idx, marche in enumerate(vision_globale, 2):
                ws_marches.cell(row_idx, 1, marche.get('marche', ''))
                ws_marches.cell(row_idx, 2, marche.get('operation', ''))
                ws_marches.cell(row_idx, 3, marche.get('libelle_marche', ''))
                ws_marches.cell(row_idx, 4, marche.get('fournisseur', ''))
                ws_marches.cell(row_idx, 5, marche.get('montant_initial_marche', 0))
                ws_marches.cell(row_idx, 6, marche.get('nb_avenants', 0))
                ws_marches.cell(row_idx, 7, marche.get('service_fait_cumule', 0))
                ws_marches.cell(row_idx, 8, marche.get('paye_cumule', 0))
                ws_marches.cell(row_idx, 9, marche.get('reste_a_realiser', 0))
                ws_marches.cell(row_idx, 10, marche.get('reste_a_mandater', 0))
                ws_marches.cell(row_idx, 11, marche.get('pourcent_consomme', 0) / 100)

                # Format numérique
                for col in [5, 7, 8, 9, 10]:
                    ws_marches.cell(row_idx, col).number_format = '#,##0.00 €'
                ws_marches.cell(row_idx, 11).number_format = '0.00%'

            # Ajuster les largeurs
            ws_marches.column_dimensions['A'].width = 20
            ws_marches.column_dimensions['B'].width = 18
            ws_marches.column_dimensions['C'].width = 40
            ws_marches.column_dimensions['D'].width = 30
            for col in ['E', 'F', 'G', 'H', 'I', 'J']:
                ws_marches.column_dimensions[col].width = 18
            ws_marches.column_dimensions['K'].width = 12

            # ===============================
            # FEUILLE 3 : TRANCHES
            # ===============================
            ws_tranches = wb.create_sheet("Tranches")

            vision_detaillee = self.get_vision_detaillee()

            # En-têtes
            headers_tranches = [
                "Marché", "Tranche", "Montant initial tranche",
                "Service fait tranche", "Payé tranche", "% consommé"
            ]

            for col_idx, header in enumerate(headers_tranches, 1):
                cell = ws_tranches.cell(1, col_idx, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            # Données
            for row_idx, tranche in enumerate(vision_detaillee, 2):
                ws_tranches.cell(row_idx, 1, tranche.get('marche', ''))
                ws_tranches.cell(row_idx, 2, tranche.get('tranche_libelle', ''))
                ws_tranches.cell(row_idx, 3, tranche.get('montant_initial_tranche', 0))
                ws_tranches.cell(row_idx, 4, tranche.get('service_fait_tranche', 0))
                ws_tranches.cell(row_idx, 5, tranche.get('paye_tranche', 0))
                ws_tranches.cell(row_idx, 6, tranche.get('pourcent_consomme_tranche', 0) / 100)

                # Format numérique
                for col in [3, 4, 5]:
                    ws_tranches.cell(row_idx, col).number_format = '#,##0.00 €'
                ws_tranches.cell(row_idx, 6).number_format = '0.00%'

            # Ajuster les largeurs
            ws_tranches.column_dimensions['A'].width = 20
            ws_tranches.column_dimensions['B'].width = 15
            for col in ['C', 'D', 'E']:
                ws_tranches.column_dimensions[col].width = 18
            ws_tranches.column_dimensions['F'].width = 12

            # ===============================
            # FEUILLE 4 : AVENANTS
            # ===============================
            ws_avenants = wb.create_sheet("Avenants")

            # Récupérer tous les avenants depuis la base de données
            avenants_list = []
            if self.db:
                # Récupérer tous les marchés
                marches_uniques = self.df_marches.iloc[:, self.COL_MARCHE].unique()
                for marche in marches_uniques:
                    avenants = self.db.get_avenants(marche)
                    if avenants:
                        for avenant in avenants:
                            avenants_list.append({
                                'marche': marche,
                                'numero': avenant.get('numero_avenant', ''),
                                'libelle': avenant.get('libelle', ''),
                                'montant': avenant.get('montant', 0),
                                'type': avenant.get('type_modification', ''),
                                'date': avenant.get('date_avenant', ''),
                                'motif': avenant.get('motif', '')
                            })

            # En-têtes
            headers_avenants = [
                "Marché", "N° Avenant", "Libellé", "Montant",
                "Type", "Date", "Motif"
            ]

            for col_idx, header in enumerate(headers_avenants, 1):
                cell = ws_avenants.cell(1, col_idx, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            # Données
            for row_idx, avenant in enumerate(avenants_list, 2):
                ws_avenants.cell(row_idx, 1, avenant.get('marche', ''))
                ws_avenants.cell(row_idx, 2, avenant.get('numero', ''))
                ws_avenants.cell(row_idx, 3, avenant.get('libelle', ''))
                ws_avenants.cell(row_idx, 4, avenant.get('montant', 0))
                ws_avenants.cell(row_idx, 5, avenant.get('type', ''))
                ws_avenants.cell(row_idx, 6, avenant.get('date', ''))
                ws_avenants.cell(row_idx, 7, avenant.get('motif', ''))

                # Format numérique
                ws_avenants.cell(row_idx, 4).number_format = '#,##0.00 €'

            # Ajuster les largeurs
            ws_avenants.column_dimensions['A'].width = 20
            ws_avenants.column_dimensions['B'].width = 12
            ws_avenants.column_dimensions['C'].width = 35
            ws_avenants.column_dimensions['D'].width = 15
            ws_avenants.column_dimensions['E'].width = 15
            ws_avenants.column_dimensions['F'].width = 12
            ws_avenants.column_dimensions['G'].width = 40

            # ===============================
            # FEUILLE 5 : HISTORIQUE
            # ===============================
            ws_historique = wb.create_sheet("Historique")

            historique = self.get_historique_factures()

            # En-têtes
            headers_historique = [
                "Marché", "Fournisseur", "Date SF", "N° Facture", "Libellé",
                "Montant SF", "Montant TTC", "N° Mandat", "Statut"
            ]

            for col_idx, header in enumerate(headers_historique, 1):
                cell = ws_historique.cell(1, col_idx, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            # Données
            for row_idx, ligne in enumerate(historique, 2):
                ws_historique.cell(row_idx, 1, ligne.get('marche', ''))
                ws_historique.cell(row_idx, 2, ligne.get('fournisseur', ''))
                ws_historique.cell(row_idx, 3, ligne.get('date_sf', ''))
                ws_historique.cell(row_idx, 4, ligne.get('num_facture', ''))
                ws_historique.cell(row_idx, 5, ligne.get('libelle', ''))
                ws_historique.cell(row_idx, 6, ligne.get('montant_sf', 0))
                ws_historique.cell(row_idx, 7, ligne.get('montant_ttc', 0))
                ws_historique.cell(row_idx, 8, ligne.get('num_mandat', ''))
                ws_historique.cell(row_idx, 9, ligne.get('statut', ''))

                # Format numérique
                for col in [6, 7]:
                    ws_historique.cell(row_idx, col).number_format = '#,##0.00 €'

            # Ajuster les largeurs
            ws_historique.column_dimensions['A'].width = 20  # Marché
            ws_historique.column_dimensions['B'].width = 30  # Fournisseur
            ws_historique.column_dimensions['C'].width = 12  # Date SF
            ws_historique.column_dimensions['D'].width = 15  # N° Facture
            ws_historique.column_dimensions['E'].width = 40  # Libellé
            ws_historique.column_dimensions['F'].width = 15  # Montant SF
            ws_historique.column_dimensions['G'].width = 15  # Montant TTC
            ws_historique.column_dimensions['H'].width = 15  # N° Mandat
            ws_historique.column_dimensions['I'].width = 20  # Statut

            # Sauvegarder le fichier
            wb.save(filepath)
            return True

        except Exception as e:
            print(f"Erreur lors de l'export Excel : {e}")
            return False

    def _enveloppe_tranche(self, marche: str, tranche_libelle: str,
                           montant_tranche: float,
                           marches_totaux: Dict[str, float]) -> Tuple[float, str]:
        """Enveloppe initiale d'un couple (marché, tranche), et sa provenance.

        L'enveloppe est un paramètre : elle est lue en base (montant initial du
        marché, tranches et avenants) et jamais reconstituée depuis les lignes
        du fichier produit. Quand le marché n'est pas renseigné en base, on
        retombe sur les montants initiaux de l'export SEDIT — la provenance est
        alors signalée en clair dans l'en-tête, car le montant reste à saisir.

        Args:
            montant_tranche: montant initial de la tranche, déjà calculé.

        Returns:
            (montant, provenance) où provenance vaut ENVELOPPE_BASE,
            ENVELOPPE_SEDIT ou ENVELOPPE_ABSENTE.
        """
        montant_marche = float(marches_totaux.get(marche, 0) or 0)
        configure_en_base = montant_marche > 0

        if tranche_libelle and montant_tranche > 0:
            return montant_tranche, (
                self.ENVELOPPE_BASE if configure_en_base else self.ENVELOPPE_SEDIT
            )
        if configure_en_base:
            return montant_marche, self.ENVELOPPE_BASE
        if montant_tranche > 0:
            return montant_tranche, self.ENVELOPPE_SEDIT
        return 0.0, self.ENVELOPPE_ABSENTE

    def collecter_ecritures_operation(
        self,
        code_operation: str,
        exercice_filter: Optional[str] = None
    ) -> Tuple[Dict, Dict, Dict, Dict, Optional[Dict]]:
        """Collecte les écritures SEDIT d'une opération, groupées par prestataire/tranche.

        Source unique des onglets FINANCIER et « A jour » : les deux vues sont
        produites à partir de ce même jeu de données, jamais de deux requêtes
        indépendantes (c'était la cause de l'écart entre les deux onglets).

        Returns:
            (groupes, montants_declares, enveloppes, provenances, operation_info)
            où `groupes` associe (fournisseur, tranche_libelle) à une liste
            d'`Ecriture` et `provenances` la source de chaque enveloppe.
        """
        from collections import OrderedDict

        operations_data = self.get_vision_operations()
        operation_info = next(
            (op for op in operations_data if op['operation'] == code_operation), None
        )
        if not operation_info:
            print(f"Opération {code_operation} non trouvée")
            return {}, {}, {}, {}, None

        marches_operation = operation_info['marches']

        # Montant de BDC déclaré, une valeur par n° de commande.
        montants_declares: Dict[str, float] = {}
        marches_types: Dict[str, str] = {}
        if self.db:
            cur = self.db.conn.cursor()
            placeholders = ','.join('?' * len(marches_operation))
            cur.execute(
                f"""SELECT num_commande, SUM(montant_ttc) AS montant_total
                    FROM commandes WHERE marche IN ({placeholders})
                    GROUP BY num_commande""",
                marches_operation
            )
            for row in cur.fetchall():
                if row['num_commande']:
                    montants_declares[str(row['num_commande']).strip()] = row['montant_total']

            cur.execute(
                f"""SELECT code_marche, type_marche FROM marches
                    WHERE code_marche IN ({placeholders})""",
                marches_operation
            )
            for row in cur.fetchall():
                try:
                    type_marche = row['type_marche'] if row['type_marche'] else 'CLASSIQUE'
                except (KeyError, IndexError):
                    type_marche = 'CLASSIQUE'
                marches_types[row['code_marche']] = type_marche

        # Enveloppe initiale de chaque marché, lue en base.
        marches_totaux: Dict[str, float] = {}
        if self.db:
            for marche in marches_operation:
                marches_totaux[marche] = self.db.get_montant_total_marche(marche)

        groupes: "OrderedDict[Tuple[str, str], List[Ecriture]]" = OrderedDict()
        enveloppes: Dict[Tuple[str, str], float] = {}
        provenances: Dict[Tuple[str, str], str] = {}
        marches_vus: Dict[Tuple[str, str], set] = {}

        def _texte(row, col) -> str:
            valeur = row.iloc[col]
            return "" if pd.isna(valeur) else str(valeur).strip()

        def _nombre(row, col) -> float:
            valeur = row.iloc[col]
            if pd.isna(valeur):
                return 0.0
            try:
                return float(valeur)
            except (TypeError, ValueError):
                return 0.0

        for marche in marches_operation:
            df_marche = self.df_marches[self.df_marches.iloc[:, self.COL_MARCHE] == marche]
            if len(df_marche) == 0:
                continue

            fournisseur = df_marche.iloc[0, self.COL_FOURNISSEUR]
            fournisseur = "" if pd.isna(fournisseur) else str(fournisseur)
            tranche = df_marche.iloc[0, self.COL_TRANCHE]

            tranche_libelle = ""
            if pd.notna(tranche):
                try:
                    tranche_num = int(float(tranche))
                    tranche_libelle = "TF" if tranche_num == 0 else f"TO{tranche_num}"
                except (TypeError, ValueError):
                    tranche_libelle = str(tranche)

            cle_groupe = (fournisseur, tranche_libelle)
            groupes.setdefault(cle_groupe, [])

            montant_initial_tranche = self.calculate_montant_initial_tranche(marche, tranche)

            # L'enveloppe du groupe cumule celles de ses marchés distincts.
            vus = marches_vus.setdefault(cle_groupe, set())
            if marche not in vus:
                vus.add(marche)
                montant_enveloppe, provenance = self._enveloppe_tranche(
                    marche, tranche_libelle, float(montant_initial_tranche or 0), marches_totaux
                )
                enveloppes[cle_groupe] = enveloppes.get(cle_groupe, 0.0) + montant_enveloppe
                # La provenance la moins fiable du groupe l'emporte : mieux vaut
                # signaler un doute que laisser croire à un montant contractuel.
                courante = provenances.get(cle_groupe, self.ENVELOPPE_BASE)
                if self.RANG_PROVENANCE[provenance] >= self.RANG_PROVENANCE[courante]:
                    provenances[cle_groupe] = provenance

            type_marche = marches_types.get(marche, 'CLASSIQUE')

            for _, row in df_marche.iterrows():
                num_commande = row.iloc[self.COL_COMMANDE]
                num_commande = "" if pd.isna(num_commande) else str(num_commande).strip()

                if exercice_filter and exercice_filter != "Tous":
                    if self.extract_exercice_from_bdc(num_commande) != exercice_filter:
                        continue

                # Une ligne sans BDC est rattachée à sa tranche : le montant de
                # référence est alors le montant initial de la tranche.
                montant_initial = _nombre(row, self.COL_MONTANT_INITIAL)
                if not num_commande and montant_initial <= 0:
                    montant_initial = float(montant_initial_tranche or 0)

                groupes[cle_groupe].append(Ecriture(
                    marche=marche,
                    fournisseur=fournisseur,
                    tranche_libelle=tranche_libelle,
                    num_commande=num_commande,
                    libelle=_texte(row, self.COL_LIBELLE),
                    num_facture=_texte(row, self.COL_FACTURE),
                    num_mandat=_texte(row, self.COL_MANDAT),
                    date_sf=_texte(row, self.COL_DATE_SF),
                    montant_ttc=_nombre(row, self.COL_MONTANT_TTC),
                    montant_sf=_nombre(row, self.COL_MONTANT_SF),
                    montant_initial=montant_initial,
                    type_marche=type_marche,
                ))

        return groupes, montants_declares, enveloppes, provenances, operation_info

    def export_suivi_financier_operation(
        self,
        code_operation: str,
        filepath: str,
        exercice_filter: Optional[str] = None,
        special_export: bool = False
    ) -> bool:
        """
        Génère un fichier Excel de suivi financier pour une opération spécifique.

        Le tableau émet **une ligne par état de bon de commande**, pas une ligne
        par écriture SEDIT : les factures mandatées telles quelles, plus une
        unique ligne d'engagement portant le reliquat non facturé. Un BDC soldé
        n'a plus de ligne d'engagement, et un report d'exercice ne crée plus de
        second engagement.

        Args:
            code_operation: Code de l'opération (ex: "2024_1", "2024_17")
            filepath: Chemin du fichier Excel à créer
            exercice_filter: Exercice à filtrer (ex: "2024") ou "Tous"/None
            special_export: Trie les BDC par numéro et écrit un log de contrôle

        Returns:
            True si succès, False sinon
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            groupes, montants_declares, enveloppes, provenances, operation_info = \
                self.collecter_ecritures_operation(code_operation, exercice_filter)
            if operation_info is None:
                return False

            # Agrégation : une ligne par état de BDC. Onglets FINANCIER,
            # « A jour », « Anomalies » et « Lignes neutralisées » en découlent.
            resultats: Dict[Tuple[str, str], ResultatSuivi] = {}
            for cle_groupe, ecritures in groupes.items():
                resultats[cle_groupe] = agreger_ecritures(
                    ecritures, montants_declares, trier_par_bdc=special_export
                )

            global_resultat = ResultatSuivi()
            for resultat in resultats.values():
                global_resultat.etendre(resultat)
            enveloppe_totale = sum(enveloppes.values())

            # Provenance de l'enveloppe affichée : la moins fiable l'emporte.
            sources = set(provenances.values())
            if self.ENVELOPPE_ABSENTE in sources or not sources:
                libelle_enveloppe = "Enveloppe initiale (NON RENSEIGNÉE EN BASE) :"
            elif self.ENVELOPPE_SEDIT in sources:
                libelle_enveloppe = (
                    "Enveloppe initiale (reconstituée depuis l'export SEDIT — à saisir en base) :"
                )
            else:
                libelle_enveloppe = "Enveloppe initiale (lue en base) :"

            wb = Workbook()
            wb.remove(wb.active)

            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border_thin = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            fill_gray = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            fill_subtotal = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            fill_engagement = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            font_subtotal = Font(bold=True, size=11)
            border_thick = Border(
                left=Side(style='thick'), right=Side(style='thick'),
                top=Side(style='thick'), bottom=Side(style='thick')
            )
            format_euro = '#,##0.00 €'

            # ===============================
            # FEUILLE 1 : FINANCIER
            # ===============================
            ws_financier = wb.create_sheet("FINANCIER")

            ws_financier.cell(1, 1, f"SUIVI FINANCIER OPÉRATION {code_operation}")
            ws_financier.cell(1, 1).font = Font(bold=True, size=14)
            ws_financier.merge_cells('A1:O1')

            ws_financier.cell(2, 1, f"Libellé: {operation_info.get('libelle', '')}")
            ws_financier.cell(3, 1, f"Fournisseurs: {operation_info.get('fournisseur', '')}")
            # L'enveloppe initiale est écrite en clair : elle ne doit plus se
            # déduire d'une addition de cellules du tableau.
            ws_financier.cell(4, 1, libelle_enveloppe)
            ws_financier.cell(4, 1).font = Font(bold=True)
            ws_financier.cell(4, 6, enveloppe_totale).number_format = format_euro
            ws_financier.cell(4, 6).font = Font(bold=True)

            headers = [
                "TYPE\nINTERVENTION",
                "DESIGNATION",
                "N° MARCHÉ",
                "NOM PRESTATAIRE",
                "N° BDC ou Tranche (TF ou TO)",
                "MONTANT TTC BDC\nou Tranche",
                "N°FACTURES\nFournisseur",
                "N°MANDAT",
                "Date de\nservice fait",
                "MONTANT FACTURE mandataire HT",
                "MONTANT FACTURE YC mandataire révision TTC",
                "MONTANT FACTURE PAR BDC YC révision-AF-RETGAR TTC",
                "MONTANT FACTURE PAR BDC HORS révision TTC",
                "MONTANT RESTANT SUR BDC Hors rev / engagement TTC",
                "STATUT",
            ]
            nb_colonnes = len(headers)
            colonnes_euro = [6, 10, 11, 12, 13, 14]

            for col_idx, header in enumerate(headers, 1):
                cell = ws_financier.cell(5, col_idx, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            log_file = None
            if special_export:
                import os
                os.makedirs("run_logs", exist_ok=True)
                log_path = os.path.join("run_logs", f"export_{code_operation}.log")
                log_file = open(log_path, "w", encoding="utf-8")
                log_file.write("ligne;marche;bdc;facture;statut;enveloppe;montant_impute;restant\n")

            row_idx = 6
            color_index = 0
            current_prestataire = None

            for (fournisseur, tranche_libelle), resultat in resultats.items():
                if current_prestataire != fournisseur:
                    current_prestataire = fournisseur
                    color_index += 1
                current_fill = fill_gray if color_index % 2 == 1 else fill_white

                enveloppe = enveloppes.get((fournisseur, tranche_libelle), 0.0)
                # Solde glissant : un seul mode de calcul, l'enveloppe initiale
                # moins le cumul des montants imputés depuis le début du groupe.
                cumul_impute = 0.0

                for ligne in resultat.lignes:
                    cumul_impute += ligne.montant_impute
                    restant = enveloppe - cumul_impute

                    ws_financier.cell(row_idx, 1, "TRAVAUX")
                    ws_financier.cell(row_idx, 2, ligne.designation)
                    ws_financier.cell(row_idx, 3, ligne.marche)
                    ws_financier.cell(row_idx, 4, ligne.fournisseur)
                    ws_financier.cell(row_idx, 5, ligne.cle_bdc)
                    # Colonne 6 : toujours le montant de référence du BDC,
                    # identique sur toutes ses lignes. Jamais un montant de facture.
                    ws_financier.cell(row_idx, 6, ligne.montant_ref)
                    ws_financier.cell(row_idx, 7, ligne.num_facture)
                    ws_financier.cell(row_idx, 8, ligne.num_mandat)
                    ws_financier.cell(row_idx, 9, ligne.date_sf)
                    ws_financier.cell(row_idx, 10, ligne.montant_ht)
                    ws_financier.cell(row_idx, 11, ligne.montant_impute)
                    ws_financier.cell(row_idx, 12, ligne.montant_impute)
                    ws_financier.cell(row_idx, 13, ligne.montant_impute)
                    ws_financier.cell(row_idx, 14, restant)
                    ws_financier.cell(row_idx, 15, ligne.statut)

                    ligne_fill = fill_engagement if ligne.statut == STATUT_ENGAGEMENT else current_fill
                    for col in range(1, nb_colonnes + 1):
                        ws_financier.cell(row_idx, col).fill = ligne_fill
                    for col in colonnes_euro:
                        ws_financier.cell(row_idx, col).number_format = format_euro

                    if log_file:
                        log_file.write(
                            f"{row_idx};{ligne.marche};{ligne.cle_bdc};{ligne.num_facture};"
                            f"{ligne.statut};{enveloppe};{ligne.montant_impute};{restant}\n"
                        )

                    row_idx += 1

                # Sous-total du groupe.
                total_impute = resultat.total_impute
                ws_financier.cell(row_idx, 2, f"Sous-total Tranche {tranche_libelle} - {fournisseur}")
                # Somme sur BDC distincts, jamais une somme de colonne.
                ws_financier.cell(row_idx, 6, resultat.total_bdc_distincts)
                ws_financier.cell(row_idx, 10, sum(l.montant_ht for l in resultat.lignes))
                ws_financier.cell(row_idx, 11, total_impute)
                ws_financier.cell(row_idx, 12, total_impute)
                ws_financier.cell(row_idx, 13, total_impute)
                # Même formule que le glissant : la dernière valeur de la
                # colonne N et ce sous-total sont nécessairement égaux.
                ws_financier.cell(row_idx, 14, enveloppe - total_impute)
                ws_financier.cell(row_idx, 15, "TOTAL")

                for col in range(1, nb_colonnes + 1):
                    cell = ws_financier.cell(row_idx, col)
                    cell.fill = fill_subtotal
                    cell.font = font_subtotal
                    cell.border = border_thick
                for col in colonnes_euro:
                    ws_financier.cell(row_idx, col).number_format = format_euro

                row_idx += 1

            largeurs = [12, 40, 20, 25, 15, 15, 15, 12, 12, 14, 14, 14, 14, 14, 32]
            for col_idx, largeur in enumerate(largeurs, 1):
                ws_financier.column_dimensions[get_column_letter(col_idx)].width = largeur

            # ===============================
            # FEUILLE 2 : A JOUR (une ligne par BDC)
            # ===============================
            ws_ajour = wb.create_sheet("A jour")

            ws_ajour.cell(1, 1, f"OPÉRATION {code_operation}")
            ws_ajour.cell(1, 1).font = Font(bold=True, size=14)
            ws_ajour.cell(2, 1, libelle_enveloppe)
            ws_ajour.cell(2, 1).font = Font(bold=True)
            ws_ajour.cell(2, 6, enveloppe_totale).number_format = format_euro
            ws_ajour.cell(2, 6).font = Font(bold=True)

            headers_ajour = [
                "TYPE INTERVENTION",
                "DESIGNATION",
                "N° MARCHÉ",
                "NOM PRESTATAIRE",
                "N° BDC ou Tranche (TF ou TO)",
                "MONTANT TTC BDC ou Tranche",
                "MONTANT FACTURÉ TTC",
                "RELIQUAT NON FACTURÉ TTC",
                "ÉTAT",
            ]
            for col_idx, header in enumerate(headers_ajour, 1):
                cell = ws_ajour.cell(4, col_idx, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            row_idx = 5
            for bdc in global_resultat.bdcs:
                ws_ajour.cell(row_idx, 1, "TRAVAUX")
                ws_ajour.cell(row_idx, 2, bdc.designation)
                ws_ajour.cell(row_idx, 3, bdc.marche)
                ws_ajour.cell(row_idx, 4, bdc.fournisseur)
                ws_ajour.cell(row_idx, 5, bdc.cle_bdc)
                ws_ajour.cell(row_idx, 6, bdc.montant_ref)
                ws_ajour.cell(row_idx, 7, bdc.montant_facture)
                ws_ajour.cell(row_idx, 8, bdc.reliquat)
                ws_ajour.cell(row_idx, 9, bdc.etat)
                for col in [6, 7, 8]:
                    ws_ajour.cell(row_idx, col).number_format = format_euro
                row_idx += 1

            ws_ajour.cell(row_idx, 5, "TOTAL")
            ws_ajour.cell(row_idx, 5).font = Font(bold=True)
            # Totaux issus du même DataFrame que l'onglet FINANCIER : l'égalité
            # avec ses sous-totaux est structurelle, pas fortuite.
            for col, valeur in (
                (6, global_resultat.total_bdc_distincts),
                (7, global_resultat.total_facture),
                (8, global_resultat.total_engagement),
            ):
                cell = ws_ajour.cell(row_idx, col, valeur)
                cell.font = Font(bold=True)
                cell.number_format = format_euro

            for col_idx, largeur in enumerate([12, 40, 20, 25, 15, 18, 18, 18, 24], 1):
                ws_ajour.column_dimensions[get_column_letter(col_idx)].width = largeur

            # ===============================
            # FEUILLE 3 : ANOMALIES
            # ===============================
            ws_anomalies = wb.create_sheet("Anomalies")
            ws_anomalies.cell(1, 1, "ANOMALIES À ARBITRER")
            ws_anomalies.cell(1, 1).font = Font(bold=True, size=14)
            for col_idx, header in enumerate(
                ["N° MARCHÉ", "N° BDC", "ANOMALIE", "VALEURS RENCONTRÉES", "VALEUR RETENUE"], 1
            ):
                cell = ws_anomalies.cell(3, col_idx, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            row_idx = 4
            for anomalie in global_resultat.anomalies:
                ws_anomalies.cell(row_idx, 1, anomalie.marche)
                ws_anomalies.cell(row_idx, 2, anomalie.cle_bdc)
                ws_anomalies.cell(row_idx, 3, anomalie.message)
                ws_anomalies.cell(row_idx, 4, anomalie.valeurs)
                ws_anomalies.cell(row_idx, 5, anomalie.valeur_retenue).number_format = format_euro
                row_idx += 1

            for col_idx, largeur in enumerate([20, 15, 55, 35, 18], 1):
                ws_anomalies.column_dimensions[get_column_letter(col_idx)].width = largeur

            # ===============================
            # FEUILLE 4 : LIGNES NEUTRALISÉES
            # ===============================
            ws_neutral = wb.create_sheet("Lignes neutralisées")
            ws_neutral.cell(1, 1, "LIGNES NEUTRALISÉES PAR L'AGRÉGATION")
            ws_neutral.cell(1, 1).font = Font(bold=True, size=14)
            for col_idx, header in enumerate(
                ["N° MARCHÉ", "N° BDC", "DÉSIGNATION D'ORIGINE", "MONTANT D'ORIGINE",
                 "MONTANT RETENU", "MONTANT NEUTRALISÉ", "MOTIF"], 1
            ):
                cell = ws_neutral.cell(3, col_idx, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin

            row_idx = 4
            for neutralisation in global_resultat.neutralisations:
                ws_neutral.cell(row_idx, 1, neutralisation.marche)
                ws_neutral.cell(row_idx, 2, neutralisation.cle_bdc)
                ws_neutral.cell(row_idx, 3, neutralisation.designation)
                ws_neutral.cell(row_idx, 4, neutralisation.montant_origine)
                ws_neutral.cell(row_idx, 5, neutralisation.montant_retenu)
                ws_neutral.cell(row_idx, 6, neutralisation.montant_neutralise)
                ws_neutral.cell(row_idx, 7, neutralisation.motif)
                for col in [4, 5, 6]:
                    ws_neutral.cell(row_idx, col).number_format = format_euro
                row_idx += 1

            ws_neutral.cell(row_idx, 3, "TOTAL")
            ws_neutral.cell(row_idx, 3).font = Font(bold=True)
            for col, valeur in (
                (4, sum(n.montant_origine for n in global_resultat.neutralisations)),
                (5, sum(n.montant_retenu for n in global_resultat.neutralisations)),
                (6, sum(n.montant_neutralise for n in global_resultat.neutralisations)),
            ):
                cell = ws_neutral.cell(row_idx, col, valeur)
                cell.font = Font(bold=True)
                cell.number_format = format_euro

            for col_idx, largeur in enumerate([20, 15, 45, 18, 18, 18, 60], 1):
                ws_neutral.column_dimensions[get_column_letter(col_idx)].width = largeur

            if log_file:
                log_file.close()

            wb.save(filepath)
            print(
                f"[OK] Export suivi financier operation {code_operation} : {filepath} "
                f"({len(global_resultat.lignes)} lignes, "
                f"{global_resultat.total_impute:.2f} € imputés, "
                f"{len(global_resultat.neutralisations)} lignes neutralisées)"
            )
            return True

        except Exception as e:
            print(f"[ERREUR] Erreur lors de l'export suivi financier : {e}")
            import traceback
            traceback.print_exc()
            return False
